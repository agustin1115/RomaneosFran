"""
china_loader.py — Lee el desposte de China/Top Meat en formato Excel
(COD / MERCADERIA / UNIDAD / ETIQUETA / NETO / DESTINO / ...).

Devuelve un dict con la MISMA forma que parse_romaneo_pdf, para que la grilla
de selección, la segmentación y el análisis lo traten igual que un romaneo.
Los cortes de China (cuartos en manta, garrón, brazuelo...) se dejan con su
nombre propio; la valorización va por el modelo de exportación (USD FOB).
"""
import openpyxl
from datetime import datetime, date


def _num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('.', '').replace(',', '.').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _fecha(v):
    if isinstance(v, (datetime, date)):
        return v.strftime('%d/%m/%Y')
    return str(v) if v else ''


def parse_china_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    # localizar fila de encabezado (tiene MERCADERIA y NETO)
    head_i = None
    for i, r in enumerate(rows):
        vals = [str(x).upper() if x is not None else '' for x in r]
        if any('MERCADERIA' in v for v in vals) and any('NETO' in v for v in vals):
            head_i = i
            cols = {v.strip(): j for j, v in enumerate(vals) if v.strip()}
            break
    if head_i is None:
        raise ValueError('No parece un desposte de China (falta encabezado MERCADERIA/NETO).')

    def col(name, default=None):
        for k, j in cols.items():
            if name in k:
                return j
        return default

    ci_cod, ci_merc = col('COD'), col('MERCADERIA')
    ci_uni, ci_neto = col('UNIDAD'), col('NETO')
    ci_dest, ci_temp = col('DESTINO'), col('TEMPERATURA')

    cortes = []
    for r in rows[head_i + 1:]:
        if r is None:
            continue
        cod = r[ci_cod] if ci_cod is not None and ci_cod < len(r) else None
        merc = r[ci_merc] if ci_merc is not None and ci_merc < len(r) else None
        neto = _num(r[ci_neto]) if ci_neto is not None and ci_neto < len(r) else 0
        if not merc or not str(merc).strip():
            break  # fila de total / vacía → fin de cortes
        if neto <= 0:
            continue
        cortes.append({
            'corte': str(merc).strip(),
            'codigo': str(cod).strip() if cod else '',
            'grupo': str(merc).strip(),  # China: el grupo ES el corte (no mapea a consumo)
            'tipo': 'CHINA',
            'piezas': int(_num(r[ci_uni])) if ci_uni is not None and ci_uni < len(r) else 0,
            'unidades': int(_num(r[ci_uni])) if ci_uni is not None and ci_uni < len(r) else 0,
            'kg': neto,
            'destino': str(r[ci_temp]).strip() if ci_temp is not None and ci_temp < len(r) and r[ci_temp] else 'CONGE',
            'contramarca': '',
            'cliente': str(r[ci_dest]).strip() if ci_dest is not None and ci_dest < len(r) and r[ci_dest] else 'CHINA',
            'nro_venta': '',
            'es_bubalino': False,
        })

    # metadata (recepción / faena / cajas / sebo / hueso) — buscar por etiqueta
    def buscar(label):
        for r in rows:
            for j, cell in enumerate(r):
                if cell and label in str(cell).upper():
                    # el valor suele estar en la celda siguiente no vacía
                    for k in range(j + 1, len(r)):
                        if r[k] not in (None, ''):
                            return _num(r[k])
        return 0.0

    kg_recepcion = buscar('KGS DE RECPECION') or buscar('KGS DE RECEPCION') or buscar('RECEPCION')
    kg_faena = buscar('KGS DE FAENA')
    kg_cajas = buscar('KGS DE CAJAS')
    sebo = buscar('SEBO')
    hueso = buscar('HUESO')
    medias = int(buscar('MEDIA RES') or buscar('MEDIA')) or 0

    # tropa / fecha / categoría desde la 2da hoja si existe
    tropa, fecha, categoria = '', '', 'Vaca'
    if len(wb.worksheets) > 1:
        for r in wb.worksheets[1].iter_rows(values_only=True):
            vals = [str(x).upper() if x is not None else '' for x in r]
            if any('VACA' in v for v in vals):
                categoria = 'Vaca'
            for x in r:
                if isinstance(x, (datetime, date)) and not fecha:
                    fecha = _fecha(x)
                if isinstance(x, int) and 100000 < x < 9999999 and not tropa:
                    tropa = str(x)

    kg_entrada = kg_recepcion or kg_faena or (kg_cajas + sebo + hueso)
    kg_carne = sum(c['kg'] for c in cortes)
    merma = max(0.0, kg_entrada - kg_carne - sebo - hueso)

    return {
        'numero': tropa,
        'fecha': fecha,
        'medias_reses': medias,
        'kg_entrada': kg_entrada,
        'kg_faena': kg_faena,
        'kg_cajas': kg_cajas,
        'categoria': categoria,
        'tipificacion': 'VA',
        'cortes': cortes,
        'grasa_kg': sebo,
        'hueso_kg': hueso,
        'merma_kg': merma,
        'origen': 'china_xlsx',
        'texto_fuente': 'TOP MEAT CHINA DELTACAR ' + ' '.join(c['corte'] for c in cortes),
        'desglose_categoria': {categoria: {'kg': kg_entrada, 'pct': 100.0, 'tip': 'VA'}},
    }
