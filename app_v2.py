"""
Romaneo TF Carnes S.A. — Aplicación de Análisis de Despostada v2.1
Streamlit app para cargar PDFs, configurar parámetros por calidad,
y generar análisis individual o acumulado.
"""
import os
import sys
import json
import tempfile
from pathlib import Path
from datetime import datetime

import streamlit as st
import pandas as pd

# Agregar directorio actual al path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import (
    COSTOS_PERFILES, REND_OBJETIVO, GRUPOS_POR_CALIDAD,
    AMARILLA_PRECIO_DEFAULT, CORTES_CAROS, AMARILLA_CONTRAMARCAS,
    calificacion_negocio, CORTE_TO_GRUPO, PRECIOS_BASE,
    calcular_neto_peya, PRECIO_COMPRA_SUGERIDO,
)
from pdf_parser import (parse_romaneo_pdf, parse_multiple_pdfs, acumular_romaneos,
                        control_cortes, detectar_cortes_faltantes, detectar_tipo_pdf,
                        parse_remanejo_pdf, es_correccion)
from report_comprador import calcular_score, generar_pdf_comprador
from excel_builder import build_analisis
import segmentacion as seg

# Registro persistente de segmentos (canal por romaneo, por contenido)
SEGMENTOS_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'ROMANEOS', 'segmentos.json')
# Historial de costos por mes (para pricing futuro)
HIST_COSTOS_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'ROMANEOS', 'historial_costos.json')

# ══════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN DE PÁGINA
# ══════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Romaneo — TF Carnes S.A.",
    page_icon="🥩",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
    /* ══════════════════════════════════════════════════════════════
       TF CARNES — Design System
       Paleta: verde #1B4D3E→#2D7D5F · dorado #C9A84C · fondo #F5F7F5
       ══════════════════════════════════════════════════════════════ */
    :root {
        --tf-green-dark: #1B4D3E;
        --tf-green: #2D7D5F;
        --tf-gold: #C9A84C;
        --tf-gold-dark: #B89840;
        --tf-bg: #F5F7F5;
        --tf-card: #FFFFFF;
        --tf-border: #DDE5DD;
        --tf-border-soft: #E8F5E9;
        --tf-text: #1A1A1A;
        --tf-text-soft: #666;
        --tf-pos: #27AE60;
        --tf-neg: #C0392B;
    }

    html, body, [class*="css"] {
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        color: var(--tf-text);
    }

    .stApp { background: var(--tf-bg); }

    /* ══════ TOP TOOLBAR (Deploy / menu) — compacto ══════ */
    header[data-testid="stHeader"] {
        height: 2.2rem !important;
        background: transparent !important;
        backdrop-filter: none !important;
    }
    [data-testid="stToolbar"] { right: 0.5rem; top: 0.2rem; }
    [data-testid="stToolbar"] button { font-size: 12px !important; padding: 2px 8px !important; }
    [data-testid="stDecoration"] { display: none; }

    /* Reducir padding superior del bloque principal */
    .main .block-container,
    [data-testid="stAppViewContainer"] .main .block-container {
        padding-top: 1rem !important;
        padding-bottom: 2rem;
    }

    /* ══════ HERO HEADER (compacto) ══════ */
    .tf-hero {
        background: linear-gradient(135deg, #1B4D3E, #2D7D5F);
        color: #fff;
        padding: 12px 20px;
        border-radius: 12px;
        margin-bottom: 14px;
        box-shadow: 0 2px 10px rgba(27,77,62,.22);
        position: relative;
        overflow: hidden;
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 16px;
    }
    .tf-hero::after {
        content: ''; position: absolute; right: -30px; top: -30px;
        width: 120px; height: 120px; border-radius: 50%;
        background: radial-gradient(circle, rgba(201,168,76,.15), transparent 70%);
    }
    .tf-hero h1 {
        font-size: 18px; font-weight: 800; letter-spacing: 2.5px;
        margin: 0; text-transform: uppercase;
    }
    .tf-hero p {
        font-size: 10.5px; color: #C9A84C; margin: 2px 0 0 0;
        letter-spacing: 1.4px; text-transform: uppercase; font-weight: 600;
    }
    .tf-hero .gl {
        width: 38px; height: 2px; background: #C9A84C;
        border-radius: 2px; margin-top: 6px;
    }
    .tf-hero-meta {
        font-size: 10.5px; color: rgba(255,255,255,.7);
        letter-spacing: 1.2px; text-transform: uppercase;
        text-align: right; font-weight: 600;
        position: relative; z-index: 1;
    }

    /* ══════ Botón nativo de colapsar sidebar (siempre visible) ══════ */
    [data-testid="stSidebarCollapseButton"],
    button[kind="header"][data-testid="baseButton-headerNoPadding"] {
        display: inline-flex !important;
        background: var(--tf-green-dark) !important;
        color: #fff !important;
        border-radius: 8px !important;
        padding: 4px 6px !important;
        box-shadow: 0 2px 6px rgba(27,77,62,.25);
    }

    /* ══════ TABS ══════ */
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px; background: transparent; border-bottom: 2px solid var(--tf-border-soft);
        padding-bottom: 0;
    }
    .stTabs [data-baseweb="tab"] {
        background: transparent;
        border-radius: 10px 10px 0 0;
        padding: 10px 18px;
        font-size: 13px; font-weight: 600;
        color: var(--tf-text-soft);
        border: none;
        letter-spacing: .3px;
        transition: all .15s ease;
    }
    .stTabs [data-baseweb="tab"]:hover { color: var(--tf-green-dark); background: rgba(45,125,95,.06); }
    .stTabs [aria-selected="true"] {
        background: var(--tf-card) !important;
        color: var(--tf-green-dark) !important;
        border-bottom: 3px solid var(--tf-gold) !important;
        font-weight: 700;
    }

    /* ══════ HEADINGS ══════ */
    h1, h2, h3, h4, h5 { color: var(--tf-green-dark); font-weight: 700; }
    h3 { font-size: 1.2rem; letter-spacing: .3px; }
    h4 {
        font-size: 13px !important;
        text-transform: uppercase;
        letter-spacing: 1.5px;
        color: var(--tf-green-dark) !important;
        border-bottom: 2px solid var(--tf-border-soft);
        padding-bottom: 6px;
        margin-top: 20px !important;
        margin-bottom: 14px !important;
        font-weight: 700 !important;
    }

    /* ══════ METRICS ══════ */
    [data-testid="stMetric"] {
        background: var(--tf-card);
        border: 1px solid var(--tf-border);
        border-radius: 14px;
        padding: 14px 16px;
        box-shadow: 0 1px 6px rgba(0,0,0,.04);
        transition: box-shadow .15s ease;
    }
    [data-testid="stMetric"]:hover {
        box-shadow: 0 2px 12px rgba(27,77,62,.08);
        border-color: rgba(45,125,95,.3);
    }
    [data-testid="stMetricLabel"] {
        font-size: 11px !important;
        text-transform: uppercase;
        letter-spacing: 1px;
        color: var(--tf-text-soft) !important;
        font-weight: 600;
    }
    [data-testid="stMetricValue"] {
        color: var(--tf-green-dark) !important;
        font-weight: 800 !important;
        font-size: 1.6rem !important;
    }

    /* ══════ INPUTS ══════ */
    .stNumberInput input, .stTextInput input, .stTextArea textarea, .stDateInput input {
        border: 1.5px solid var(--tf-border) !important;
        border-radius: 10px !important;
        background: var(--tf-card) !important;
        color: var(--tf-text) !important;
        font-size: 14px !important;
    }
    .stNumberInput input:focus, .stTextInput input:focus, .stTextArea textarea:focus {
        border-color: var(--tf-green) !important;
        box-shadow: 0 0 0 3px rgba(45,125,95,.12) !important;
    }
    .stSelectbox > div > div {
        border: 1.5px solid var(--tf-border) !important;
        border-radius: 10px !important;
        background: var(--tf-card) !important;
    }
    label p { font-size: 12px !important; color: var(--tf-text-soft) !important;
              font-weight: 600 !important; letter-spacing: .3px; }

    /* ══════ BUTTONS ══════ */
    .stButton > button {
        border-radius: 10px;
        font-weight: 700;
        letter-spacing: .5px;
        padding: 10px 18px;
        border: 1.5px solid var(--tf-green-dark);
        background: var(--tf-card);
        color: var(--tf-green-dark);
        transition: all .15s ease;
    }
    .stButton > button:hover {
        background: var(--tf-green-dark);
        color: #fff;
        border-color: var(--tf-green-dark);
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, var(--tf-gold), var(--tf-gold-dark));
        color: #fff;
        border: none;
        text-transform: uppercase;
        letter-spacing: 1.2px;
        box-shadow: 0 4px 12px rgba(201,168,76,.35);
    }
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, var(--tf-gold-dark), var(--tf-gold));
        transform: translateY(-1px);
        box-shadow: 0 6px 16px rgba(201,168,76,.45);
    }
    .stDownloadButton > button {
        background: var(--tf-green-dark);
        color: #fff;
        border: none;
        border-radius: 10px;
        font-weight: 700;
        letter-spacing: .5px;
    }
    .stDownloadButton > button:hover {
        background: var(--tf-green);
    }

    /* ══════ RADIO / CHECKBOX ══════ */
    .stRadio [data-baseweb="radio"] {
        background: var(--tf-card);
        border: 1.5px solid var(--tf-border);
        border-radius: 10px;
        padding: 8px 14px;
        margin-right: 6px;
        transition: all .15s ease;
    }
    .stRadio [data-baseweb="radio"]:hover { border-color: var(--tf-green); }

    /* ══════ DATAFRAMES ══════ */
    [data-testid="stDataFrame"] {
        border: 1px solid var(--tf-border);
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 1px 6px rgba(0,0,0,.04);
    }

    /* ══════ EXPANDERS ══════ */
    .streamlit-expanderHeader, [data-testid="stExpander"] summary {
        background: var(--tf-card);
        border: 1px solid var(--tf-border);
        border-radius: 10px;
        font-weight: 600;
        color: var(--tf-green-dark);
    }
    [data-testid="stExpander"] {
        border: 1px solid var(--tf-border);
        border-radius: 12px;
        background: var(--tf-card);
        margin-bottom: 10px;
        box-shadow: 0 1px 4px rgba(0,0,0,.03);
    }

    /* ══════ ALERTS ══════ */
    [data-testid="stAlert"] {
        border-radius: 12px;
        border: none;
        box-shadow: 0 1px 6px rgba(0,0,0,.04);
    }

    .alert-amarilla {
        background-color: #FFF8E1;
        border-left: 4px solid var(--tf-gold);
        border-radius: 10px;
        padding: 14px 16px;
        margin: 10px 0;
    }

    /* ══════ DIVIDERS ══════ */
    hr {
        border: none;
        border-top: 1px solid var(--tf-border-soft);
        margin: 18px 0;
    }

    /* ══════ SIDEBAR ══════ */
    div[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #FAFCFA 0%, #F0F4F0 100%);
        border-right: 1px solid var(--tf-border);
    }
    div[data-testid="stSidebar"] h2,
    div[data-testid="stSidebar"] h3 {
        color: var(--tf-green-dark) !important;
        font-weight: 700;
    }
    div[data-testid="stSidebar"] [data-testid="stMetric"] {
        background: var(--tf-card);
    }

    /* ══════ SLIDERS ══════ */
    .stSlider [data-baseweb="slider"] [role="slider"] {
        background: var(--tf-green-dark) !important;
        border-color: var(--tf-green-dark) !important;
    }

    /* ══════ FILE UPLOADER ══════ */
    [data-testid="stFileUploader"] section {
        border: 2px dashed var(--tf-border) !important;
        border-radius: 12px;
        background: var(--tf-card);
    }
    [data-testid="stFileUploader"] section:hover { border-color: var(--tf-green) !important; }

    /* Legacy header (fallback) */
    .main-header { display: none; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
# ESTADO DE SESIÓN
# ══════════════════════════════════════════════════════════════════════
if 'parsed_files' not in st.session_state:
    st.session_state.parsed_files = []
if 'analysis_results' not in st.session_state:
    st.session_state.analysis_results = []
if 'html_produccion_results' not in st.session_state:
    st.session_state.html_produccion_results = []

# ══════════════════════════════════════════════════════════════════════
# CONSTANTES GOOGLE SHEETS
# ══════════════════════════════════════════════════════════════════════
SPREADSHEET_COMPRAS = '1OVoP_2QE2gWnX6CPL7ys2vjcisMbuSc9kUpfTcWy66M'
SPREADSHEET_PRECIOS = '1u5RWP5thzfo9ZxSnBqkEk-Usn29qHc3qla43PqnvyP0'
CREDENTIALS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'credentials.json')

# ── Bootstrap credenciales para Streamlit Cloud ───────────────────────────
# En la nube no existe credentials.json (está gitignored). Si hay secrets
# configurados en Streamlit Cloud, los escribimos a un archivo en runtime
# para que todo el código que usa from_service_account_file siga funcionando.
if not os.path.exists(CREDENTIALS_PATH):
    try:
        if 'gcp_service_account' in st.secrets:
            with open(CREDENTIALS_PATH, 'w') as _f:
                json.dump(dict(st.secrets['gcp_service_account']), _f)
    except Exception:
        pass
# ──────────────────────────────────────────────────────────────────────────

DESCUENTO_PEYA = 1 - (1 - 0.05) * (1 - 0.02) * (1 - 0.045)
DESCUENTO_MAXI_MERMA = 0.035

# ══════════════════════════════════════════════════════════════════════
# SIDEBAR — PARÁMETROS EDITABLES
# ══════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## ⚙️ Parámetros")

    calidad = st.selectbox(
        "**Perfil de calidad**",
        ['Standard', 'Búfalo', 'Premium Black', 'Exportación'],
        help="Cada perfil tiene costos, rendimientos y % de cortes distintos"
    )

    perfil_desc = {
        'Standard': '🐄 Mercado interno — Vaca, Novillo, Novillito estándar',
        'Búfalo': '🦬 Carne de búfalo — Más magra, rendimientos distintos',
        'Premium Black': '🏷️ Selección premium — Cortes de alta gama',
        'Exportación': '🌍 Mercado internacional — Hilton, Cuota, export',
    }
    st.info(perfil_desc[calidad])

    st.markdown("---")
    st.markdown("### 💰 Costos variables ($/kg carne)")

    costos_base = COSTOS_PERFILES[calidad].copy()
    costos_edit = {}
    costos_edit['mo'] = st.number_input("Mano de obra", value=costos_base['mo'], step=10, key='mo')
    costos_edit['insumos'] = st.number_input("Insumos", value=costos_base['insumos'], step=10, key='ins')
    costos_edit['flete'] = st.number_input("Flete", value=costos_base['flete'], step=10, key='fl')
    costos_edit['senasa'] = st.number_input("SENASA + cuarteo", value=costos_base['senasa'], step=10, key='sen')
    _iibb_pct = st.number_input("IIBB (% sobre venta)", value=costos_base['iibb']*100,
                                 step=0.1, format="%.2f", key='iibb_v2')
    costos_edit['iibb'] = _iibb_pct / 100

    costo_total_var = costos_edit['mo'] + costos_edit['insumos'] + costos_edit['flete'] + costos_edit['senasa']
    st.metric("Total costo variable/kg", f"${costo_total_var:,.0f}")

    st.markdown("---")
    st.markdown("### 🟡 Precio amarilla")
    precio_amarilla = st.number_input("$/kg amarilla (ctm 47/73/74)",
                                       value=AMARILLA_PRECIO_DEFAULT, step=100, key='am')

    st.markdown("### 🦬 Precio bubalino")
    from config import BUBALINO_PRECIO_DEFAULT
    precio_bubalino = st.number_input("$/kg bubalino (estandarizado)",
                                       value=BUBALINO_PRECIO_DEFAULT, step=100, key='bub')

    st.markdown("---")
    st.markdown("### 🎯 Rendimiento objetivo")
    rend_tabla = REND_OBJETIVO[calidad]
    for cat, rend in rend_tabla.items():
        st.markdown(f"- **{cat}**: {rend*100:.0f}%")

    st.markdown("---")
    precio_sugerido = PRECIO_COMPRA_SUGERIDO.get(calidad, 6800)
    precio_compra = st.number_input(
        "**💲 Precio compra $/kg** (media res s/IVA)",
        value=precio_sugerido, step=100, key='pc',
        help="Se puede sobreescribir con la planilla de compras"
    )

    st.markdown("---")
    st.markdown("### 📦 Planilla de compras")
    st.markdown("Se lee automáticamente de Google Sheets.")
    if os.path.exists(CREDENTIALS_PATH):
        st.success("🔗 Conectado a Google Sheets")
    else:
        st.error("❌ Falta credentials.json")

    st.markdown("---")
    st.markdown("### 📋 Precios de facturación")
    st.markdown("Se leen en vivo desde Google Sheets.")

# ══════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════
def _parse_ar_number(s):
    """Parsea número en formato AR ('1.234.567,89') o US. Devuelve float o None."""
    if s is None:
        return None
    s = str(s).replace('$', '').replace(' ', '').strip()
    if not s or s in {'0', '-', '—'}:
        return None
    try:
        if '.' in s and ',' in s:
            return float(s.replace('.', '').replace(',', '.'))
        if ',' in s:
            return float(s.replace('.', '').replace(',', '.'))
        if '.' in s:
            parts = s.split('.')
            # US decimal: "1.23" (un solo punto y 1-2 cifras después)
            if len(parts) == 2 and 1 <= len(parts[1]) <= 2:
                return float(s)
            return float(s.replace('.', ''))
        return float(s)
    except (ValueError, TypeError):
        return None


@st.cache_data(ttl=300)  # Cache 5 minutos
def cargar_compras_google_sheets():
    """
    Lee la solapa 'Compras' de Google Sheets vía Service Account.
    Retorna dict {nro_tropa: [lista de entradas]} — una tropa puede tener varias
    filas (partidas); cada entrada lleva precio, monto y kg_derivado para poder
    calcular el promedio ponderado aguas abajo.
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        if not os.path.exists(CREDENTIALS_PATH):
            return None, 'No se encontró credentials.json'

        SCOPES = [
            'https://www.googleapis.com/auth/spreadsheets.readonly',
            'https://www.googleapis.com/auth/drive.metadata.readonly',
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(SPREADSHEET_COMPRAS)
        ws = sh.worksheet('Compras')
        all_data = ws.get_all_values()

        if len(all_data) < 2:
            return None, 'Planilla vacía'

        # Columnas conocidas:
        # 3=Fecha faena, 7=Tipo, 41/42/43=TROPA, 53=$/kg
        # Para monto de cada partida: escaneamos cols 8..52 y tomamos el valor
        # numérico más grande cuyo kg derivado (monto/precio) sea razonable.
        from datetime import date
        hoy = date.today()

        # Primera pasada: parseamos todas las filas sin asignar año todavía.
        filas_parsed = []  # lista de (dd, mm, yyyy_explicito_o_None, entrada, tropas)
        for row in all_data[1:]:
            if len(row) < 54:
                continue

            fecha = row[3].strip() if row[3] else ''
            tipo = row[7].strip() if row[7] else ''

            precio = _parse_ar_number(row[53])
            if not precio or precio <= 0:
                continue

            monto = None
            kg_der = None
            best_score = 0
            for c in range(8, 53):
                if c in {41, 42, 43}:
                    continue
                val = _parse_ar_number(row[c])
                if not val or val <= 0:
                    continue
                kg_try = val / precio
                if 50 <= kg_try <= 100000 and val > best_score:
                    best_score = val
                    monto = val
                    kg_der = kg_try

            dd = mm = yyyy = None
            if fecha:
                try:
                    parts = fecha.replace('-', '/').split('/')
                    if len(parts) >= 2:
                        dd = int(parts[0])
                        mm = int(parts[1])
                        if len(parts) >= 3 and parts[2].strip():
                            raw_y = int(parts[2])
                            yyyy = raw_y + 2000 if raw_y < 100 else raw_y
                except (ValueError, IndexError):
                    dd = mm = yyyy = None

            # Tropas que empiezan con "7" → medias compradas a matarife
            # (no faena propia). Se marca para segmentar acumulado.
            tropas_en_fila = set()
            for col_t in [41, 42, 43]:
                tropa = row[col_t].strip() if row[col_t] else ''
                if tropa and tropa != '0' and tropa != 'XX':
                    tropas_en_fila.add(tropa)
            es_compra_media = any(t.startswith('7') and len(t) >= 4 for t in tropas_en_fila)

            entrada = {
                'precio': precio,
                'monto': monto,
                'kg_sheet': kg_der,
                'tipo': tipo,
                'fecha': fecha,
                'fecha_faena_dt': None,  # se completa abajo
                'es_compra_media': es_compra_media,
            }

            filas_parsed.append((dd, mm, yyyy, entrada, tropas_en_fila))

        # Segunda pasada: inferir el año asumiendo orden cronológico ascendente
        # (la planilla va agregando filas con faenas más recientes abajo).
        # Caminamos de abajo hacia arriba partiendo del año actual; sólo
        # decrementamos cuando el mes SALTA ≥6 (cruce real de fin de año
        # Dic→Ene). Saltos chicos son desórdenes locales normales en la planilla
        # y NO deben dispararnos años hacia atrás.
        # Además: si una fecha cae en el futuro, retrocedemos un año por
        # seguridad (no existen faenas futuras).
        year_cursor = hoy.year
        prev_mm = None
        for i in range(len(filas_parsed) - 1, -1, -1):
            dd, mm, yyyy, entrada, _ = filas_parsed[i]
            if mm is None:
                continue
            if yyyy is not None:
                year_cursor = yyyy
                prev_mm = mm
                continue
            if prev_mm is not None and (mm - prev_mm) >= 6:
                year_cursor -= 1
            try:
                cand = date(year_cursor, mm, dd)
            except ValueError:
                cand = None
            if cand and cand > hoy:
                try:
                    cand = date(year_cursor - 1, mm, dd)
                except ValueError:
                    cand = None
            entrada['fecha_faena_dt'] = cand
            prev_mm = mm

        # Aplicar los años explícitos que sí tenían yyyy
        for dd, mm, yyyy, entrada, _ in filas_parsed:
            if yyyy is not None and dd is not None and mm is not None and entrada['fecha_faena_dt'] is None:
                try:
                    entrada['fecha_faena_dt'] = date(yyyy, mm, dd)
                except ValueError:
                    pass

        # Construir dict final por tropa
        compras = {}
        for _, _, _, entrada, tropas_en_fila in filas_parsed:
            for tropa in tropas_en_fila:
                compras.setdefault(tropa, []).append(entrada)

        return compras, f'{len(compras)} tropas cargadas'

    except Exception as e:
        return None, str(e)


def extraer_tropas_romaneo(parsed_data):
    """
    Extrae los números de tropa de un romaneo parseado.
    Las tropas aparecen en las líneas de entrada del PDF como "28037- 26".
    El número antes del guión (28037) es la tropa en la planilla de compras.
    """
    import pdfplumber
    tropas = set()

    # Las tropas están en las líneas de entrada, no en los cortes.
    # El parser ya extrajo el texto — buscar en los datos raw si disponibles.
    # Pero también podemos extraerlas de los nro_venta de los cortes:
    # Formato 1: "87- 215" → 87 es contramarca, 215 es nro romaneo (NO tropa)
    # Formato 2: "87-100215" → 87 es contramarca

    # Las tropas reales están en la sección de Entrada del PDF.
    # El parser no las guarda, pero las vimos en el texto raw:
    # "1 1/2 RES SHOFAR REC VA 28036- 26 73 8,214.00"
    # Acá 28036 es la tropa y 26 es otro identificador.

    # Vamos a buscar en los datos de entrada si el parser los guardó
    # Si no, retornamos vacío y se usa el archivo PDF directamente

    return tropas


def _extraer_tropas_con_kg(pdf_path):
    """
    Extrae tropas + kg de cada tropa desde la sección Entrada del PDF.
    Formato típico: "1 1/2 RES SHOFAR REC VA 28037- 26 59 7,330.00"
    donde 28037 es tropa, 59 es medias, 7330 es kg.
    Retorna dict {tropa: kg_totales}.
    """
    import pdfplumber, re
    tropas_kg = {}
    if not pdf_path or not os.path.exists(pdf_path):
        return tropas_kg
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                in_entrada = False
                for line in text.split('\n'):
                    if 'Entrada Despostada' in line:
                        in_entrada = True
                        continue
                    if in_entrada and 'Total >>>' in line:
                        return tropas_kg
                    if in_entrada:
                        # Buscar tropa y kg en la misma línea
                        m_tropa = re.search(r'(\d{4,6})-\s*\d+', line)
                        if not m_tropa:
                            continue
                        tropa = m_tropa.group(1)
                        # Últimos número de la línea = kg (formato 7,330.00 o 7.330,00 o 7330)
                        m_kg = re.findall(r'([\d.,]+)', line)
                        # El último o penúltimo grande es kg
                        kg_val = 0
                        for n in reversed(m_kg):
                            try:
                                # Limpiar comas de miles (formato "7,330.00")
                                clean = n.replace(',', '')
                                v = float(clean)
                                if v > 50:  # kg probable
                                    kg_val = v
                                    break
                            except ValueError:
                                continue
                        if kg_val > 0:
                            tropas_kg[tropa] = tropas_kg.get(tropa, 0) + kg_val
    except Exception:
        pass
    return tropas_kg


def cruzar_precio_compra_gsheets(parsed_data, pdf_path=None):
    """
    Cruza el romaneo con la planilla de compras de Google Sheets.
    Si una tropa tiene varias partidas en la planilla (misma tropa, varias
    filas con distinto $/kg), se resuelve con promedio ponderado por monto:
        precio_tropa = Σ monto / Σ kg
    Luego se pondera entre tropas por kg del PDF.
    """
    compras, status = cargar_compras_google_sheets()
    if not compras:
        return None, None, status

    tropas_kg = _extraer_tropas_con_kg(pdf_path)
    tropas_encontradas = set(tropas_kg.keys())

    if not tropas_encontradas:
        return None, None, 'No se encontraron tropas en el PDF'

    # Fecha del romaneo para filtrar matches por proximidad temporal
    # (los nros de tropa se reciclan entre años → distinguir por fecha)
    fecha_rom_dt = None
    fecha_str_rom = (parsed_data or {}).get('fecha', '')
    if fecha_str_rom:
        try:
            from datetime import datetime as _dt
            fecha_rom_dt = _dt.strptime(fecha_str_rom.strip(), '%d/%m/%Y').date()
        except Exception:
            fecha_rom_dt = None

    matches = []
    tropas_sin_match = []
    for tropa in tropas_encontradas:
        kg_t = tropas_kg.get(tropa, 0)
        entradas = compras.get(tropa)
        if not entradas:
            tropas_sin_match.append({'tropa': tropa, 'kg': kg_t})
            continue

        # Filtro por proximidad: si tenemos fecha romaneo, descartar partidas
        # con faena > 120 días antes del romaneo o posteriores al romaneo.
        if fecha_rom_dt:
            cercanas = [
                e for e in entradas
                if e.get('fecha_faena_dt')
                and 0 <= (fecha_rom_dt - e['fecha_faena_dt']).days <= 120
            ]
            if cercanas:
                entradas = cercanas
            else:
                # Ningún match dentro del rango razonable → tropa no liquidada
                tropas_sin_match.append({
                    'tropa': tropa, 'kg': kg_t,
                    'razon': 'sin partida cercana a fecha romaneo',
                })
                continue

        # Promedio ponderado intra-tropa (suma monto / suma kg)
        partidas = len(entradas)
        monto_total = sum(e['monto'] for e in entradas if e.get('monto'))
        kg_sheet_total = sum(e['kg_sheet'] for e in entradas if e.get('kg_sheet'))
        if monto_total > 0 and kg_sheet_total > 0:
            precio_tropa = monto_total / kg_sheet_total
        else:
            precio_tropa = sum(e['precio'] for e in entradas) / len(entradas)

        primera = entradas[0]
        # Tropa empieza con 7 (4+ dígitos) → media comprada
        es_compra = (tropa.startswith('7') and len(tropa) >= 4) \
                    or any(e.get('es_compra_media') for e in entradas)
        matches.append({
            'tropa': tropa,
            'precio': round(precio_tropa),
            'tipo': primera['tipo'],
            'fecha': primera['fecha'],
            'fecha_faena_dt': primera.get('fecha_faena_dt'),
            'kg': kg_t,
            'partidas': partidas,
            'precios_partidas': [round(e['precio']) for e in entradas],
            'es_compra_media': es_compra,
        })

    if not matches:
        return (None, tropas_encontradas,
                f'Tropas {tropas_encontradas} no encontradas en planilla')

    total_kg_match = sum(m['kg'] for m in matches) or 1
    precio_ponderado = sum(m['precio'] * m['kg'] for m in matches) / total_kg_match

    return {
        'precio_promedio': round(precio_ponderado),
        'matches': matches,
        'tropas_romaneo': tropas_encontradas,
        'tropas_sin_match': tropas_sin_match,
        'kg_matcheado': total_kg_match,
    }, tropas_encontradas, 'ok'


def generar_nombre_archivo(parsed_data, calidad):
    """
    Genera nombre descriptivo: DDMM CAT MEDIAS CALIDAD
    Ejemplo: 1003 VQ 120 Standard
    """
    fecha = parsed_data.get('fecha', '')
    # Extraer DD y MM
    if '/' in fecha:
        parts = fecha.split('/')
        dd = parts[0]
        mm = parts[1]
    else:
        dd = '00'
        mm = '00'

    cat = parsed_data.get('categoria', 'XX')
    cat_code = {
        'Vaca': 'VA', 'Novillo': 'NO', 'Novillito': 'NT',
        'Vaquillona': 'VQ', 'Toro': 'TO', 'Bubalino': 'BU',
    }.get(cat, 'XX')

    medias = parsed_data.get('medias_reses', 0)
    kg_entrada = parsed_data.get('kg_entrada', 0)
    # Peso promedio por media
    peso_prom = round(kg_entrada / medias) if medias > 0 else 0

    calidad_short = {
        'Standard': 'Std', 'Búfalo': 'Buf',
        'Premium Black': 'PBlack', 'Exportación': 'Export',
    }.get(calidad, calidad)

    return f"{dd}{mm} {cat_code} {medias}m {peso_prom}kg {calidad_short}"


@st.cache_data(ttl=600)  # Cache 10 minutos
def cargar_precios_google_sheets():
    """
    Lee precios en vivo desde Google Sheets.
    Aplica descuentos a PEDIDOS YA y MAXICONSUMO.
    Retorna (price_matrix, fecha_actualizacion, status).
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        if not os.path.exists(CREDENTIALS_PATH):
            return None, None, 'Sin credentials.json'

        SCOPES = [
            'https://www.googleapis.com/auth/spreadsheets.readonly',
            'https://www.googleapis.com/auth/drive.metadata.readonly',
        ]
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(SPREADSHEET_PRECIOS)

        # Fecha última modificación
        fecha_mod = None
        try:
            fecha_mod = sh.lastUpdateTime
        except Exception:
            pass

        ws = sh.worksheet('CORTES DE CARNE')
        all_data = ws.get_all_values()

        if len(all_data) < 3:
            return None, fecha_mod, 'Hoja vacía'

        # Row 1 = categorías, Row 2 = nombres clientes (col 0 = producto)
        headers = [str(h).strip() for h in all_data[1]]

        from pdf_parser import clasificar_corte
        price_matrix = {}
        for row in all_data[2:]:
            prod = row[0].strip() if row[0] else ''
            if not prod:
                continue
            grupo = clasificar_corte(prod)
            if not grupo or grupo in ('SIN CLASIFICAR', 'GRASA'):
                continue

            if grupo not in price_matrix:
                price_matrix[grupo] = {}

            for col_idx in range(1, len(row)):
                if col_idx >= len(headers):
                    break
                hdr = headers[col_idx].strip()
                if not hdr:
                    continue
                val_str = row[col_idx].strip()
                if not val_str:
                    continue
                # Parsear precio: "$18,261.00" o "18,592" o "$11,052.60" o "18261"
                val_str = val_str.replace('$', '').replace(' ', '').strip()
                if not val_str:
                    continue
                try:
                    # Siempre sacar comas (separador de miles) y convertir
                    val = float(val_str.replace(',', ''))
                except (ValueError, TypeError):
                    continue

                if val <= 0:
                    continue

                precio_final = round(val)
                from config import (DESCUENTO_PEYA_LOGISTICA,
                                     DESCUENTO_PEYA_MERMA,
                                     DESCUENTO_PEYA_COMISION)

                # PEDIDOS YA: SIEMPRE recalculamos el NETO con la fórmula vigente
                # (5% logística + 2% merma + 4,2% comisión, compuestos).
                # Ignoramos la columna "NETOS PEYA" de la planilla porque suele
                # tener valores con descuentos viejos.
                if hdr == 'PEDIDOS YA':
                    # Tomar el MAYOR PEDIDOS YA bruto entre filas del mismo grupo
                    if 'PEDIDOS YA' not in price_matrix[grupo] \
                       or precio_final > price_matrix[grupo]['PEDIDOS YA']:
                        price_matrix[grupo]['PEDIDOS YA'] = precio_final
                        neto = round(val * (1 - DESCUENTO_PEYA_LOGISTICA)
                                           * (1 - DESCUENTO_PEYA_MERMA)
                                           * (1 - DESCUENTO_PEYA_COMISION))
                        price_matrix[grupo]['NETOS PEYA'] = neto
                    continue

                # PEYA BLACK
                if hdr == 'PEYA BLACK':
                    if 'PEYA BLACK' not in price_matrix[grupo] \
                       or precio_final > price_matrix[grupo]['PEYA BLACK']:
                        price_matrix[grupo]['PEYA BLACK'] = precio_final
                        neto_black = round(val * (1 - DESCUENTO_PEYA_LOGISTICA)
                                                  * (1 - DESCUENTO_PEYA_MERMA)
                                                  * (1 - DESCUENTO_PEYA_COMISION))
                        price_matrix[grupo]['NETOS PEYA BLACK'] = neto_black
                    continue

                # Columna "NETOS PEYA" de la planilla → fallback solo si NO
                # tenemos PEDIDOS YA bruto del cual recalcular.
                if hdr == 'NETOS PEYA':
                    if 'NETOS PEYA' not in price_matrix[grupo] \
                       and 'PEDIDOS YA' not in price_matrix[grupo]:
                        price_matrix[grupo]['NETOS PEYA'] = precio_final
                    continue

                # Descuentos por cliente (merma/comisión)
                if 'MAXICONSUMO' in hdr:
                    precio_final = round(val * (1 - 0.035))
                elif hdr == 'LIBERTAD':
                    precio_final = round(val * (1 - 0.05))
                elif 'DIARCO' in hdr:
                    precio_final = round(val * (1 - 0.015))
                elif 'VITAL' in hdr:
                    precio_final = round(val * (1 - 0.03))

                # Guardar el mayor precio por grupo/cliente (merge de filas)
                if hdr not in price_matrix[grupo] or precio_final > price_matrix[grupo][hdr]:
                    price_matrix[grupo][hdr] = precio_final

        return price_matrix, fecha_mod, f'{len(price_matrix)} grupos cargados'

    except Exception as e:
        return None, None, str(e)


FECHA_CORTE_PRECIOS = datetime(2026, 3, 30).date()  # A partir de esta fecha se usan precios nuevos


def parsear_fecha_romaneo(fecha_str):
    """Parsea fecha del romaneo (DD/MM/YYYY) a date."""
    if not fecha_str:
        return None
    try:
        parts = fecha_str.strip().split('/')
        if len(parts) == 3:
            return datetime(int(parts[2]), int(parts[1]), int(parts[0])).date()
        elif len(parts) == 2:
            # DD/MM sin año → asumir 2026
            return datetime(2026, int(parts[1]), int(parts[0])).date()
    except (ValueError, IndexError):
        pass
    return None


def cargar_price_matrix(fecha_romaneo_str=None):
    """
    Fecha romaneo >= 30/03/2026 → precios live de Google Sheets.
    Fecha romaneo < 30/03/2026 → PRECIOS_BASE (históricos).
    """
    import copy

    fecha = parsear_fecha_romaneo(fecha_romaneo_str)
    usar_nuevos = fecha is not None and fecha >= FECHA_CORTE_PRECIOS

    if usar_nuevos:
        precios_gs, fecha_mod, status_gs = cargar_precios_google_sheets()
        if precios_gs:
            return precios_gs, True, f'Precios nuevos (Google Sheets)', fecha_mod
        # Si falla GSheets, fallback
        return copy.deepcopy(PRECIOS_BASE), False, 'Error leyendo GSheets, usando precargados', None
    else:
        motivo = f'Romaneo anterior al {FECHA_CORTE_PRECIOS.strftime("%d/%m/%Y")}' if fecha else 'Sin fecha'
        return copy.deepcopy(PRECIOS_BASE), False, f'Precios anteriores ({motivo})', None


# ══════════════════════════════════════════════════════════════════════
# CONTENIDO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="tf-hero">
  <div>
    <h1>TF Carnes</h1>
    <p>Romaneo · Pricing · Revoleo</p>
  </div>
  <div class="tf-hero-meta">Orgullosa Carne Argentina</div>
</div>
""", unsafe_allow_html=True)

tab_upload, tab_analisis, tab_control, tab_comprador, tab_pricing, tab_revoleo, tab_export, tab_historial, tab_costos = st.tabs([
    "📤 Cargar archivos",
    "📊 Análisis",
    "🔍 Control de cortes",
    "🛒 Reporte Comprador",
    "💰 Pricing",
    "🔄 Revoleo",
    "🌍 Exportación",
    "📈 Historial",
    "🏭 Costos Operativos"
])

# ── TAB 1: CARGA DE ARCHIVOS ──
with tab_upload:
    from drive_loader import listar_pdfs_drive, descargar_pdf, organizar_por_mes_semana

    fuente = st.radio(
        "**Fuente de romaneos**",
        ['📁 Google Drive (automático)', '📤 Subir archivos manualmente'],
        horizontal=True, key='fuente_romaneos'
    )

    if fuente == '📁 Google Drive (automático)':
        st.markdown("### 📁 Romaneos desde Google Drive")
        st.markdown("Se leen automáticamente de la carpeta compartida.")

        # Listar archivos (cacheable) para sacar los meses disponibles
        if st.button("🔍 Listar archivos de Drive", key='listar_drive'):
            with st.spinner("Listando..."):
                st.session_state.drive_files_cache = listar_pdfs_drive(CREDENTIALS_PATH)

        drive_files_all = st.session_state.get('drive_files_cache', [])

        # Filtro de mes basado en el nombre del archivo (formato DDMM) o createdTime
        def _mes_archivo(f):
            """Devuelve 'YYYY-MM Mes' o 'Sin fecha'."""
            import re
            from datetime import datetime as _dt
            meses_n = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
                       5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
                       9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}
            # Patrón DDMM al inicio o DD/MM o DD-MM
            nombre = f.get('name', '')
            m = re.search(r'(?:^|\s)(\d{2})(\d{2})(?:\s|_|[A-Za-z])', ' ' + nombre)
            if m:
                dd, mm = int(m.group(1)), int(m.group(2))
                if 1 <= mm <= 12 and 1 <= dd <= 31:
                    yyyy = 2026  # default año actual del sistema
                    # Si createdTime disponible, usar ese año
                    ct = f.get('createdTime', '')
                    if ct:
                        try:
                            yyyy = _dt.fromisoformat(ct.replace('Z', '+00:00')).year
                        except Exception:
                            pass
                    return f"{yyyy}-{mm:02d} {meses_n.get(mm, '?')}"
            # Fallback: createdTime
            ct = f.get('createdTime', '')
            if ct:
                try:
                    d = _dt.fromisoformat(ct.replace('Z', '+00:00'))
                    return f"{d.year}-{d.month:02d} {meses_n.get(d.month, '?')}"
                except Exception:
                    pass
            return 'Sin fecha'

        if drive_files_all:
            # Agrupar por mes
            meses_dict = {}
            for f in drive_files_all:
                k = _mes_archivo(f)
                meses_dict.setdefault(k, []).append(f)

            meses_ordenados = sorted([m for m in meses_dict.keys() if m != 'Sin fecha'], reverse=True)
            if 'Sin fecha' in meses_dict:
                meses_ordenados.append('Sin fecha')

            col_m1, col_m2 = st.columns([2, 1])
            sel_mes_drive = col_m1.selectbox(
                "Mes a cargar",
                ['Todos'] + meses_ordenados,
                key='sel_mes_drive',
                help="Filtrar por mes antes de descargar"
            )
            col_m2.markdown(f"**Total en Drive:** {len(drive_files_all)}")

            # Filtrar
            if sel_mes_drive == 'Todos':
                drive_files_filtrados = drive_files_all
            else:
                drive_files_filtrados = meses_dict.get(sel_mes_drive, [])

            # Resumen de los meses disponibles
            with st.expander(f"📊 Resumen por mes ({len(meses_ordenados)} meses detectados)", expanded=False):
                for m in meses_ordenados:
                    st.caption(f"**{m}**: {len(meses_dict[m])} archivos")

            st.info(f"📋 {len(drive_files_filtrados)} archivo(s) a procesar con el filtro actual")
        else:
            drive_files_filtrados = []

        if st.button("🔄 Cargar desde Drive", type="primary", use_container_width=True,
                     disabled=not drive_files_filtrados):
            drive_files = drive_files_filtrados if drive_files_filtrados else listar_pdfs_drive(CREDENTIALS_PATH)

            if not drive_files:
                st.warning("No se encontraron PDFs en la carpeta de Drive.")
            else:
                st.success(f"✅ {len(drive_files)} PDF(s) a procesar")

                # Descargar, clasificar y parsear cada PDF
                parsed = []
                remanejos = []
                descartados = []
                progress = st.progress(0)
                temp_paths = {}
                for i, df in enumerate(drive_files):
                    with st.spinner(f"Descargando {df['name'][:40]}..."):
                        try:
                            tmp_path = descargar_pdf(CREDENTIALS_PATH, df['id'])
                            tipo_pdf = detectar_tipo_pdf(tmp_path)

                            if tipo_pdf == 'entrada':
                                descartados.append({'archivo': df['name'], 'motivo': 'Romaneo de ENTRADA (sin salidas)'})
                                os.unlink(tmp_path)
                            elif tipo_pdf == 'remanejo':
                                result = parse_remanejo_pdf(tmp_path)
                                result['archivo'] = df['name']
                                result['drive_created'] = df.get('createdTime', '')
                                remanejos.append(result)
                                os.unlink(tmp_path)
                            elif tipo_pdf == 'romaneo':
                                result = parse_romaneo_pdf(tmp_path)
                                # Verificar que tenga cortes reales
                                meat = [c for c in result.get('cortes', []) if c.get('grupo') != 'GRASA']
                                if not meat:
                                    descartados.append({'archivo': df['name'], 'motivo': '0 cortes de carne'})
                                    os.unlink(tmp_path)
                                else:
                                    result['archivo'] = df['name']
                                    result['drive_id'] = df['id']
                                    result['drive_created'] = df.get('createdTime', '')
                                    result['precio_compra'] = precio_compra
                                    temp_paths[df['name']] = tmp_path
                                    parsed.append(result)
                            else:
                                # Desconocido — intentar parsear como romaneo
                                result = parse_romaneo_pdf(tmp_path)
                                meat = [c for c in result.get('cortes', []) if c.get('grupo') != 'GRASA']
                                if not meat:
                                    descartados.append({'archivo': df['name'], 'motivo': 'Sin cortes detectados'})
                                    os.unlink(tmp_path)
                                else:
                                    result['archivo'] = df['name']
                                    result['drive_id'] = df['id']
                                    result['drive_created'] = df.get('createdTime', '')
                                    result['precio_compra'] = precio_compra
                                    temp_paths[df['name']] = tmp_path
                                    parsed.append(result)
                        except Exception as e:
                            parsed.append({'archivo': df['name'], 'error': str(e)})
                    progress.progress((i + 1) / len(drive_files))

                # Guardar remanejos en session
                if 'remanejos' not in st.session_state:
                    st.session_state.remanejos = []
                st.session_state.remanejos = remanejos

                # Mostrar descartados
                if descartados:
                    with st.expander(f"⏭️ {len(descartados)} archivo(s) descartados", expanded=False):
                        for d in descartados:
                            st.caption(f"- {d['archivo']}: {d['motivo']}")

                # Mostrar remanejos
                if remanejos:
                    st.info(f"🔄 {len(remanejos)} remanejo(s) detectados (ver abajo)")

                # Manejar correcciones: reemplazar romaneo original con la corrección
                correcciones = []
                no_correcciones = []
                for p in parsed:
                    if 'error' in p:
                        no_correcciones.append(p)
                        continue
                    es_corr, nombre_base = es_correccion(p.get('archivo', ''))
                    if es_corr:
                        p['es_correccion'] = True
                        p['corrige_a'] = nombre_base
                        correcciones.append(p)
                    else:
                        no_correcciones.append(p)

                # Aplicar correcciones: buscar el original y reemplazarlo
                for corr in correcciones:
                    base = corr['corrige_a'].upper().replace('.PDF', '')
                    reemplazado = False
                    for i, orig in enumerate(no_correcciones):
                        if 'error' in orig:
                            continue
                        orig_name = orig.get('archivo', '').upper().replace('.PDF', '')
                        if base in orig_name or orig_name in base:
                            # Guardar referencia al original
                            corr['original_archivo'] = orig['archivo']
                            corr['original_cortes'] = len(orig.get('cortes', []))
                            no_correcciones[i] = corr  # reemplazar
                            reemplazado = True
                            break
                    if not reemplazado:
                        # No encontró original, agregar la corrección sola
                        no_correcciones.append(corr)

                if correcciones:
                    st.info(f"📝 {len(correcciones)} corrección(es) aplicadas")
                    for c in correcciones:
                        orig = c.get('original_archivo', '?')
                        st.caption(f"- {c['archivo']} → reemplaza a {orig}")

                parsed = no_correcciones

                # Cruzar con planilla de compras
                with st.spinner("Cruzando con planilla de compras..."):
                    for p in parsed:
                        if 'error' in p:
                            continue
                        pdf_tmp = temp_paths.get(p.get('archivo'))
                        cruce, tropas, status = cruzar_precio_compra_gsheets(p, pdf_tmp)
                        if cruce:
                            p['precio_compra_auto'] = cruce['precio_promedio']
                            p['tropas_match'] = cruce['matches']
                            p['tropas_sin_match'] = cruce.get('tropas_sin_match', [])
                            p['precio_compra'] = cruce['precio_promedio']
                            # Tipo de origen: faena propia vs media comprada
                            kg_compra = sum(m.get('kg', 0) for m in cruce['matches']
                                            if m.get('es_compra_media'))
                            kg_faena = sum(m.get('kg', 0) for m in cruce['matches']
                                           if not m.get('es_compra_media'))
                            p['kg_compra_media'] = kg_compra
                            p['kg_faena_propia'] = kg_faena
                            if kg_compra > 0 and kg_faena == 0:
                                p['origen_carcaza'] = 'compra_media'
                            elif kg_faena > 0 and kg_compra == 0:
                                p['origen_carcaza'] = 'faena_propia'
                            elif kg_compra > 0:
                                p['origen_carcaza'] = 'mixto'
                            else:
                                p['origen_carcaza'] = 'desconocido'
                        p['tropas_encontradas'] = tropas if tropas else set()
                        p['cruce_status'] = status

                        # Días faena→producción
                        if cruce and cruce.get('matches'):
                            fecha_rom_dt = parsear_fecha_romaneo(p.get('fecha', ''))
                            if fecha_rom_dt:
                                dias_list = []
                                for m in cruce['matches']:
                                    ff = m.get('fecha_faena_dt')
                                    if ff:
                                        dias_list.append((fecha_rom_dt - ff).days)
                                if dias_list:
                                    p['dias_faena_produccion'] = round(sum(dias_list) / len(dias_list), 1)

                # Limpiar temporales
                for tmp_path in temp_paths.values():
                    try:
                        os.unlink(tmp_path)
                    except Exception:
                        pass

                st.session_state.parsed_files = parsed
                st.rerun()

    else:
        st.markdown("### 📤 Subir PDFs manualmente")
        st.markdown("Podés cargar **uno o varios** archivos. Si cargás más de uno, "
                    "elegís análisis **individual** o **acumulado**.")

    uploaded_files = st.file_uploader(
        "Arrastrá los PDFs acá",
        type=['pdf'],
        accept_multiple_files=True,
        key='pdfs',
    )

    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} archivo(s) cargado(s)")

        if st.button("🔍 Parsear archivos", type="primary", use_container_width=True):
            parsed = []
            temp_paths = {}  # Guardar paths temporales para cruce de tropas
            progress = st.progress(0)
            for i, uf in enumerate(uploaded_files):
                with st.spinner(f"Parseando {uf.name}..."):
                    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                        tmp.write(uf.read())
                        tmp_path = tmp.name
                    try:
                        result = parse_romaneo_pdf(tmp_path)
                        result['archivo'] = uf.name
                        result['precio_compra'] = precio_compra
                        temp_paths[uf.name] = tmp_path
                        parsed.append(result)
                    except Exception as e:
                        parsed.append({'archivo': uf.name, 'error': str(e)})
                        os.unlink(tmp_path)
                progress.progress((i + 1) / len(uploaded_files))

            # Cruzar con planilla de compras de Google Sheets (precio + fecha faena)
            with st.spinner("Cruzando con planilla de compras..."):
                for p in parsed:
                    if 'error' in p:
                        continue
                    pdf_tmp = temp_paths.get(p.get('archivo'))
                    cruce, tropas, status = cruzar_precio_compra_gsheets(p, pdf_tmp)
                    if cruce:
                        p['precio_compra_auto'] = cruce['precio_promedio']
                        p['tropas_match'] = cruce['matches']
                        p['precio_compra'] = cruce['precio_promedio']
                        kg_compra = sum(m.get('kg', 0) for m in cruce['matches']
                                        if m.get('es_compra_media'))
                        kg_faena = sum(m.get('kg', 0) for m in cruce['matches']
                                       if not m.get('es_compra_media'))
                        p['kg_compra_media'] = kg_compra
                        p['kg_faena_propia'] = kg_faena
                        if kg_compra > 0 and kg_faena == 0:
                            p['origen_carcaza'] = 'compra_media'
                        elif kg_faena > 0 and kg_compra == 0:
                            p['origen_carcaza'] = 'faena_propia'
                        elif kg_compra > 0:
                            p['origen_carcaza'] = 'mixto'
                        else:
                            p['origen_carcaza'] = 'desconocido'
                    p['tropas_encontradas'] = tropas if tropas else set()
                    p['cruce_status'] = status

                    # Calcular días entre faena y producción (fecha romaneo)
                    if cruce and cruce.get('matches'):
                        fecha_rom_str = p.get('fecha', '')
                        fecha_rom_dt = parsear_fecha_romaneo(fecha_rom_str)
                        if fecha_rom_dt:
                            dias_list = []
                            for m in cruce['matches']:
                                ff = m.get('fecha_faena_dt')
                                if ff:
                                    dias = (fecha_rom_dt - ff).days
                                    dias_list.append(dias)
                            if dias_list:
                                p['dias_faena_produccion'] = round(sum(dias_list) / len(dias_list), 1)

            # Limpiar archivos temporales
            for tmp_path in temp_paths.values():
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

            st.session_state.parsed_files = parsed
            st.rerun()

    # Mostrar archivos parseados con opción de renombrar
    if st.session_state.parsed_files:
        st.markdown("---")
        st.markdown("### 📋 Archivos parseados")

        # ── Selección y segmentación por romaneo (por CONTENIDO, no por nombre) ──
        valid_all = [p for p in st.session_state.parsed_files if 'error' not in p]
        registro = seg.cargar_registro(SEGMENTOS_PATH)

        # Identidad estable + canal (sugerido o guardado) para cada romaneo
        for p in valid_all:
            p['rid'] = seg.romaneo_id(p)
            s_sug, conf = seg.segmento_de(p, registro)
            if 'segmento' not in p:
                p['segmento'] = s_sug
            p['segmento_confirmado'] = conf

        st.caption("Marcá qué romaneos entran en el análisis y confirmá/corregí el canal. "
                   "El canal queda guardado para la próxima vez (aunque el archivo se llame igual).")

        df_sel = pd.DataFrame([{
            'Incluir': not p.get('excluido', False),
            'Fecha': seg.resumen(p)['fecha'],
            'Categoría': p.get('categoria', '—'),
            'Medias': p.get('medias_reses', 0),
            'Kg': int(round(p.get('kg_entrada', 0) or 0)),
            'Cortes': len(seg.grupos_distintos(p)),
            'Destino': seg.destino_principal(p),
            'Segmento': p.get('segmento', 'Consumo'),
        } for p in valid_all])

        edited = st.data_editor(
            df_sel,
            key='grid_segmentos',
            hide_index=True,
            use_container_width=True,
            num_rows='fixed',
            column_config={
                'Incluir': st.column_config.CheckboxColumn('✔', help='Incluir en el análisis'),
                'Segmento': st.column_config.SelectboxColumn('Canal', options=seg.SEGMENTOS, required=True),
                'Fecha': st.column_config.TextColumn('Fecha', disabled=True),
                'Categoría': st.column_config.TextColumn('Categoría', disabled=True),
                'Medias': st.column_config.NumberColumn('Medias', disabled=True),
                'Kg': st.column_config.NumberColumn('Kg entrada', disabled=True),
                'Cortes': st.column_config.NumberColumn('N° cortes', disabled=True),
                'Destino': st.column_config.TextColumn('Destino ppal', disabled=True),
            },
        )

        # Aplicar cambios de la grilla (mapeo por posición; el orden no cambia)
        for i, p in enumerate(valid_all):
            row = edited.iloc[i]
            p['excluido'] = not bool(row['Incluir'])
            nuevo_seg = row['Segmento']
            if nuevo_seg != p.get('segmento'):
                p['segmento'] = nuevo_seg
                registro = seg.set_segmento(SEGMENTOS_PATH, registro, p, nuevo_seg)

        valid_parsed = [p for p in st.session_state.parsed_files
                        if 'error' not in p and not p.get('excluido')]

        n_excl = len(valid_all) - len(valid_parsed)
        if n_excl:
            st.caption(f"Incluidos {len(valid_parsed)} de {len(valid_all)} ({n_excl} fuera del análisis)")

        organizado = organizar_por_mes_semana(valid_parsed)

        if organizado:
            for mes, semanas in organizado.items():
                st.markdown(f"#### 📅 {mes}")
                for semana, archs in semanas.items():
                    kg_sem = sum(a.get('kg_entrada', 0) for a in archs)
                    medias_sem = sum(a.get('medias_reses', 0) for a in archs)
                    st.markdown(f"**{semana}** — {len(archs)} romaneos, {medias_sem} medias, {kg_sem:,.0f} kg")
            st.markdown("---")

        for i, p in enumerate(st.session_state.parsed_files):
            if 'error' in p:
                st.error(f"❌ **{p['archivo']}**: {p['error']}")
                continue

            kg_entrada = p.get('kg_entrada', 0)
            meat_kg = sum(c['kg'] for c in p.get('cortes', []) if c.get('grupo') != 'GRASA')
            rend = (meat_kg / kg_entrada * 100) if kg_entrada > 0 else 0

            # Nombre sugerido
            nombre_sugerido = generar_nombre_archivo(p, calidad)

            # Asegurar precio de compra siempre seteado
            if not p.get('precio_compra') or p['precio_compra'] == 0:
                p['precio_compra'] = p.get('precio_compra_auto', precio_compra)

            # Días faena→producción (visible sin expandir)
            dias_fp = p.get('dias_faena_produccion')
            dias_txt = ''
            if dias_fp is not None:
                if dias_fp < 3:
                    dias_txt = f' — 🟢 {dias_fp:.0f}d'
                elif dias_fp < 5:
                    dias_txt = f' — 🟡 {dias_fp:.0f}d'
                else:
                    dias_txt = f' — 🔴 {dias_fp:.0f}d'

            with st.expander(
                f"📄 {p['archivo']} — {p.get('categoria', '?')} — "
                f"{p.get('medias_reses', 0)} medias — "
                f"{kg_entrada:,.0f} kg — Rend: {rend:.1f}% — "
                f"${p['precio_compra']:,.0f}/kg{dias_txt}",
                expanded=False
            ):
                # Renombrar
                col_name, col_btn = st.columns([3, 1])
                nuevo_nombre = col_name.text_input(
                    "Nombre para el análisis",
                    value=nombre_sugerido,
                    key=f'rename_{i}',
                    help="Formato sugerido: DDMM CAT MEDIAS PESO CALIDAD"
                )
                if nuevo_nombre:
                    p['nombre_analisis'] = nuevo_nombre

                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("Medias", p.get('medias_reses', 0))
                col2.metric("Kg entrada", f"{kg_entrada:,.0f}")
                col3.metric("Kg carne", f"{meat_kg:,.0f}")
                col4.metric("Rendimiento", f"{rend:.1f}%")
                n_cortes = len([c for c in p.get('cortes', []) if c.get('grupo') != 'GRASA'])
                col5.metric("Líneas de corte", n_cortes)

                peso_prom = round(kg_entrada / p.get('medias_reses', 1)) if p.get('medias_reses', 0) > 0 else 0
                st.markdown(f"**Peso promedio por media:** {peso_prom} kg")

                # Precio de compra — cruce automático con Google Sheets
                if p.get('precio_compra_auto'):
                    tropas_info = p.get('tropas_match', [])
                    sin_m = p.get('tropas_sin_match', [])

                    # ── Inputs manuales para tropas sin match ──
                    precios_manuales_tropas = {}
                    if sin_m:
                        st.warning(
                            f"⚠️ {len(sin_m)} tropa(s) sin match en planilla de compras "
                            "(aún no liquidadas). Podés cargar el precio manualmente:"
                        )
                        for t_sm in sin_m:
                            t_id = t_sm['tropa'] if isinstance(t_sm, dict) else str(t_sm)
                            t_kg = t_sm.get('kg', 0) if isinstance(t_sm, dict) else 0
                            col_t1, col_t2 = st.columns([1, 2])
                            col_t1.markdown(f"**Tropa {t_id}** ({t_kg:,.0f} kg)")
                            precio_manual_t = col_t2.number_input(
                                f"Precio $/kg — tropa {t_id}",
                                min_value=0, value=0, step=100,
                                key=f'precio_tropa_{i}_{t_id}',
                                label_visibility='collapsed',
                            )
                            if precio_manual_t > 0:
                                precios_manuales_tropas[t_id] = {
                                    'tropa': t_id,
                                    'precio': precio_manual_t,
                                    'kg': t_kg,
                                    'fecha': 'manual',
                                    'manual': True,
                                }

                    # ── Recalcular ponderado incluyendo los manuales ──
                    todas_tropas = list(tropas_info) + list(precios_manuales_tropas.values())
                    kg_total = sum(m.get('kg', 0) for m in todas_tropas)
                    if kg_total > 0:
                        precio_ponderado = sum(m['precio'] * m.get('kg', 0)
                                               for m in todas_tropas) / kg_total
                        precio_ponderado = round(precio_ponderado)
                    else:
                        precio_ponderado = p['precio_compra_auto']

                    # Si hubo manuales, actualizar la data
                    if precios_manuales_tropas:
                        p['tropas_match'] = todas_tropas
                        # Quitar de sin_match los que ahora tienen precio manual
                        p['tropas_sin_match'] = [
                            t for t in sin_m
                            if (t['tropa'] if isinstance(t, dict) else str(t))
                               not in precios_manuales_tropas
                        ]

                    p['precio_compra_auto'] = precio_ponderado

                    # Mostrar detalle del cálculo
                    def _etiqueta_tropa(m):
                        etiqueta = f"**{m['tropa']}**"
                        if m.get('manual'):
                            etiqueta += '📝'
                        n_part = m.get('partidas', 1) or 1
                        if n_part > 1:
                            etiqueta += f" (×{n_part} partidas → promedio ponderado)"
                        return (f"{etiqueta}: ${m['precio']:,.0f} × "
                                f"{m.get('kg', 0):,.0f} kg")
                    tropas_txt = ' | '.join(
                        [_etiqueta_tropa(m) for m in todas_tropas if m.get('precio', 0) > 0]
                    )
                    st.success(
                        f"💲 **Precio ponderado: ${precio_ponderado:,.0f}/kg**\n\n"
                        f"{tropas_txt}"
                    )

                    usar_auto = st.checkbox(
                        "Usar precio ponderado", value=True, key=f'auto_precio_{i}')
                    if usar_auto:
                        p['precio_compra'] = precio_ponderado
                    else:
                        precio_manual_g = st.number_input(
                            "Precio manual general $/kg", value=precio_compra,
                            step=100, key=f'precio_manual_{i}')
                        p['precio_compra'] = precio_manual_g
                else:
                    status = p.get('cruce_status', '')
                    if status and status != 'ok':
                        st.warning(f"⚠️ {status}")
                    p['precio_compra'] = precio_compra

                # Días entre faena y producción
                dias_fp = p.get('dias_faena_produccion')
                if dias_fp is not None:
                    if dias_fp < 3:
                        st.markdown(f"🟢 **Faena → Producción: {dias_fp:.0f} días**")
                    elif dias_fp < 5:
                        st.markdown(f"🟡 **Faena → Producción: {dias_fp:.0f} días**")
                    else:
                        st.markdown(f"🔴 **Faena → Producción: {dias_fp:.0f} días** — demasiado tiempo")

                if p.get('cortes'):
                    df = pd.DataFrame(p['cortes'])
                    cols_show = ['corte', 'grupo', 'tipo', 'piezas', 'unidades', 'kg', 'cliente', 'contramarca']
                    cols_avail = [c for c in cols_show if c in df.columns]
                    st.dataframe(df[cols_avail], use_container_width=True, hide_index=True)

                # Alertas de cortes faltantes
                faltantes = detectar_cortes_faltantes(p)
                caros_faltantes = [f for f in faltantes if f['es_caro']]
                otros_faltantes = [f for f in faltantes if not f['es_caro']]
                if caros_faltantes:
                    st.error(
                        f"🔴 **CORTES CAROS FALTANTES:** "
                        f"{', '.join(f['grupo'] for f in caros_faltantes)}"
                    )
                if otros_faltantes:
                    st.caption(
                        f"Otros faltantes: {', '.join(f['grupo'] for f in otros_faltantes)}"
                    )

                # Cortes sin clasificar — opción de reclasificar
                sin_clasificar = [c for c in p.get('cortes', []) if c.get('grupo') == 'SIN CLASIFICAR']
                if sin_clasificar:
                    st.warning(f"⚠️ **{len(sin_clasificar)} corte(s) sin clasificar**")
                    from config import GRUPOS_POR_CALIDAD
                    opciones_grupo = ['SIN CLASIFICAR'] + [g for g, _ in GRUPOS_POR_CALIDAD['Standard']]
                    for j, sc in enumerate(sin_clasificar):
                        col_sc1, col_sc2 = st.columns([2, 1])
                        col_sc1.markdown(f"**{sc['corte']}** ({sc['kg']:.2f} kg)")
                        nuevo_grupo = col_sc2.selectbox(
                            "Asignar a", opciones_grupo,
                            key=f'reclas_{i}_{j}',
                            help=f"Reclasificar '{sc['corte']}'"
                        )
                        if nuevo_grupo != 'SIN CLASIFICAR':
                            sc['grupo'] = nuevo_grupo

        # ── Alertar faenas sin romaneo (a partir de abril 2026) ──
        if st.session_state.parsed_files:
            from datetime import date
            hoy = date.today()
            if hoy >= date(2026, 4, 1):
                compras_data, compras_status = cargar_compras_google_sheets()
                if compras_data:
                    def _resumen_tropa(entradas):
                        primera = entradas[0]
                        monto_tot = sum(e['monto'] for e in entradas if e.get('monto'))
                        kg_tot = sum(e['kg_sheet'] for e in entradas if e.get('kg_sheet'))
                        if monto_tot > 0 and kg_tot > 0:
                            precio = monto_tot / kg_tot
                        else:
                            precio = sum(e['precio'] for e in entradas) / len(entradas)
                        return {
                            'fecha': primera['fecha'],
                            'fecha_faena_dt': primera.get('fecha_faena_dt'),
                            'tipo': primera['tipo'],
                            'precio': precio,
                            'partidas': len(entradas),
                            'monto': monto_tot,
                            'kg': kg_tot,
                        }

                    # Resumen por tropa (con año inferido)
                    resumen_por_tropa = {t: _resumen_tropa(e) for t, e in compras_data.items()}

                    # Tropas ya procesadas en romaneos cargados
                    tropas_procesadas = set()
                    for p in st.session_state.parsed_files:
                        if 'error' in p:
                            continue
                        for t in p.get('tropas_encontradas', set()):
                            tropas_procesadas.add(t)
                        for m in p.get('tropas_match', []):
                            tropas_procesadas.add(m['tropa'])

                    st.markdown("---")
                    st.markdown("#### 🗓️ Control de faenas vs romaneos")

                    # Selector de mes + año
                    meses_nombres = ['Ene','Feb','Mar','Abr','May','Jun',
                                     'Jul','Ago','Sep','Oct','Nov','Dic']
                    años_disponibles = sorted({r['fecha_faena_dt'].year
                                               for r in resumen_por_tropa.values()
                                               if r['fecha_faena_dt']}, reverse=True)
                    if not años_disponibles:
                        años_disponibles = [hoy.year]
                    col_sel1, col_sel2 = st.columns(2)
                    año_sel = col_sel1.selectbox(
                        "Año", años_disponibles,
                        index=años_disponibles.index(hoy.year) if hoy.year in años_disponibles else 0,
                        key='faena_anio_sel'
                    )
                    mes_sel = col_sel2.selectbox(
                        "Mes", list(range(1, 13)),
                        format_func=lambda m: meses_nombres[m-1],
                        index=hoy.month - 1,
                        key='faena_mes_sel'
                    )

                    # Filtrar tropas del mes/año seleccionado
                    tropas_mes = {
                        t: r for t, r in resumen_por_tropa.items()
                        if r['fecha_faena_dt']
                        and r['fecha_faena_dt'].month == mes_sel
                        and r['fecha_faena_dt'].year == año_sel
                    }
                    sin_romaneo = {t: info for t, info in tropas_mes.items()
                                   if t not in tropas_procesadas}

                    if not tropas_mes:
                        st.info(f"No hay faenas cargadas en {meses_nombres[mes_sel-1]} {año_sel}.")
                    elif sin_romaneo:
                        st.warning(
                            f"⚠️ **{len(sin_romaneo)} de {len(tropas_mes)} faena(s) "
                            f"de {meses_nombres[mes_sel-1]} {año_sel} sin romaneo cargado:**"
                        )
                        for tropa, info in sorted(sin_romaneo.items()):
                            extra = f" · {info['partidas']} partidas" if info['partidas'] > 1 else ''
                            st.markdown(
                                f"- Tropa **{tropa}** — Faena {info['fecha']} — "
                                f"{info['tipo']} — ${info['precio']:,.0f}/kg{extra}"
                            )
                    else:
                        st.success(
                            f"✅ Las {len(tropas_mes)} faena(s) de "
                            f"{meses_nombres[mes_sel-1]} {año_sel} tienen romaneo cargado"
                        )

                    # ── Resumen anual 2026 ──
                    st.markdown("---")
                    st.markdown("#### 📊 Resumen año 2026")
                    tropas_2026 = {t: r for t, r in resumen_por_tropa.items()
                                    if r['fecha_faena_dt'] and r['fecha_faena_dt'].year == 2026}
                    if tropas_2026:
                        # Segmentar por origen: tropas que empiezan con "7" → media comprada
                        def _es_compra(tropa):
                            return tropa.startswith('7') and len(tropa) >= 4
                        tropas_compra = {t: r for t, r in tropas_2026.items() if _es_compra(t)}
                        tropas_faena = {t: r for t, r in tropas_2026.items() if not _es_compra(t)}

                        total_tropas = len(tropas_2026)
                        total_kg_ano = sum(r['kg'] for r in tropas_2026.values() if r.get('kg'))
                        total_monto_ano = sum(r['monto'] for r in tropas_2026.values() if r.get('monto'))
                        precio_pond_ano = (total_monto_ano / total_kg_ano) if total_kg_ano else 0
                        tropas_con_romaneo = sum(1 for t in tropas_2026 if t in tropas_procesadas)
                        pct_cobertura = (tropas_con_romaneo / total_tropas * 100) if total_tropas else 0

                        col_a1, col_a2, col_a3, col_a4 = st.columns(4)
                        col_a1.metric("Tropas total", f"{total_tropas:,}".replace(',', '.'))
                        col_a2.metric("Kg total", f"{total_kg_ano:,.0f}".replace(',', '.'))
                        col_a3.metric("Precio prom. ponderado", f"${precio_pond_ano:,.0f}/kg".replace(',', '.'))
                        col_a4.metric("Cobertura romaneo",
                                       f"{tropas_con_romaneo}/{total_tropas}",
                                       delta=f"{pct_cobertura:.0f}%",
                                       delta_color='normal' if pct_cobertura >= 90 else 'inverse')

                        # Segmentación faena propia vs compra de medias
                        st.markdown("##### Faena propia vs Compra de medias")
                        kg_faena = sum(r['kg'] for r in tropas_faena.values() if r.get('kg'))
                        kg_compra = sum(r['kg'] for r in tropas_compra.values() if r.get('kg'))
                        monto_faena = sum(r['monto'] for r in tropas_faena.values() if r.get('monto'))
                        monto_compra = sum(r['monto'] for r in tropas_compra.values() if r.get('monto'))
                        precio_faena = (monto_faena / kg_faena) if kg_faena else 0
                        precio_compra_pond = (monto_compra / kg_compra) if kg_compra else 0

                        # Rendimiento real vs objetivo de los romaneos cargados,
                        # separado por origen (faena vs compra)
                        rend_faena_list = []
                        rend_compra_list = []
                        for pp in st.session_state.parsed_files:
                            if pp.get('error'): continue
                            kgi = pp.get('kg_entrada', 0)
                            kgc = pp.get('kg_carne', 0)
                            if kgi <= 0: continue
                            origen = pp.get('origen_carcaza', 'desconocido')
                            r_pct = kgc / kgi * 100
                            if origen == 'compra_media':
                                rend_compra_list.append(r_pct)
                            elif origen == 'faena_propia':
                                rend_faena_list.append(r_pct)
                        rend_faena_avg = (sum(rend_faena_list) / len(rend_faena_list)) if rend_faena_list else None
                        rend_compra_avg = (sum(rend_compra_list) / len(rend_compra_list)) if rend_compra_list else None

                        col_s1, col_s2, col_s3 = st.columns(3)
                        with col_s1:
                            st.markdown("**🐂 Faena propia**")
                            st.metric("Tropas", len(tropas_faena))
                            st.metric("Kg", f"{kg_faena:,.0f}".replace(',', '.'))
                            st.metric("$/kg prom.", f"${precio_faena:,.0f}".replace(',', '.'))
                            if rend_faena_avg is not None:
                                st.metric("Rend. real prom.", f"{rend_faena_avg:.2f}%".replace('.', ','))
                        with col_s2:
                            st.markdown("**🚚 Compra de medias**")
                            st.metric("Tropas", len(tropas_compra))
                            st.metric("Kg", f"{kg_compra:,.0f}".replace(',', '.'))
                            st.metric("$/kg prom.", f"${precio_compra_pond:,.0f}".replace(',', '.'))
                            if rend_compra_avg is not None:
                                st.metric("Rend. real prom.", f"{rend_compra_avg:.2f}%".replace('.', ','))
                        with col_s3:
                            st.markdown("**Δ Diferencia**")
                            if kg_faena and kg_compra:
                                diff_precio = precio_compra_pond - precio_faena
                                st.metric("Δ $/kg",
                                          f"${diff_precio:+,.0f}".replace(',', '.'),
                                          delta_color='inverse' if diff_precio > 0 else 'normal')
                                pct_compra = kg_compra / (kg_faena + kg_compra) * 100
                                st.metric("% kg comprado", f"{pct_compra:.1f}%".replace('.', ','))
                                if rend_faena_avg is not None and rend_compra_avg is not None:
                                    drend = rend_compra_avg - rend_faena_avg
                                    st.metric("Δ Rendimiento",
                                              f"{drend:+.2f}%".replace('.', ','),
                                              delta_color='normal' if drend > 0 else 'inverse')
                            else:
                                st.caption("Necesita datos en ambos tipos para comparar.")

                        # Desglose por mes
                        import pandas as pd
                        filas_mes = []
                        for m in range(1, 13):
                            tropas_m = [r for t, r in tropas_2026.items()
                                        if r['fecha_faena_dt'].month == m]
                            if not tropas_m:
                                continue
                            kg_m = sum(r['kg'] for r in tropas_m if r.get('kg'))
                            monto_m = sum(r['monto'] for r in tropas_m if r.get('monto'))
                            precio_m = (monto_m / kg_m) if kg_m else 0
                            con_rom = sum(1 for t, r in tropas_2026.items()
                                          if r['fecha_faena_dt'].month == m and t in tropas_procesadas)
                            tropas_m_compra = [t for t in tropas_2026
                                                if tropas_2026[t]['fecha_faena_dt'].month == m
                                                and _es_compra(t)]
                            filas_mes.append({
                                'Mes': meses_nombres[m-1],
                                'Tropas': len(tropas_m),
                                '— Faena': len(tropas_m) - len(tropas_m_compra),
                                '— Compra': len(tropas_m_compra),
                                'Kg total': f"{kg_m:,.0f}".replace(',', '.'),
                                '$/kg prom.': f"${precio_m:,.0f}".replace(',', '.'),
                                'Con romaneo': f"{con_rom}/{len(tropas_m)}",
                            })
                        if filas_mes:
                            st.markdown("##### Desglose por mes")
                            st.dataframe(pd.DataFrame(filas_mes),
                                         use_container_width=True, hide_index=True)
                    else:
                        st.info("Aún no hay faenas cargadas para 2026.")

        # ── Remanejos ──
        if st.session_state.get('remanejos'):
            st.markdown("---")
            st.markdown("### 🔄 Remanejos")
            for rem in st.session_state.remanejos:
                with st.expander(
                    f"🔄 {rem['archivo']} — Entrada: {rem['kg_entrada']:,.0f} kg → "
                    f"Salida: {rem['kg_salida']:,.0f} kg — "
                    f"Merma: {rem['merma_kg']:,.0f} kg ({rem['merma_pct']:.1f}%)"
                ):
                    col_r1, col_r2, col_r3, col_r4 = st.columns(4)
                    col_r1.metric("Kg entrada", f"{rem['kg_entrada']:,.0f}")
                    col_r2.metric("Kg salida", f"{rem['kg_salida']:,.0f}")
                    col_r3.metric("Merma kg", f"{rem['merma_kg']:,.0f}")
                    if rem['merma_pct'] <= 5:
                        col_r4.metric("Merma %", f"{rem['merma_pct']:.1f}%")
                    elif rem['merma_pct'] <= 12:
                        col_r4.metric("Merma %", f"🟡 {rem['merma_pct']:.1f}%")
                    else:
                        col_r4.metric("Merma %", f"🔴 {rem['merma_pct']:.1f}%")

                    if rem.get('cortes_salida'):
                        df_rem = pd.DataFrame(rem['cortes_salida'])
                        cols = [c for c in ['corte', 'grupo', 'kg', 'piezas'] if c in df_rem.columns]
                        st.dataframe(df_rem[cols], use_container_width=True, hide_index=True)

# ── TAB 2: ANÁLISIS ──
with tab_analisis:
    if not st.session_state.parsed_files:
        st.info("👆 Primero cargá y parseá los archivos en la pestaña anterior.")
    else:
        all_valid = [p for p in st.session_state.parsed_files if 'error' not in p and not p.get('excluido')]

        if len(all_valid) == 0:
            st.error("No hay archivos válidos para analizar.")
        else:
            st.markdown(f"### 📊 Análisis — Perfil: **{calidad}**")

            # ── Filtros de selección ──
            organizado_analisis = organizar_por_mes_semana(all_valid)
            meses_disp = list(organizado_analisis.keys())
            filtro_col1, filtro_col2 = st.columns(2)

            sel_mes = filtro_col1.selectbox(
                "Mes", ['Todos'] + meses_disp, key='filtro_mes_analisis')

            semanas_disp = []
            if sel_mes != 'Todos' and sel_mes in organizado_analisis:
                semanas_disp = list(organizado_analisis[sel_mes].keys())
            sel_semana = filtro_col2.selectbox(
                "Semana", ['Todas'] + semanas_disp, key='filtro_sem_analisis')

            # Filtrar
            if sel_mes == 'Todos':
                valid_files = all_valid
            elif sel_semana == 'Todas':
                valid_files = []
                for sem_archs in organizado_analisis[sel_mes].values():
                    valid_files.extend(sem_archs)
            else:
                valid_files = organizado_analisis.get(sel_mes, {}).get(sel_semana, [])

            # Selección individual de archivos
            archivos_nombres = [p.get('archivo', '?') for p in valid_files]
            sel_archivos = st.multiselect(
                "Archivos (dejar vacío = todos)",
                archivos_nombres, key='filtro_arch_analisis')
            if sel_archivos:
                valid_files = [p for p in valid_files if p.get('archivo') in sel_archivos]

            st.markdown(f"**{len(valid_files)}** archivo(s) seleccionado(s) de {len(all_valid)} totales.")

            if len(valid_files) == 0:
                st.warning("No hay archivos en el filtro seleccionado.")
            elif len(valid_files) == 1:
                modo = 'Individual'
                st.info("📄 Un solo archivo → Análisis individual")
            else:
                modo = st.radio(
                    "**Modo de análisis**",
                    ['Individual (por archivo)', 'Acumulado (todos juntos)', 'Ambos'],
                    horizontal=True,
                )

            col_btn1, col_btn2 = st.columns(2)
            run_excel = col_btn1.button("🚀 Generar análisis (Excel comercial)",
                                          type="primary", use_container_width=True)
            run_html_prod = col_btn2.button("📄 Generar reportes HTML producción",
                                              use_container_width=True,
                                              help=("Reporte para gerente de planta — sin info comercial. "
                                                    "Incluye rendimiento, cortes faltantes, calidad de "
                                                    "despostada, mermas y sugerencias."))

            if run_excel:
                output_dir = tempfile.mkdtemp(prefix='romaneo_')
                generated_files = []

                with st.spinner("Generando análisis..."):
                    # Individual
                    if modo in ['Individual', 'Individual (por archivo)', 'Ambos']:
                        for p in valid_files:
                            precio_este = p.get('precio_compra', precio_compra)
                            fecha_rom = p.get('fecha', '')

                            # Precios según fecha del romaneo
                            pm, pm_live, pm_status, pm_mod = cargar_price_matrix(fecha_rom)
                            p['precios_status'] = pm_status
                            p['precios_live'] = pm_live

                            data = {
                                'romaneo': {
                                    'numero': p.get('numero', ''),
                                    'fecha': fecha_rom,
                                    'medias_reses': p.get('medias_reses', 0),
                                    'kg_entrada': p.get('kg_entrada', 0),
                                    'categoria': p.get('categoria', 'Vaca'),
                                    'tipificacion': p.get('tipificacion', ''),
                                    'precio_compra': precio_este,
                                    'precio_mag': 0,
                                    'desglose_categoria': p.get('desglose_categoria', {}),
                                    'desglose_tipificacion': p.get('desglose_tipificacion', {}),
                                },
                                'cortes': p.get('cortes', []),
                            }
                            fname = p.get('nombre_analisis', generar_nombre_archivo(p, calidad))
                            output_path = os.path.join(output_dir, f'{fname}_Analisis.xlsx')
                            try:
                                refs = build_analisis(
                                    data, output_path,
                                    calidad=calidad,
                                    costos_override=costos_edit,
                                    precio_amarilla=precio_amarilla,
                                    price_matrix=pm,
                                    extra_info={
                                        'tropas_match': p.get('tropas_match', []),
                                        'tropas_sin_match': p.get('tropas_sin_match', []),
                                        'dias_faena_produccion': p.get('dias_faena_produccion'),
                                    },
                                )
                                generated_files.append({
                                    'path': output_path,
                                    'name': f'{fname}_Analisis.xlsx',
                                    'tipo': 'Individual',
                                    'data': p,
                                    'refs': refs,
                                })
                            except Exception as e:
                                st.error(f"Error generando {fname}: {e}")

                    # Acumulado
                    if modo in ['Acumulado', 'Acumulado (todos juntos)', 'Ambos']:
                        acum = acumular_romaneos(valid_files)
                        precio_acum = acum.get('precio_compra', precio_compra)

                        # Precios: usar fecha más reciente del lote
                        fechas_acum = [p.get('fecha', '') for p in valid_files if p.get('fecha')]
                        fecha_acum = max(fechas_acum) if fechas_acum else ''
                        pm_acum, _, pm_acum_status, _ = cargar_price_matrix(fecha_acum)

                        data_acum = {
                            'romaneo': {
                                'numero': 'ACUMULADO',
                                'fecha': acum.get('fecha', ''),
                                'medias_reses': acum.get('medias_reses', 0),
                                'kg_entrada': acum.get('kg_entrada', 0),
                                'categoria': acum.get('categoria', 'Vaca'),
                                'tipificacion': '',
                                'precio_compra': precio_acum,
                                'precio_mag': 0,
                            },
                            'cortes': acum.get('cortes', []),
                        }
                        output_path = os.path.join(output_dir, f'ACUMULADO_{calidad}_Analisis.xlsx')
                        try:
                            refs = build_analisis(
                                data_acum, output_path,
                                calidad=calidad,
                                costos_override=costos_edit,
                                precio_amarilla=precio_amarilla,
                                price_matrix=pm_acum,
                            )
                            generated_files.append({
                                'path': output_path,
                                'name': f'ACUMULADO_{calidad}_Analisis.xlsx',
                                'tipo': 'Acumulado',
                                'data': acum,
                                'refs': refs,
                            })
                        except Exception as e:
                            st.error(f"Error generando acumulado: {e}")

                st.session_state.analysis_results = generated_files

                # Actualizar historial
                try:
                    historial_path = os.path.join(
                        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        'ROMANEOS', 'historial_romaneos.json'
                    )
                    historial = []
                    if os.path.exists(historial_path):
                        with open(historial_path, 'r') as f:
                            historial = json.load(f)

                    for p in valid_files:
                        meat_kg = sum(c['kg'] for c in p.get('cortes', [])
                                      if c.get('grupo') != 'GRASA')
                        kg_ent = p.get('kg_entrada', 0)
                        rend_val = (meat_kg / kg_ent * 100) if kg_ent > 0 else 0
                        cat = p.get('categoria', 'Vaca').split('(')[0].strip()
                        rend_obj_base = REND_OBJETIVO.get(calidad, {}).get(cat, 0.66)
                        if isinstance(rend_obj_base, float) and rend_obj_base < 1:
                            rend_obj_base *= 100

                        kg_am = sum(c['kg'] for c in p.get('cortes', [])
                                    if str(c.get('contramarca', '')) in AMARILLA_CONTRAMARCAS)
                        medias_val = p.get('medias_reses', 0)
                        peso_prom = round(kg_ent / medias_val) if medias_val > 0 else 0

                        # Usar el precio real de este romaneo (del cruce o manual)
                        precio_real = p.get('precio_compra', precio_compra)
                        costo_hacienda = round(kg_ent * precio_real) if kg_ent > 0 else 0

                        entry = {
                            'archivo': p.get('nombre_analisis', generar_nombre_archivo(p, calidad)),
                            'archivo_original': p.get('archivo', ''),
                            'fecha': p.get('fecha', ''),
                            'categoria': p.get('categoria', ''),
                            'calidad': calidad,
                            'medias': medias_val,
                            'peso_promedio_media': peso_prom,
                            'kg_entrada': kg_ent,
                            'kg_carne': round(meat_kg, 2),
                            'rendimiento_pct': round(rend_val, 2),
                            'rend_objetivo_pct': round(rend_obj_base, 1),
                            'rend_vs_obj': round(rend_val - rend_obj_base, 2),
                            'kg_amarilla': round(kg_am, 2),
                            'pct_amarilla': round(kg_am / meat_kg * 100, 1) if meat_kg > 0 else 0,
                            'precio_compra': precio_real,
                            'costo_hacienda': costo_hacienda,
                        }

                        ya_existe = any(h.get('archivo_original') == p.get('archivo', '') for h in historial)
                        if not ya_existe:
                            historial.append(entry)

                    os.makedirs(os.path.dirname(historial_path), exist_ok=True)
                    with open(historial_path, 'w') as f:
                        json.dump(historial, f, indent=2, ensure_ascii=False)
                except Exception:
                    pass

                st.rerun()

            # ── HTML PRODUCCIÓN (gerente de planta) ──
            if run_html_prod:
                from html_builder_produccion import build_html_produccion
                html_results = []

                # Cargar histórico para comparativas / mermas
                historial_data = []
                try:
                    hist_path = os.path.join(
                        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        'ROMANEOS', 'historial_romaneos.json')
                    if os.path.exists(hist_path):
                        with open(hist_path, 'r') as fh:
                            historial_data = json.load(fh)
                except Exception:
                    historial_data = []

                with st.spinner("Generando reportes HTML para producción..."):
                    # Individual
                    if modo in ['Individual', 'Individual (por archivo)', 'Ambos']:
                        for p in valid_files:
                            try:
                                data_prod = {
                                    'archivo': p.get('archivo', 'Romaneo'),
                                    'fecha': p.get('fecha', ''),
                                    'numero': p.get('numero', ''),
                                    'medias_reses': p.get('medias_reses', 0),
                                    'kg_entrada': p.get('kg_entrada', 0),
                                    'kg_carne': p.get('kg_carne', 0),
                                    'categoria': p.get('categoria', 'Vaca'),
                                    'tipificacion': p.get('tipificacion', ''),
                                    'precio_compra': p.get('precio_compra', precio_compra),
                                    'pct_amarilla': p.get('pct_amarilla', 0),
                                    'dias_faena_produccion': p.get('dias_faena_produccion'),
                                    'tropas_match': p.get('tropas_match', []),
                                    'tropas_sin_match': p.get('tropas_sin_match', []),
                                    'cortes': p.get('cortes', []),
                                    'merma_kg': p.get('merma_kg', 0),
                                    'grasa_kg': p.get('grasa_kg', 0),
                                }
                                html_str = build_html_produccion(
                                    data_prod, calidad=calidad,
                                    historial=historial_data,
                                )
                                fname = p.get('nombre_analisis',
                                                generar_nombre_archivo(p, calidad))
                                html_results.append({
                                    'name': f'{fname}_PRODUCCION.html',
                                    'tipo': 'Individual',
                                    'html': html_str,
                                    'data': p,
                                })
                            except Exception as e:
                                st.error(f"Error generando HTML de {p.get('archivo')}: {e}")

                    # Acumulado
                    if modo in ['Acumulado', 'Acumulado (todos juntos)', 'Ambos']:
                        try:
                            acum = acumular_romaneos(valid_files)
                            data_acum_prod = {
                                'archivo': f'ACUMULADO_{calidad}',
                                'fecha': acum.get('fecha', ''),
                                'numero': 'ACUMULADO',
                                'medias_reses': acum.get('medias_reses', 0),
                                'kg_entrada': acum.get('kg_entrada', 0),
                                'kg_carne': acum.get('kg_carne', 0),
                                'categoria': acum.get('categoria', 'Vaca'),
                                'precio_compra': acum.get('precio_compra', precio_compra),
                                'pct_amarilla': acum.get('pct_amarilla', 0),
                                'cortes': acum.get('cortes', []),
                                'tropas_match': acum.get('tropas_match', []),
                            }
                            html_str = build_html_produccion(
                                data_acum_prod, calidad=calidad,
                                historial=historial_data,
                                titulo_extra='ACUMULADO',
                            )
                            html_results.append({
                                'name': f'ACUMULADO_{calidad}_PRODUCCION.html',
                                'tipo': 'Acumulado',
                                'html': html_str,
                                'data': acum,
                            })
                        except Exception as e:
                            st.error(f"Error generando HTML acumulado: {e}")

                st.session_state.html_produccion_results = html_results
                if html_results:
                    st.success(f"✅ {len(html_results)} reporte(s) HTML generado(s)")

            # ── Mostrar HTMLs producción si existen ──
            if st.session_state.get('html_produccion_results'):
                st.markdown("---")
                st.markdown("### 📄 Reportes HTML — Gerente de Producción")
                st.caption("Sin información comercial. Foco en rendimiento, calidad y mermas.")

                for hres in st.session_state.html_produccion_results:
                    with st.expander(
                        f"{'📊' if hres['tipo']=='Acumulado' else '📄'} {hres['name']}",
                        expanded=True,
                    ):
                        d = hres['data']
                        kg_e = d.get('kg_entrada', 0) or 0
                        meat = [c for c in d.get('cortes', []) if c.get('grupo') != 'GRASA']
                        kg_c = d.get('kg_carne', 0) or sum(c['kg'] for c in meat)
                        rend = (kg_c / kg_e * 100) if kg_e else 0
                        n_cortes = len(meat)

                        col_h1, col_h2, col_h3, col_h4 = st.columns(4)
                        col_h1.metric("Kg entrada", f"{kg_e:,.0f}".replace(',', '.'))
                        col_h2.metric("Kg carne", f"{kg_c:,.0f}".replace(',', '.'))
                        col_h3.metric("Rendimiento", f"{rend:.2f}%".replace('.', ','))
                        col_h4.metric("Líneas de corte", n_cortes)

                        st.download_button(
                            "⬇️ Descargar HTML producción",
                            hres['html'].encode('utf-8'),
                            file_name=hres['name'],
                            mime='text/html',
                            use_container_width=True,
                            key=f"dl_htmlprod_{hres['name']}_{id(hres)}",
                        )

            # Mostrar resultados
            if st.session_state.analysis_results:
                st.markdown("---")
                st.markdown("### 📥 Resultados generados")

                for res in st.session_state.analysis_results:
                    with st.expander(
                        f"{'📊' if res['tipo'] == 'Acumulado' else '📄'} "
                        f"{res['name']} ({res['tipo']})", expanded=True
                    ):
                        data = res['data']
                        kg_entrada = data.get('kg_entrada', 0)
                        meat_cortes = [c for c in data.get('cortes', []) if c.get('grupo') != 'GRASA']
                        kg_carne = sum(c['kg'] for c in meat_cortes)
                        rend = (kg_carne / kg_entrada * 100) if kg_entrada > 0 else 0
                        medias = data.get('medias_reses', 0)

                        col1, col2, col3, col4, col5 = st.columns(5)
                        col1.metric("Medias", f"{medias:,}")
                        col2.metric("Kg entrada", f"{kg_entrada:,.0f}")
                        col3.metric("Kg carne", f"{kg_carne:,.0f}")
                        col4.metric("Rendimiento", f"{rend:.1f}%")

                        rend_obj = res['refs'].get('rend_obj', 0.66) * 100
                        delta_rend = rend - rend_obj
                        col5.metric("vs Objetivo", f"{delta_rend:+.1f}%",
                                    delta=f"{delta_rend:+.1f}%", delta_color="normal")

                        # Info de precios usados
                        data_rom = res.get('data', {})
                        precios_status = data_rom.get('precios_status', '')
                        if precios_status:
                            if data_rom.get('precios_live'):
                                st.caption(f"📋 {precios_status}")
                            else:
                                st.caption(f"📋 {precios_status}")
                        precio_usado = data_rom.get('precio_compra', 0)
                        if not precio_usado and 'romaneo' in data:
                            precio_usado = data.get('precio_compra', 0)

                        # Desglose de costos por romaneo (solo en acumulado)
                        if res['tipo'] == 'Acumulado' and data_rom.get('detalle_compras'):
                            st.markdown("#### 💲 Desglose de costos de hacienda")
                            df_costos = pd.DataFrame(data_rom['detalle_compras'])
                            df_costos.columns = ['Archivo', 'Kg entrada', '$/kg', 'Costo total', 'Categoría', 'Medias']
                            df_costos['$/kg'] = df_costos['$/kg'].apply(lambda x: f"${x:,.0f}")
                            df_costos['Costo total'] = df_costos['Costo total'].apply(lambda x: f"${x:,.0f}")
                            df_costos['Kg entrada'] = df_costos['Kg entrada'].apply(lambda x: f"{x:,.0f}")
                            st.dataframe(df_costos, use_container_width=True, hide_index=True)

                            costo_total = data_rom.get('costo_hacienda_total', 0)
                            precio_prom = data_rom.get('precio_compra', 0)
                            col_c1, col_c2 = st.columns(2)
                            col_c1.metric("Costo total hacienda", f"${costo_total:,.0f}")
                            col_c2.metric("Precio promedio ponderado", f"${precio_prom:,.0f}/kg")

                            # Promedio días faena→producción en acumulado
                            valid_parsed = [pp for pp in st.session_state.parsed_files
                                            if 'error' not in pp and pp.get('dias_faena_produccion') is not None]
                            if valid_parsed:
                                dias_prom = sum(pp['dias_faena_produccion'] for pp in valid_parsed) / len(valid_parsed)
                                if dias_prom < 3:
                                    st.markdown(f"🟢 **Promedio faena → producción: {dias_prom:.1f} días**")
                                elif dias_prom < 5:
                                    st.markdown(f"🟡 **Promedio faena → producción: {dias_prom:.1f} días**")
                                else:
                                    st.markdown(f"🔴 **Promedio faena → producción: {dias_prom:.1f} días** — revisar tiempos")

                        elif res['tipo'] == 'Individual' and precio_usado:
                            st.markdown(f"**💲 Precio de compra:** ${precio_usado:,.0f}/kg")

                        # Amarillas
                        kg_am = sum(c['kg'] for c in data.get('cortes', [])
                                    if str(c.get('contramarca', '')) in AMARILLA_CONTRAMARCAS)
                        if kg_am > 0:
                            pct_am = kg_am / kg_carne * 100 if kg_carne > 0 else 0
                            st.markdown(
                                f'<div class="alert-amarilla">'
                                f'⚠️ <b>Amarillas:</b> {kg_am:,.0f} kg ({pct_am:.1f}% de carne) '
                                f'— Precio fijo ${precio_amarilla:,.0f}/kg</div>',
                                unsafe_allow_html=True
                            )

                        col_dl1, col_dl2 = st.columns(2)
                        if os.path.exists(res['path']):
                            with open(res['path'], 'rb') as f:
                                col_dl1.download_button(
                                    f"⬇️ Excel completo (Comercial)",
                                    f.read(),
                                    file_name=res['name'],
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    key=f"dl_{res['name']}_{id(res)}",
                                )

                        # ── HTML para gerente de producción ──
                        try:
                            from html_builder_produccion import build_html_produccion
                            # Cargar histórico para comparativas / mermas
                            historial_data = []
                            try:
                                hist_path = os.path.join(
                                    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                    'ROMANEOS', 'historial_romaneos.json')
                                if os.path.exists(hist_path):
                                    with open(hist_path, 'r') as fh:
                                        historial_data = json.load(fh)
                            except Exception:
                                historial_data = []

                            data_html = res['data']
                            # Asegurar que el dict tenga 'fecha' y 'archivo' al nivel raíz
                            if 'fecha' not in data_html and 'romaneo' in data_html:
                                data_html = {**data_html,
                                              'fecha': data_html['romaneo'].get('fecha', ''),
                                              'archivo': res['name']}
                            elif 'archivo' not in data_html:
                                data_html = {**data_html, 'archivo': res['name']}

                            titulo_extra = 'ACUMULADO' if res['tipo'] == 'Acumulado' else None
                            html_str = build_html_produccion(
                                data_html, calidad=calidad,
                                historial=historial_data,
                                titulo_extra=titulo_extra,
                            )
                            html_name = res['name'].replace('.xlsx', '_PRODUCCION.html')
                            col_dl2.download_button(
                                "📄 HTML Producción (Gerente Planta)",
                                html_str.encode('utf-8'),
                                file_name=html_name,
                                mime='text/html',
                                use_container_width=True,
                                key=f"dl_html_{res['name']}_{id(res)}",
                                help=("Reporte para gerente de producción: rendimiento, cortes "
                                      "faltantes, calidad de despostada, mermas y sugerencias. "
                                      "Sin info comercial.")
                            )
                        except Exception as e:
                            col_dl2.error(f"Error generando HTML: {e}")

# ── TAB 3: CONTROL DE CORTES ──
with tab_control:
    if not st.session_state.parsed_files:
        st.info("👆 Primero cargá y parseá los archivos.")
    else:
        all_files_ctrl = [p for p in st.session_state.parsed_files
                          if 'error' not in p and not p.get('excluido')]

        st.markdown("### 🔍 Control de piezas por media res")
        st.markdown(
            "Compara las **piezas reales** que salieron vs las **esperadas** según la cantidad de medias. "
            "Para cortes porcionados/feteados muestra el **ratio por media** vs el esperado."
        )

        # ── Filtros: mes, romaneo, categoría ──
        org_ctrl = organizar_por_mes_semana(all_files_ctrl)
        meses_ctrl = list(org_ctrl.keys())
        cats_ctrl = sorted({p.get('categoria', '?') for p in all_files_ctrl if p.get('categoria')})

        fcol1, fcol2, fcol3 = st.columns(3)
        sel_mes_ctrl = fcol1.selectbox("Mes", ['Todos'] + meses_ctrl, key='ctrl_mes')
        sel_cat_ctrl = fcol2.selectbox("Categoría", ['Todas'] + cats_ctrl, key='ctrl_cat')

        # Filtrar por mes y categoría primero
        filtrados = all_files_ctrl
        if sel_mes_ctrl != 'Todos':
            filtrados = []
            for sa in org_ctrl.get(sel_mes_ctrl, {}).values():
                filtrados.extend(sa)
        if sel_cat_ctrl != 'Todas':
            filtrados = [p for p in filtrados if p.get('categoria') == sel_cat_ctrl]

        # Filtro por romaneos específicos
        nombres_ctrl = [p.get('archivo', '?') for p in filtrados]
        sel_archs_ctrl = fcol3.multiselect(
            "Romaneos (vacío = todos)", nombres_ctrl, key='ctrl_archs')
        if sel_archs_ctrl:
            filtrados = [p for p in filtrados if p.get('archivo') in sel_archs_ctrl]

        st.caption(f"Mostrando {len(filtrados)} romaneo(s) de {len(all_files_ctrl)} totales.")

        if not filtrados:
            st.warning("No hay romaneos con los filtros seleccionados.")
            valid_files = []
        else:
            valid_files = filtrados

        for p in valid_files:
            nombre = p.get('nombre_analisis', p.get('archivo', 'S/N'))
            medias = p.get('medias_reses', 0)

            with st.expander(f"📄 {nombre} — {medias} medias", expanded=True):
                ctrl = control_cortes(p)

                if not ctrl:
                    st.warning("No se pudo generar el control (sin medias o sin cortes)")
                    continue

                # Separar por tipo
                anat = [c for c in ctrl if c['tipo'] == 'ANATÓMICO']
                porc = [c for c in ctrl if c['tipo'] == 'PORCIONADO']
                fete = [c for c in ctrl if c['tipo'] == 'FETEADO']
                sinreg = [c for c in ctrl if c['tipo'] == 'SIN REGISTRO']

                def color_estado(val):
                    s = str(val)
                    if 'FALTAN' in s or 'SIN REGISTRO' in s or 'PESADO' in s or 'LIVIANO' in s:
                        return 'background-color: #FCE4EC; color: #C62828'
                    elif 'SOBRAN' in s:
                        return 'background-color: #FFF3E0; color: #E65100'
                    elif 'OK' in s:
                        return 'background-color: #E8F5E9; color: #2E7D32'
                    return ''

                if anat:
                    st.markdown("#### Cortes anatómicos")
                    st.caption("Piezas esperadas según el % de medias que fue anatómico")
                    df_anat = pd.DataFrame(anat)
                    cols_show = ['grupo', 'piezas_reales', 'piezas_esperadas', 'diferencia',
                                 'peso_promedio', 'alerta', 'detalle']
                    cols_avail = [c for c in cols_show if c in df_anat.columns]
                    df_anat = df_anat[cols_avail]
                    df_anat.columns = ['Corte', 'Reales', 'Esperadas', 'Dif',
                                       'Peso prom (kg)', 'Estado', 'Detalle'][:len(cols_avail)]
                    st.dataframe(
                        df_anat.style.applymap(color_estado, subset=['Estado']),
                        use_container_width=True, hide_index=True
                    )

                if porc:
                    st.markdown("#### Cortes porcionados")
                    st.caption("Piezas esperadas = (medias × %porc) × ratio. Peso ideal: 1.0-1.5 kg por pieza")
                    df_porc = pd.DataFrame(porc)
                    cols_show = ['grupo', 'piezas_reales', 'piezas_esperadas', 'diferencia',
                                 'peso_promedio', 'peso_nota', 'alerta']
                    cols_avail = [c for c in cols_show if c in df_porc.columns]
                    df_porc = df_porc[cols_avail]
                    df_porc.columns = ['Corte', 'Reales', 'Esperadas', 'Dif',
                                       'Peso prom (kg)', 'Peso eval', 'Estado'][:len(cols_avail)]
                    subset_cols = [c for c in ['Estado', 'Peso eval'] if c in df_porc.columns]
                    st.dataframe(
                        df_porc.style.applymap(color_estado, subset=subset_cols),
                        use_container_width=True, hide_index=True
                    )

                if fete:
                    st.markdown("#### Cortes feteados")
                    st.caption("Piezas esperadas = (medias × %fet) × ratio. Peso ideal: 1.0-1.5 kg por pieza")
                    df_fet = pd.DataFrame(fete)
                    cols_show = ['grupo', 'piezas_reales', 'piezas_esperadas', 'diferencia',
                                 'peso_promedio', 'peso_nota', 'alerta']
                    cols_avail = [c for c in cols_show if c in df_fet.columns]
                    df_fet = df_fet[cols_avail]
                    df_fet.columns = ['Corte', 'Reales', 'Esperadas', 'Dif',
                                      'Peso prom (kg)', 'Peso eval', 'Estado'][:len(cols_avail)]
                    subset_cols = [c for c in ['Estado', 'Peso eval'] if c in df_fet.columns]
                    st.dataframe(
                        df_fet.style.applymap(color_estado, subset=subset_cols),
                        use_container_width=True, hide_index=True
                    )

                if sinreg:
                    st.markdown("#### Cortes sin registro")
                    df_sr = pd.DataFrame(sinreg)
                    df_sr = df_sr[['grupo', 'piezas_esperadas', 'alerta']]
                    df_sr.columns = ['Corte', 'Esperadas', 'Estado']
                    st.dataframe(
                        df_sr.style.applymap(color_estado, subset=['Estado']),
                        use_container_width=True, hide_index=True
                    )

                # Resumen
                total_alertas = sum(1 for c in ctrl
                                    if 'FALTAN' in c.get('alerta', '')
                                    or 'BAJO' in c.get('alerta', '')
                                    or 'SIN REGISTRO' in c.get('alerta', ''))
                if total_alertas == 0:
                    st.success(f"✅ Sin alertas — todos los cortes dentro de rango para {medias} medias")
                else:
                    st.warning(f"⚠️ {total_alertas} alerta(s) detectada(s) — revisar faltantes")

# ── TAB 4: REPORTE COMPRADOR ──
with tab_comprador:
    st.markdown("### 🛒 Reporte para el comprador")

    all_comprador = [p for p in st.session_state.parsed_files if 'error' not in p and not p.get('excluido')]

    if not all_comprador:
        st.info("👆 Cargá y parseá romaneos para generar el reporte.")
    else:
        # Filtros
        org_comp = organizar_por_mes_semana(all_comprador)
        meses_comp = list(org_comp.keys())
        fc1, fc2 = st.columns(2)
        sel_mes_c = fc1.selectbox("Mes", ['Todos'] + meses_comp, key='filtro_mes_comp')
        sem_comp = list(org_comp.get(sel_mes_c, {}).keys()) if sel_mes_c != 'Todos' else []
        sel_sem_c = fc2.selectbox("Semana", ['Todas'] + sem_comp, key='filtro_sem_comp')

        if sel_mes_c == 'Todos':
            valid_comprador = all_comprador
        elif sel_sem_c == 'Todas':
            valid_comprador = []
            for sa in org_comp[sel_mes_c].values():
                valid_comprador.extend(sa)
        else:
            valid_comprador = org_comp.get(sel_mes_c, {}).get(sel_sem_c, [])

        st.markdown(f"**{len(valid_comprador)}** romaneo(s) seleccionados.")

    if not all_comprador or not valid_comprador:
        pass
    else:
        from config import (CORTES_CAROS, COSTOS_PERFILES, REND_OBJETIVO,
                            AMARILLA_CONTRAMARCAS, PRECIOS_FIJOS_RECORTE)
        from collections import defaultdict

        costos_std = COSTOS_PERFILES['Standard']
        costo_var_kg = costos_std['mo'] + costos_std['insumos'] + costos_std['flete'] + costos_std['senasa']
        iibb = costos_std['iibb']

        # Cargar precios para calcular ingresos a nivel corte
        pm_rep, _, _, _ = cargar_price_matrix('30/03/2026')

        # ── Calcular CM detallada por romaneo ──
        reportes = []
        for p in valid_comprador:
            cortes = p.get('cortes', [])
            meat = [c for c in cortes if c.get('grupo') not in ('GRASA', 'SIN CLASIFICAR')]
            kg_entrada = p.get('kg_entrada', 0)
            kg_carne = sum(c['kg'] for c in meat)
            precio_compra = p.get('precio_compra', 6800)

            # Ingreso por corte
            ingreso_total = 0
            ingreso_caros = 0
            kg_caros = 0
            kg_picada = 0
            kg_amarilla = 0
            cortes_detalle = defaultdict(lambda: {'kg': 0, 'ingreso': 0})

            for c in meat:
                grupo = c['grupo']
                kg = c['kg']
                cm_str = str(c.get('contramarca', ''))

                # Resolver precio
                if grupo in PRECIOS_FIJOS_RECORTE:
                    precio = PRECIOS_FIJOS_RECORTE[grupo]
                elif cm_str in AMARILLA_CONTRAMARCAS:
                    precio = 9500
                elif c.get('es_bubalino'):
                    precio = precio_bubalino if 'precio_bubalino' in dir() else 10500
                else:
                    from config import CONTRAMARCA_MAP
                    _, pricing_col = CONTRAMARCA_MAP.get(cm_str, ('', 'RESTO CLIENTES AMBA'))
                    precio = 0
                    if grupo in pm_rep and pricing_col in pm_rep[grupo]:
                        precio = pm_rep[grupo][pricing_col]
                    elif grupo in pm_rep:
                        for fb in ['NETOS PEYA', 'RESTO CLIENTES AMBA']:
                            if fb in pm_rep[grupo]:
                                precio = pm_rep[grupo][fb]; break

                ingreso_corte = kg * precio
                ingreso_total += ingreso_corte
                cortes_detalle[grupo]['kg'] += kg
                cortes_detalle[grupo]['ingreso'] += ingreso_corte

                if grupo in CORTES_CAROS:
                    ingreso_caros += ingreso_corte
                    kg_caros += kg
                if grupo == 'CARNE PICADA':
                    kg_picada += kg
                if cm_str in AMARILLA_CONTRAMARCAS:
                    kg_amarilla += kg

            # P&L
            costo_hacienda = kg_entrada * precio_compra
            costo_var = kg_carne * costo_var_kg
            costo_iibb = ingreso_total * iibb
            costo_total = costo_hacienda + costo_var + costo_iibb
            ingreso_neto = ingreso_total * (1 - iibb)
            cm = ingreso_neto - costo_hacienda - costo_var
            margen = (cm / ingreso_neto * 100) if ingreso_neto > 0 else 0

            rend = (kg_carne / kg_entrada * 100) if kg_entrada > 0 else 0
            pct_caros = (kg_caros / kg_carne * 100) if kg_carne > 0 else 0
            pct_picada = (kg_picada / kg_carne * 100) if kg_carne > 0 else 0
            pct_amarilla = (kg_amarilla / kg_carne * 100) if kg_carne > 0 else 0
            pvp = (ingreso_total / kg_carne) if kg_carne > 0 else 0
            pvp_caros = (ingreso_caros / kg_caros) if kg_caros > 0 else 0

            reportes.append({
                'archivo': p.get('archivo', '?'),
                'nombre': p.get('nombre_analisis', p.get('archivo', '?')),
                'categoria': p.get('categoria', '?'),
                'medias': p.get('medias_reses', 0),
                'kg_entrada': kg_entrada,
                'kg_carne': kg_carne,
                'rend': rend,
                'precio_compra': precio_compra,
                'ingreso_total': ingreso_total,
                'costo_total': costo_total,
                'cm': cm,
                'margen': margen,
                'pvp': pvp,
                'pvp_caros': pvp_caros,
                'pct_caros': pct_caros,
                'pct_picada': pct_picada,
                'pct_amarilla': pct_amarilla,
                'kg_amarilla': kg_amarilla,
                'cortes_detalle': dict(cortes_detalle),
            })

        reportes.sort(key=lambda x: x['cm'], reverse=True)
        mejor = reportes[0]
        peor = reportes[-1]

        # ══════ MEJOR Y PEOR COMPRA ══════
        st.markdown("#### 🏆 Mejor compra vs 🔴 Peor compra (por Contribución Marginal)")

        col_m, col_p = st.columns(2)
        with col_m:
            st.success(f"**🏆 {mejor['nombre']}**")
            st.metric("CM", f"${mejor['cm']:,.0f}", delta=f"{mejor['margen']:.1f}%")
            st.markdown(f"""
            | | Valor |
            |---|---|
            | Categoría | **{mejor['categoria']}** |
            | Medias | {mejor['medias']} |
            | Rendimiento | **{mejor['rend']:.1f}%** |
            | Precio compra | ${mejor['precio_compra']:,.0f}/kg |
            | Cortes caros | **{mejor['pct_caros']:.1f}%** del total |
            | Carne picada | {mejor['pct_picada']:.1f}% |
            | Amarilla | **{mejor['pct_amarilla']:.1f}%** |
            | Precio venta prom. | ${mejor['pvp']:,.0f}/kg |
            """)

        with col_p:
            st.error(f"**🔴 {peor['nombre']}**")
            st.metric("CM", f"${peor['cm']:,.0f}", delta=f"{peor['margen']:.1f}%")
            st.markdown(f"""
            | | Valor |
            |---|---|
            | Categoría | **{peor['categoria']}** |
            | Medias | {peor['medias']} |
            | Rendimiento | **{peor['rend']:.1f}%** |
            | Precio compra | ${peor['precio_compra']:,.0f}/kg |
            | Cortes caros | **{peor['pct_caros']:.1f}%** del total |
            | Carne picada | {peor['pct_picada']:.1f}% |
            | Amarilla | **{peor['pct_amarilla']:.1f}%** |
            | Precio venta prom. | ${peor['pvp']:,.0f}/kg |
            """)

        # ── Explicación de la diferencia ──
        st.markdown("---")
        st.markdown("#### ¿Por qué fue mejor/peor?")

        razones_mejor = []
        razones_peor = []

        # Rendimiento
        diff_rend = mejor['rend'] - peor['rend']
        if abs(diff_rend) > 1:
            razones_mejor.append(f"Rindió **{diff_rend:+.1f}% más** de carne ({mejor['rend']:.1f}% vs {peor['rend']:.1f}%)")

        # Cortes caros
        diff_caros = mejor['pct_caros'] - peor['pct_caros']
        if abs(diff_caros) > 2:
            razones_mejor.append(f"**{diff_caros:+.1f}% más cortes caros** ({mejor['pct_caros']:.1f}% vs {peor['pct_caros']:.1f}%)")

        # Amarilla
        diff_am = peor['pct_amarilla'] - mejor['pct_amarilla']
        if diff_am > 3:
            razones_peor.append(f"La peor tuvo **{diff_am:.1f}% más amarilla** ({peor['pct_amarilla']:.1f}% vs {mejor['pct_amarilla']:.1f}%) — vende a $10.500/kg en vez de ~${mejor['pvp']:,.0f}")

        # Picada
        diff_picada = peor['pct_picada'] - mejor['pct_picada']
        if diff_picada > 3:
            razones_peor.append(f"La peor tuvo **{diff_picada:.1f}% más picada** (corte más barato del mix)")

        # Precio compra
        diff_precio = peor['precio_compra'] - mejor['precio_compra']
        if abs(diff_precio) > 100:
            if diff_precio > 0:
                razones_peor.append(f"Se pagó **${diff_precio:,.0f}/kg más** por la hacienda")
            else:
                razones_mejor.append(f"Se pagó **${abs(diff_precio):,.0f}/kg menos** por la hacienda")

        # PVP
        diff_pvp = mejor['pvp'] - peor['pvp']
        if abs(diff_pvp) > 500:
            razones_mejor.append(f"Precio venta promedio **${diff_pvp:,.0f}/kg mayor** (mejores clientes o mix de cortes)")

        if razones_mejor:
            st.markdown("**La mejor compra ganó porque:**")
            for r in razones_mejor:
                st.markdown(f"- ✅ {r}")
        if razones_peor:
            st.markdown("**La peor compra perdió porque:**")
            for r in razones_peor:
                st.markdown(f"- 🔴 {r}")

        st.markdown("---")

        # ══════ RANKING COMPLETO ══════
        # Calcular scoring para cada romaneo
        from config import REND_OBJETIVO as ROBJ
        reportes_con_score = []
        for r in reportes:
            cat = r['categoria']
            rend_obj = ROBJ.get('Standard', {}).get(cat, 0.66) * 100
            sc = calcular_score(r, rend_objetivo=rend_obj)
            r['puntaje'] = sc['puntaje']
            reportes_con_score.append({'reporte': r, 'score': sc})

        reportes.sort(key=lambda x: x['puntaje'], reverse=True)

        st.markdown("#### Ranking de todas las compras")
        df_rank = pd.DataFrame([{
            'Romaneo': r['nombre'],
            'Puntaje': r['puntaje'],
            'Cat.': r['categoria'],
            'Medias': r['medias'],
            'Rend %': round(r['rend'], 1),
            '$/kg compra': r['precio_compra'],
            'Caros %': round(r['pct_caros'], 1),
            'Amarilla %': round(r['pct_amarilla'], 1),
            'PVP $/kg': round(r['pvp']),
            'Margen %': round(r['margen'], 1),
        } for r in reportes])

        def color_puntaje(val):
            if val >= 7: return 'background-color: #E8F5E9; color: #2E7D32; font-weight: bold'
            elif val >= 5: return 'background-color: #FFF3E0; color: #E65100; font-weight: bold'
            else: return 'background-color: #FCE4EC; color: #C62828; font-weight: bold'

        st.dataframe(
            df_rank.style.applymap(color_puntaje, subset=['Puntaje']),
            use_container_width=True, hide_index=True
        )

        # ── Botón descargar PDF ──
        st.markdown("---")
        st.markdown("#### 📄 Descargar reporte para el comprador (PDF)")
        st.markdown("_Sin datos de facturación. Solo precio compra, puntaje y fundamentos._")

        reportes_con_score_sorted = sorted(
            reportes_con_score, key=lambda x: x['score']['puntaje'], reverse=True)

        pdf_bytes = bytes(generar_pdf_comprador(reportes_con_score_sorted))
        fecha_hoy = datetime.now().strftime('%Y%m%d')
        st.download_button(
            "⬇️ Descargar PDF Reporte Comprador",
            pdf_bytes,
            file_name=f"Reporte_Comprador_{fecha_hoy}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

        st.markdown("---")

        # ══════ VALORES DE EQUILIBRIO PARA LA PRÓXIMA SEMANA ══════
        st.markdown("#### 📊 Valores de equilibrio para la próxima compra")
        st.markdown(
            "_Basado en los **peores resultados observados** (pesimista). "
            "Muestra el precio máximo de compra para lograr 8-12% CM. "
            "Categoría superior = mejor calidad de carne (terneza), pero hay que pagarla._"
        )

        # Usar mediana (conservador pero alcanzable) para rendimiento y PVP
        # P60 para amarilla (no P75 que agarra outliers extremos)
        rend_obj_std = REND_OBJETIVO.get('Standard', {})
        categorias_ref = ['Vaca', 'Vaquillona', 'Novillito', 'Novillo']

        # Historial para tener más datos
        historial_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'ROMANEOS', 'historial_romaneos.json'
        )
        df_hist_comp = pd.DataFrame()
        if os.path.exists(historial_path):
            with open(historial_path, 'r') as f:
                hist = json.load(f)
            if hist:
                df_hist_comp = pd.DataFrame(hist)
                for col in ['rendimiento_pct', 'pct_amarilla', 'precio_compra']:
                    if col in df_hist_comp.columns:
                        df_hist_comp[col] = pd.to_numeric(df_hist_comp[col], errors='coerce')

        # Combinar datos cargados + historial
        datos_por_cat = defaultdict(lambda: {'rend': [], 'amarilla': [], 'pvp': [], 'pct_caros': []})

        for r in reportes:
            cat = r['categoria']
            datos_por_cat[cat]['rend'].append(r['rend'])
            datos_por_cat[cat]['amarilla'].append(r['pct_amarilla'])
            datos_por_cat[cat]['pvp'].append(r['pvp'])
            datos_por_cat[cat]['pct_caros'].append(r['pct_caros'])

        if not df_hist_comp.empty and 'categoria' in df_hist_comp.columns:
            for _, row in df_hist_comp.iterrows():
                cat = row.get('categoria', '')
                if cat and 'rendimiento_pct' in row and pd.notna(row['rendimiento_pct']):
                    datos_por_cat[cat]['rend'].append(row['rendimiento_pct'])
                if cat and 'pct_amarilla' in row and pd.notna(row['pct_amarilla']):
                    datos_por_cat[cat]['amarilla'].append(row['pct_amarilla'])
                if cat and 'precio_venta_promedio' in row and pd.notna(row.get('precio_venta_promedio')):
                    datos_por_cat[cat]['pvp'].append(row['precio_venta_promedio'])

        import numpy as np

        # Escenarios de amarilla por categoría
        # Vaca: 2 opciones (25% amarilla y <10%)
        # Novillito, Novillo, Vaquillona: solo <5%
        escenarios_amarilla = {
            'Vaca': [
                ('Vaca (amarilla <10%)', 8),
                ('Vaca (amarilla ~25%)', 25),
            ],
            'Vaquillona': [('Vaquillona (amarilla <5%)', 4)],
            'Novillito': [('Novillito (amarilla <5%)', 4)],
            'Novillo': [('Novillo (amarilla <5%)', 4)],
        }

        calidad_txt_map = {
            'Vaca': 'Básica',
            'Vaquillona': 'Buena (más tierna)',
            'Novillito': 'Muy buena',
            'Novillo': 'Premium (máx. terneza)',
        }

        equilibrio_rows = []
        for cat in categorias_ref:
            data = datos_por_cat.get(cat)
            if not data or not data['rend']:
                rend_conserv = rend_obj_std.get(cat, 0.66) * 100 - 1
                pvp_base = 13000
            else:
                rend_arr = np.array(data['rend'])
                pvp_arr = np.array(data['pvp']) if data['pvp'] else np.array([13000])
                rend_conserv = float(np.median(rend_arr))
                pvp_base = float(np.median(pvp_arr)) if len(pvp_arr) > 0 else 13000

            rend_dec = rend_conserv / 100
            obj = rend_obj_std.get(cat, 0.66) * 100
            calidad_txt = calidad_txt_map.get(cat, '')

            for escenario_label, amarilla_pct in escenarios_amarilla.get(cat, [(cat, 8)]):
                am_dec = amarilla_pct / 100

                # Ajustar PVP por amarilla: la parte amarilla vende a $9.500
                pvp_ajustado = pvp_base * (1 - am_dec) + 9500 * am_dec
                ingreso_neto_kg_ent = rend_dec * pvp_ajustado * (1 - iibb)
                costo_var_kg_ent = rend_dec * costo_var_kg

                for margen_target in [0.08, 0.10, 0.12]:
                    precio_max = ingreso_neto_kg_ent * (1 - margen_target) - costo_var_kg_ent

                    equilibrio_rows.append({
                        'Escenario': escenario_label,
                        'Calidad carne': calidad_txt,
                        'Rend. esperado': f"{rend_conserv:.1f}%",
                        'Amarilla': f"{amarilla_pct}%",
                        'PVP ajustado': f"${pvp_ajustado:,.0f}",
                        'CM objetivo': f"{margen_target*100:.0f}%",
                        'Precio máx compra': f"${precio_max:,.0f}",
                    })

        df_equil = pd.DataFrame(equilibrio_rows)
        # Pivot para mejor lectura
        df_pivot_eq = df_equil.pivot_table(
            index=['Escenario', 'Calidad carne', 'Rend. esperado', 'Amarilla', 'PVP ajustado'],
            columns='CM objetivo',
            values='Precio máx compra',
            aggfunc='first'
        )
        st.dataframe(df_pivot_eq, use_container_width=True)

        st.markdown("""
        **Cómo leer esta tabla:**
        - **Rend. esperado** = mediana de los rendimientos reales observados
        - **Amarilla** = escenario fijo por categoría (Vaca tiene 2 opciones, el resto <5%)
        - **PVP ajustado** = precio de venta promedio corregido por % de amarilla a $9.500/kg
        - **Precio máx compra** = lo máximo que podés pagar por kg de media res para lograr ese margen
        - Categoría superior = **mejor terneza y mejor precio de venta**, pero mayor costo de compra
        - La vaca con 25% amarilla necesita comprarse mucho más barata para dar el mismo margen
        """)

        st.markdown("---")

        # ══════ QUÉ BUSCAR EN CADA CATEGORÍA ══════
        st.markdown("#### 🎯 Qué buscar al comprar cada categoría")

        for cat in categorias_ref:
            data = datos_por_cat.get(cat)
            if not data or not data['rend']:
                continue

            rend_arr = np.array(data['rend'])
            am_arr = np.array(data['amarilla'])

            obj = rend_obj_std.get(cat, 0.66) * 100
            rend_prom = rend_arr.mean()
            rend_mejor = rend_arr.max()
            rend_peor = rend_arr.min()
            am_prom = am_arr.mean()

            calidad_txt = {
                'Vaca': '🐄 Calidad básica — precio más accesible, mayor volumen',
                'Vaquillona': '🐄 Buena calidad — más tierna que vaca, buen equilibrio precio/calidad',
                'Novillito': '🐂 Muy buena calidad — tierna, buen rendimiento',
                'Novillo': '🐂 Premium — máxima terneza, mejor precio de venta, mayor costo',
            }.get(cat, '')

            with st.expander(f"**{cat}** — {calidad_txt}", expanded=True):
                st.markdown(f"""
                | Indicador | Peor caso | Promedio | Mejor caso | Objetivo |
                |---|---|---|---|---|
                | Rendimiento | {rend_peor:.1f}% | {rend_prom:.1f}% | {rend_mejor:.1f}% | {obj:.0f}% |
                | Amarilla | {am_arr.max():.1f}% | {am_prom:.1f}% | {am_arr.min():.1f}% | <10% |
                """)

                recomendaciones = []
                if rend_prom < obj:
                    recomendaciones.append(f"⚠️ Rendimiento promedio ({rend_prom:.1f}%) **por debajo del objetivo** ({obj:.0f}%). Buscar tropas con mejor terminación.")
                else:
                    recomendaciones.append(f"✅ Rendimiento promedio ({rend_prom:.1f}%) cumple el objetivo ({obj:.0f}%).")

                if am_prom > 15:
                    recomendaciones.append(f"🔴 Amarilla promedio alta ({am_prom:.1f}%). Castiga fuerte el PVP. Buscar tropas con menos grasa amarilla.")
                elif am_prom > 8:
                    recomendaciones.append(f"🟡 Amarilla moderada ({am_prom:.1f}%). Controlar proveedores con alta incidencia.")
                else:
                    recomendaciones.append(f"✅ Amarilla controlada ({am_prom:.1f}%).")

                if data.get('pct_caros'):
                    caros_prom = np.mean(data['pct_caros'])
                    if caros_prom < 35:
                        recomendaciones.append(f"⚠️ Solo {caros_prom:.0f}% de kg van a cortes caros. Buscar tropas que den más bife, lomo, ojo de bife.")
                    else:
                        recomendaciones.append(f"✅ {caros_prom:.0f}% de kg en cortes caros — buen mix.")

                for rec in recomendaciones:
                    st.markdown(f"- {rec}")

        st.markdown("---")
        st.markdown(
            "_Este reporte usa escenarios **pesimistas** (percentil 25/75) para que las decisiones "
            "de compra no dependan de tener suerte. Siempre es preferible comprar categoría superior "
            "(novillo > novillito > vaquillona > vaca) si el precio de equilibrio lo permite, "
            "porque la terneza mejora el valor percibido y facilita la venta._"
        )

# ── TAB: PRICING ──
with tab_pricing:
    from pricing import (COSTOS_PRICING_DEFAULT, calcular_media_res, calcular_cuartos,
                         calcular_cortes, construir_lista_cortes,
                         aplicar_precios_a_romaneo, rendimiento_real_de_romaneos,
                         CORTES_POR_CUARTO, REND_CATEGORIA,
                         fmt_num, fmt_pct, fmt_dinero)
    import pandas as pd

    st.markdown("### 💰 Pricing — Modelo basado en Costeo MADRE")
    st.markdown("_Precio sugerido según modalidad de venta y rentabilidad objetivo._")

    # ── Selector de modalidad ARRIBA ──
    modo_pricing = st.radio(
        "**Modalidad**",
        ['🥩 Media res entera', '🔪 Por cuartos', '📋 Por cortes', '🍔 Hamburguesas'],
        horizontal=True, key='modo_pricing_top'
    )

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════
    # MODALIDADES CARNE: Media / Cuartos / Cortes (parámetros compartidos)
    # ══════════════════════════════════════════════════════════════════
    if modo_pricing != '🍔 Hamburguesas':
        # ── 1. Parámetros de la media res ──
        st.markdown("#### 1. Parámetros de la media res")
        col_a, col_b, col_c = st.columns(3)
        pr_peso = col_a.number_input("Peso media (kg)", value=COSTOS_PRICING_DEFAULT['peso_media'],
                                      step=5, key='pr_peso')
        pr_cat = col_b.selectbox("Categoría", list(REND_CATEGORIA.keys()), index=3, key='pr_cat')
        pr_rend_pct = col_c.number_input("Rendimiento (%)", value=REND_CATEGORIA[pr_cat]*100,
                                          step=0.5, format="%.2f", key='pr_rend_v2')
        pr_rend = pr_rend_pct / 100

        # Precio de compra: ARS s/IVA, ARS c/IVA (precio final), o USD × TC
        st.markdown("##### Precio de compra")
        moneda_compra = st.radio(
            "Modalidad",
            ['ARS s/IVA ($/kg)', 'ARS c/IVA ($/kg final)', 'USD ($/kg) + TC'],
            horizontal=True, key='pr_moneda_compra',
            label_visibility='collapsed',
            help=("Muchos proveedores pasan el precio FINAL c/IVA. "
                  "Elegí la modalidad correspondiente y el sistema calcula el neto.")
        )
        # IVA de la compra (común a todas las modalidades — algunos cortes/lotes
        # tienen IVA distinto al estándar)
        iva_compra_pct_default = 10.5
        if moneda_compra.startswith('ARS s/IVA'):
            col_p1, col_p2, col_p3 = st.columns([2, 1, 1])
            pr_compra = col_p1.number_input(
                "Precio compra ARS $/kg (s/IVA)",
                value=COSTOS_PRICING_DEFAULT['precio_compra_kg'],
                step=100, key='pr_compra'
            )
            iva_compra_pct = col_p2.number_input(
                "IVA compra (%)", value=iva_compra_pct_default,
                step=0.5, format="%.2f", key='pr_iva_compra',
                help="Sólo informativo en este modo (no se aplica al neto)."
            )
            tc_default = st.session_state.get('pr_tc', 1300)
            precio_c_iva = pr_compra * (1 + iva_compra_pct/100)
            col_p3.caption("c/IVA equivale:")
            col_p3.markdown(
                f"**${precio_c_iva:,.0f}/kg**  ·  USD {pr_compra/tc_default:.2f}/kg"
                .replace(',', '.')
            )
        elif moneda_compra.startswith('ARS c/IVA'):
            col_p1, col_p2, col_p3 = st.columns([2, 1, 1])
            pr_compra_c_iva = col_p1.number_input(
                "Precio FINAL ARS $/kg (c/IVA)",
                value=int(round(COSTOS_PRICING_DEFAULT['precio_compra_kg'] * 1.105)),
                step=100, key='pr_compra_c_iva',
                help="Precio que te pasa el proveedor incluyendo IVA."
            )
            iva_compra_pct = col_p2.number_input(
                "IVA compra (%)", value=iva_compra_pct_default,
                step=0.5, format="%.2f", key='pr_iva_compra',
                help="Se descuenta del precio final para obtener el neto s/IVA."
            )
            pr_compra = int(round(pr_compra_c_iva / (1 + iva_compra_pct/100)))
            col_p3.metric("Neto s/IVA $/kg", f"${pr_compra:,.0f}".replace(',', '.'))
        else:  # USD + TC
            col_p1, col_p2, col_p3 = st.columns(3)
            pr_compra_usd = col_p1.number_input(
                "Precio compra USD/kg",
                value=float(COSTOS_PRICING_DEFAULT['precio_compra_kg']) / 1300,
                step=0.10, format="%.2f", key='pr_compra_usd'
            )
            pr_tc = col_p2.number_input(
                "TC ARS/USD", value=1300, step=10, key='pr_tc',
                help="Tipo de cambio para convertir USD → ARS"
            )
            iva_compra_pct = iva_compra_pct_default
            pr_compra = int(round(pr_compra_usd * pr_tc))
            col_p3.metric("Neto s/IVA $/kg", f"${pr_compra:,.0f}".replace(',', '.'))

        # ── 2. Costos operativos ──
        st.markdown("#### 2. Costos operativos")
        st.caption("⚠️ MO, insumos, congelado y faena **solo aplican a venta por cortes**")
        col_e, col_f, col_g, col_h = st.columns(4)
        pr_mo = col_e.number_input("MO $/kg", value=COSTOS_PRICING_DEFAULT['mo_kg'], step=50, key='pr_mo')
        pr_flete = col_f.number_input("Flete $/kg", value=COSTOS_PRICING_DEFAULT['flete_kg'], step=10, key='pr_flete')
        pr_faena = col_g.number_input("Faena/Cuarteo $/media", value=COSTOS_PRICING_DEFAULT['faena_media'],
                                       step=100, key='pr_faena')
        pr_senasa = col_h.number_input("SENASA $/kg", value=COSTOS_PRICING_DEFAULT['senasa_kg'],
                                        step=0.1, format="%.2f", key='pr_senasa')

        col_i, col_j, col_k, col_l = st.columns(4)
        pr_insumos = col_i.number_input("Insumos $/kg", value=COSTOS_PRICING_DEFAULT['insumos_kg'], step=10, key='pr_insumos')
        pr_congelado = col_j.number_input("Congelado $/kg", value=COSTOS_PRICING_DEFAULT['congelado_kg'], step=10, key='pr_congelado')
        pr_tna_pct = col_k.number_input("TNA (%)", value=COSTOS_PRICING_DEFAULT['tna']*100,
                                         step=1.0, format="%.2f", key='pr_tna_v2')
        pr_tna = pr_tna_pct / 100
        pr_dias = col_l.number_input("Días financ.", value=COSTOS_PRICING_DEFAULT['dias_financiamiento'],
                                      step=1, key='pr_dias')

        # ── 3. Impuestos y mermas ──
        st.markdown("#### 3. Impuestos y mermas")
        col_m, col_n, col_o, col_p = st.columns(4)
        pr_iibb_pct = col_m.number_input("IIBB+Ganancias (%)",
                                          value=COSTOS_PRICING_DEFAULT['iibb_ganancias']*100,
                                          step=0.1, format="%.2f", key='pr_iibb_v2')
        pr_iibb = pr_iibb_pct / 100
        pr_cheque_pct = col_n.number_input("Imp. Cheque (%)",
                                            value=COSTOS_PRICING_DEFAULT['imp_cheque']*100,
                                            step=0.1, format="%.2f", key='pr_cheque_v2')
        pr_cheque = pr_cheque_pct / 100
        pr_merma_media_pct = col_o.number_input("Merma media (%)",
                                                 value=COSTOS_PRICING_DEFAULT['merma_media']*100,
                                                 step=0.25, format="%.2f", key='pr_merma_media_v2',
                                                 help="Merma adicional al vender media entera (2-3%)")
        pr_merma_media = pr_merma_media_pct / 100
        pr_merma_cuarteo_pct = col_p.number_input("Merma cuarteo (%)",
                                                   value=COSTOS_PRICING_DEFAULT['merma_cuarteo']*100,
                                                   step=0.25, format="%.2f", key='pr_merma_cuarteo_v2')
        pr_merma_cuarteo = pr_merma_cuarteo_pct / 100

        # ── 4. Rentabilidad objetivo ──
        st.markdown("#### 4. Rentabilidad objetivo")
        col_q, col_r, col_s = st.columns(3)
        pr_margen_media_pct = col_q.number_input("🥩 Margen media (%)",
                                                  value=COSTOS_PRICING_DEFAULT['margen_media']*100,
                                                  step=0.25, format="%.2f", key='pr_margen_media_v2',
                                                  help="Rango sugerido: 3-5%")
        pr_margen_media = pr_margen_media_pct / 100
        pr_margen_cuartos_pct = col_r.number_input("🔪 Margen cuartos (%)",
                                                    value=COSTOS_PRICING_DEFAULT['margen_cuartos']*100,
                                                    step=0.25, format="%.2f", key='pr_margen_cuartos_v2',
                                                    help="Rango sugerido: 4-6%")
        pr_margen_cuartos = pr_margen_cuartos_pct / 100
        pr_margen_cortes_pct = col_s.number_input("📋 Margen cortes (%)",
                                                   value=COSTOS_PRICING_DEFAULT['margen_cortes']*100,
                                                   step=0.5, format="%.2f", key='pr_margen_cortes_v2',
                                                   help="Rango sugerido: 8-12%")
        pr_margen_cortes = pr_margen_cortes_pct / 100

        # ── 5. Carne amarilla (solo aplica a venta por cortes) ──
        st.markdown("#### 5. Carne amarilla")
        st.caption("_Porción vendida a precio especial. Solo afecta venta por cortes._")
        col_am1, col_am2 = st.columns(2)
        pr_pct_amarilla_p = col_am1.number_input(
            "% carne amarilla",
            value=COSTOS_PRICING_DEFAULT['pct_amarilla'] * 100,
            step=0.5, format="%.2f", min_value=0.0, max_value=100.0,
            key='pr_pct_am_v2'
        )
        pr_pct_amarilla = pr_pct_amarilla_p / 100
        pr_precio_amarilla = col_am2.number_input(
            "Precio amarilla $/kg",
            value=COSTOS_PRICING_DEFAULT['precio_amarilla'],
            step=100, key='pr_precio_am_v2'
        )

        params = {
            'peso_media': pr_peso, 'rendimiento': pr_rend, 'precio_compra_kg': pr_compra,
            'mo_kg': pr_mo, 'flete_kg': pr_flete, 'faena_media': pr_faena,
            'senasa_kg': pr_senasa, 'insumos_kg': pr_insumos, 'congelado_kg': pr_congelado,
            'tna': pr_tna, 'dias_financiamiento': pr_dias,
            'iibb_ganancias': pr_iibb, 'imp_cheque': pr_cheque,
            'merma_media': pr_merma_media, 'merma_cuarteo': pr_merma_cuarteo,
            'margen_media': pr_margen_media, 'margen_cuartos': pr_margen_cuartos,
            'margen_cortes': pr_margen_cortes,
            'pct_amarilla': pr_pct_amarilla, 'precio_amarilla': pr_precio_amarilla,
        }

        st.markdown("---")

        # ── OUTPUT según modalidad ──
        if modo_pricing == '🥩 Media res entera':
            st.markdown("#### Venta de media res entera")
            st.caption(f"Solo costos: hacienda + SENASA + flete (sin MO/insumos/faena/congelado). Merma: {fmt_pct(pr_merma_media*100, 2)}")

            rm = calcular_media_res(params)
            b = rm['base']
            col_m1, col_m2, col_m3, col_m4 = st.columns(4)
            col_m1.metric("Kg vendibles", fmt_num(b['kg_vendibles'], 1))
            col_m2.metric("Precio $/kg", fmt_dinero(rm['precio_kg']))
            col_m3.metric("Venta total", fmt_dinero(rm['venta_total']))
            col_m4.metric("Resultado", fmt_dinero(rm['resultado']), delta=fmt_pct(rm['margen_pct'], 2))

            with st.expander("Desglose de costos", expanded=False):
                df_cm = pd.DataFrame([
                    ('Compra hacienda', b['costo_hacienda']),
                    ('SENASA', b['costo_senasa']),
                    ('Flete', b['costo_flete']),
                    ('Costos directos', b['costos_directos']),
                    ('IIBB + Ganancias', b['costo_iibb']),
                    ('Imp. Cheque', b['costo_cheque']),
                    ('Costo financiero', b['costo_financiero']),
                    ('TOTAL', b['costo_total']),
                ], columns=['Concepto', 'Monto'])
                df_cm['Monto'] = df_cm['Monto'].apply(fmt_dinero)
                st.dataframe(df_cm, use_container_width=True, hide_index=True)

        elif modo_pricing == '🔪 Por cuartos':
            st.markdown("#### Venta por cuartos")
            st.caption(f"Solo costos: hacienda + SENASA + flete. Merma cuarteo: {fmt_pct(pr_merma_cuarteo*100, 2)}")

            col_cq1, col_cq2 = st.columns(2)
            lista_cq = col_cq1.selectbox("Lista de precios", ['Standard', 'Black', 'Bufalo'], key='lista_cuartos_v2')
            ajuste_cq = col_cq2.number_input("Ajuste general",
                                              value=0.0, step=0.01, format="%.2f",
                                              key='ajuste_cuartos_v2',
                                              help="Ej: 0.05 = +5%")

            rq = calcular_cuartos(params, lista=lista_cq, ajuste_pct=ajuste_cq)
            b = rq['base']

            col_q1, col_q2, col_q3, col_q4 = st.columns(4)
            col_q1.metric("Kg vendibles", fmt_num(rq['kg_total'], 1))
            col_q2.metric("Precio prom $/kg",
                          fmt_dinero(rq['venta_total'] / rq['kg_total']) if rq['kg_total'] > 0 else '$0')
            col_q3.metric("Venta total", fmt_dinero(rq['venta_total']))
            col_q4.metric("Resultado", fmt_dinero(rq['resultado']),
                          delta=fmt_pct(rq['margen_pct'], 2))

            st.markdown(f"**Factor de ajuste aplicado:** {fmt_pct((rq['factor_ajuste']-1)*100, 2)} "
                        f"_(lo que hay que sumar/restar a la lista para lograr el margen objetivo)_")

            df_rq = pd.DataFrame(rq['cuartos'])
            df_rq['Kg'] = df_rq['Kg'].apply(lambda x: fmt_num(x, 1))
            df_rq['Valor estimado'] = df_rq['Valor estimado'].apply(fmt_dinero)
            df_rq['Precio $/kg (lista)'] = df_rq['Precio $/kg (lista)'].apply(fmt_dinero)
            df_rq['Precio $/kg sugerido'] = df_rq['Precio $/kg sugerido'].apply(fmt_dinero)
            st.dataframe(df_rq, use_container_width=True, hide_index=True)

            with st.expander("Desglose de costos", expanded=False):
                df_cq = pd.DataFrame([
                    ('Compra hacienda', b['costo_hacienda']),
                    ('SENASA', b['costo_senasa']),
                    ('Flete', b['costo_flete']),
                    ('Costos directos', b['costos_directos']),
                    ('IIBB + Ganancias', b['costo_iibb']),
                    ('Imp. Cheque', b['costo_cheque']),
                    ('Costo financiero', b['costo_financiero']),
                    ('TOTAL', b['costo_total']),
                ], columns=['Concepto', 'Monto'])
                df_cq['Monto'] = df_cq['Monto'].apply(fmt_dinero)
                st.dataframe(df_cq, use_container_width=True, hide_index=True)

        else:  # 📋 Por cortes
            st.markdown("#### Venta por cortes individuales")
            st.caption("Incluye todos los costos. Precios base: **NETOS PEYA** (editables).")

            col_cc1, col_cc2 = st.columns(2)
            lista_cc = col_cc1.selectbox("Lista base", ['Standard', 'Black', 'Bufalo'], key='lista_cortes_v2')
            ajuste_cc = col_cc2.number_input("Ajuste general",
                                              value=0.0, step=0.01, format="%.2f",
                                              key='ajuste_cortes_v2')

            # ── Toggle: precios LIVE de Google Sheets ──
            col_pl1, col_pl2 = st.columns([3, 1])
            usar_precios_live = col_pl1.checkbox(
                "📡 Usar precios actualizados de Google Sheets (Precios facturación)",
                value=True, key='precios_live_cortes',
                help=("Lee la planilla de Precios facturación y aplica descuentos: "
                      "PEYA -2% merma -5% logística -4,2% comisión, MAXICONSUMO -3,5%, "
                      "LIBERTAD -5%, DIARCO -1,5%.")
            )
            if col_pl2.button("🔄 Refrescar", key='refresh_precios_live',
                                help="Limpiar caché y volver a leer la planilla"):
                cargar_precios_google_sheets.clear()
                st.rerun()

            price_matrix_live = None
            if usar_precios_live and lista_cc == 'Standard':
                with st.spinner("Cargando precios..."):
                    pmx, fecha_mod, status_pmx = cargar_precios_google_sheets()
                if pmx:
                    price_matrix_live = pmx
                    st.success(
                        f"✅ Precios live: {status_pmx}"
                        + (f" · actualizada: {fecha_mod[:10]}" if fecha_mod else '')
                    )
                else:
                    st.warning(f"⚠️ No se pudo leer la planilla ({status_pmx}). "
                               "Usando precios precargados.")

            # ── Rendimiento: hardcodeado vs real de los romaneos cargados ──
            from datetime import datetime as _dtnow
            rom_cargados = [_p for _p in st.session_state.get('parsed_files', [])
                             if not _p.get('error') and _p.get('cortes')]

            usar_real = st.checkbox(
                "📊 Usar rendimiento real de los romaneos cargados",
                value=False, key='rend_real_cortes',
                help=("Reemplaza la distribución hardcodeada por la calculada "
                      "agregando los cortes de los romaneos cargados.")
            )
            grupos_override = None
            rend_info = ""
            if usar_real:
                if not rom_cargados:
                    st.warning("No hay romaneos cargados. Cargá archivos en 📤 para activar.")
                else:
                    años_r = sorted({_dtnow.strptime(p['fecha'], '%d/%m/%Y').year
                                      for p in rom_cargados if p.get('fecha')}, reverse=True)
                    meses_n = ['Ene','Feb','Mar','Abr','May','Jun',
                               'Jul','Ago','Sep','Oct','Nov','Dic']
                    col_r1, col_r2, col_r3 = st.columns(3)
                    año_r = col_r1.selectbox("Año", años_r,
                                              index=0, key='rend_real_anio')
                    mes_r = col_r2.selectbox(
                        "Mes", [None] + list(range(1, 13)),
                        format_func=lambda m: 'Todos' if m is None else meses_n[m-1],
                        index=_dtnow.now().month, key='rend_real_mes'
                    )
                    cal_r = col_r3.selectbox(
                        "Calidad", ['Todas'] + ['Standard', 'Búfalo', 'Premium Black', 'Exportación'],
                        index=0 if lista_cc != 'Standard' else 1,
                        key='rend_real_cal'
                    )
                    cal_filt = None if cal_r == 'Todas' else cal_r
                    grupos_override = rendimiento_real_de_romaneos(
                        rom_cargados, mes=mes_r, año=año_r, calidad=cal_filt
                    )
                    if grupos_override:
                        n_rom = sum(1 for p in rom_cargados
                                    if (mes_r is None or
                                        _dtnow.strptime(p['fecha'], '%d/%m/%Y').month == mes_r)
                                    and _dtnow.strptime(p['fecha'], '%d/%m/%Y').year == año_r
                                    and (cal_filt is None or p.get('calidad') == cal_filt))
                        rend_info = f"✅ Rendimiento de **{n_rom} romaneo(s)** aplicado."
                        st.success(rend_info)
                    else:
                        st.warning("No hay datos suficientes para el filtro elegido.")
                        grupos_override = None

            if 'precios_custom' not in st.session_state:
                st.session_state.precios_custom = {}

            lista_base = construir_lista_cortes(lista=lista_cc, ajuste_pct=0,
                                                  grupos_override=grupos_override,
                                                  price_matrix=price_matrix_live)

            st.markdown("##### Precios base editables")
            st.caption("Editá el precio de cada corte para ver cómo impacta en el total.")

            df_edit = pd.DataFrame(lista_base)
            for i, row in df_edit.iterrows():
                if row['Corte'] in st.session_state.precios_custom:
                    df_edit.at[i, 'Precio base'] = st.session_state.precios_custom[row['Corte']]

            edited = st.data_editor(
                df_edit[['Cuarto', 'Corte', 'Precio base']].rename(columns={'Precio base': 'Precio $/kg'}),
                column_config={
                    'Cuarto': st.column_config.TextColumn(disabled=True),
                    'Corte': st.column_config.TextColumn(disabled=True),
                    'Precio $/kg': st.column_config.NumberColumn(min_value=0, step=100, format='$%d'),
                },
                hide_index=True, use_container_width=True, key='editor_precios'
            )

            precios_custom = {}
            for _, row in edited.iterrows():
                precios_custom[row['Corte']] = row['Precio $/kg']
            st.session_state.precios_custom = precios_custom

            rc = calcular_cortes(params, lista=lista_cc, precios_custom=precios_custom,
                                  ajuste_pct=ajuste_cc, grupos_override=grupos_override,
                                  price_matrix=price_matrix_live)
            b = rc['base']

            st.markdown("---")
            col_co1, col_co2, col_co3, col_co4 = st.columns(4)
            col_co1.metric("Kg vendibles", fmt_num(rc['kg_total'], 1))
            col_co2.metric("Precio prom $/kg", fmt_dinero(rc['precio_prom']))
            col_co3.metric("Venta total", fmt_dinero(rc['venta_total']))
            col_co4.metric("Resultado", fmt_dinero(rc['resultado']),
                           delta=fmt_pct(rc['margen_pct'], 2))

            diff_margen = rc['margen_pct'] - rc['margen_objetivo_pct']
            if diff_margen >= 0:
                st.success(
                    f"✅ Margen real {fmt_pct(rc['margen_pct'], 2)} supera al objetivo "
                    f"({fmt_pct(rc['margen_objetivo_pct'], 2)}). Podés bajar precios hasta "
                    f"{fmt_pct(rc['ajuste_necesario_pct'], 2)} para llegar al objetivo."
                )
            else:
                st.warning(
                    f"⚠️ Margen real {fmt_pct(rc['margen_pct'], 2)} por debajo del objetivo "
                    f"({fmt_pct(rc['margen_objetivo_pct'], 2)}). Necesitás subir {fmt_pct(rc['ajuste_necesario_pct'], 2)} "
                    f"a toda la lista para llegar."
                )

            st.markdown("---")
            st.markdown("##### 📊 Peso de cada corte en el total")
            st.caption(
                "**% kg total**: cuánto pesa en volumen.  "
                "**% valor total**: cuánto pesa en facturación.  "
                "**Impacto +1% precio**: si subís 1% este corte, el precio promedio general sube X%."
            )

            df_rc = pd.DataFrame(rc['cortes'])
            df_rc = df_rc.sort_values('% del valor total', ascending=False).reset_index(drop=True)

            df_show = df_rc[['Cuarto', 'Corte', 'Kg', 'Precio $/kg', 'Valor',
                             '% del kg total', '% del valor total', 'Impacto +1% precio']].copy()
            df_show['Kg'] = df_show['Kg'].apply(lambda x: fmt_num(x, 1))
            df_show['Precio $/kg'] = df_show['Precio $/kg'].apply(fmt_dinero)
            df_show['Valor'] = df_show['Valor'].apply(fmt_dinero)
            df_show['% del kg total'] = df_show['% del kg total'].apply(lambda x: fmt_pct(x, 2))
            df_show['% del valor total'] = df_show['% del valor total'].apply(lambda x: fmt_pct(x, 2))
            df_show['Impacto +1% precio'] = df_show['Impacto +1% precio'].apply(lambda x: fmt_pct(x, 3))

            def color_cuarto(val):
                return {
                    'Bife': 'background-color: #FFEBEE',
                    'Asado': 'background-color: #FFF3E0',
                    'Pecho': 'background-color: #E3F2FD',
                    'Mocho': 'background-color: #E8F5E9',
                }.get(val, '')
            st.dataframe(
                df_show.style.applymap(color_cuarto, subset=['Cuarto']),
                use_container_width=True, hide_index=True
            )

            if st.button("🔄 Resetear a precios base", key='reset_precios'):
                st.session_state.precios_custom = {}
                st.rerun()

            st.markdown("---")
            if st.button("📥 Exportar lista a Excel", key='exp_pricing_v2'):
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                import io

                wb = Workbook()
                ws = wb.active
                ws.title = f"Lista {lista_cc}"

                ws['A1'] = f'LISTA DE PRECIOS — {lista_cc.upper()}'
                ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
                ws.merge_cells('A1:H1')
                ws['A2'] = (f'Ajuste: {fmt_pct(ajuste_cc*100, 2)}  |  '
                            f'Margen objetivo: {fmt_pct(pr_margen_cortes*100, 2)}  |  '
                            f'Margen real: {fmt_pct(rc["margen_pct"], 2)}')
                ws['A2'].font = Font(italic=True, size=10, color='666666')
                ws.merge_cells('A2:H2')

                headers_x = ['Cuarto', 'Corte', 'Kg', 'Precio $/kg', 'Valor', '% kg', '% valor', 'Impacto +1%']
                for i, h in enumerate(headers_x, 1):
                    cc = ws.cell(4, i, h)
                    cc.font = Font(bold=True, color='FFFFFF')
                    cc.fill = PatternFill('solid', fgColor='1F4E79')
                    cc.alignment = Alignment(horizontal='center')

                for r_idx, row in enumerate(rc['cortes'], 5):
                    ws.cell(r_idx, 1, row['Cuarto'])
                    ws.cell(r_idx, 2, row['Corte'])
                    ws.cell(r_idx, 3, row['Kg']).number_format = '#,##0.0'
                    ws.cell(r_idx, 4, row['Precio $/kg']).number_format = '$#,##0'
                    ws.cell(r_idx, 5, row['Valor']).number_format = '$#,##0'
                    ws.cell(r_idx, 6, row['% del kg total'] / 100).number_format = '0.00%'
                    ws.cell(r_idx, 7, row['% del valor total'] / 100).number_format = '0.00%'
                    ws.cell(r_idx, 8, row['Impacto +1% precio'] / 100).number_format = '0.000%'

                for col, w in [('A', 12), ('B', 28), ('C', 10), ('D', 14),
                               ('E', 16), ('F', 10), ('G', 10), ('H', 14)]:
                    ws.column_dimensions[col].width = w

                ws.sheet_view.showGridLines = False

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                from datetime import datetime as dt
                st.download_button(
                    "⬇️ Descargar",
                    buf.getvalue(),
                    file_name=f"Pricing_{lista_cc}_{dt.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key='dl_pricing_v2'
                )

            # ══════════════════════════════════════════════════════════════
            # APLICAR PRICING A UN ROMANEO REAL
            # ══════════════════════════════════════════════════════════════
            st.markdown("---")
            st.markdown("#### 🎯 Aplicar pricing a un romaneo real")
            st.caption(
                "Toma los **kg reales** de los cortes de un romaneo y los multiplica "
                "por los precios de la lista actual para ver cómo daría el resultado real."
            )

            romaneos_disp = [p for p in st.session_state.get('parsed_files', [])
                              if not p.get('error') and p.get('cortes')]
            if not romaneos_disp:
                st.info("Cargá romaneos en la solapa **📤 Cargar archivos** "
                        "para usar esta función.")
            else:
                from datetime import datetime as _dt
                # Parsear fecha y ordenar (más reciente primero)
                def _fecha_dt(p):
                    try:
                        return _dt.strptime(p.get('fecha', ''), '%d/%m/%Y')
                    except Exception:
                        return _dt.min
                romaneos_disp = sorted(romaneos_disp, key=_fecha_dt, reverse=True)

                # Filtro mes/año
                años_rom = sorted({_fecha_dt(p).year for p in romaneos_disp
                                    if _fecha_dt(p) != _dt.min}, reverse=True)
                if not años_rom:
                    años_rom = [_dt.now().year]

                meses_n = ['Ene','Feb','Mar','Abr','May','Jun',
                           'Jul','Ago','Sep','Oct','Nov','Dic']
                col_fa, col_fm = st.columns(2)
                año_aplica = col_fa.selectbox("Año", años_rom, index=0, key='aplica_anio')
                mes_aplica = col_fm.selectbox(
                    "Mes", list(range(1, 13)),
                    format_func=lambda m: meses_n[m-1],
                    index=_dt.now().month - 1,
                    key='aplica_mes'
                )

                # Romaneos del mes/año
                romaneos_mes = [p for p in romaneos_disp
                                 if _fecha_dt(p).month == mes_aplica
                                 and _fecha_dt(p).year == año_aplica]

                if not romaneos_mes:
                    st.warning(f"No hay romaneos cargados de "
                               f"{meses_n[mes_aplica-1]} {año_aplica}.")
                else:
                    opciones_rom = {
                        f"{p.get('fecha', '?')} · {p.get('archivo', '?')} · "
                        f"{p.get('kg_carne', 0):,.0f} kg carne": idx
                        for idx, p in enumerate(romaneos_mes)
                    }
                    sel_label = st.selectbox(
                        "Romaneo (último primero)",
                        list(opciones_rom.keys()),
                        key='sel_rom_aplica'
                    )
                    p_sel = romaneos_mes[opciones_rom[sel_label]]

                    # Params para el cálculo: usamos los datos REALES del romaneo
                    medias = p_sel.get('medias_reses', 1) or 1
                    kg_ent = p_sel.get('kg_entrada', 0) or 0
                    peso_real = (kg_ent / medias) if medias > 0 else pr_peso
                    rend_real = (p_sel.get('kg_carne', 0) / kg_ent) if kg_ent > 0 else pr_rend
                    precio_compra_real = p_sel.get('precio_compra') or pr_compra

                    # Sumar para TODAS las medias del romaneo: usamos kg_entrada total
                    params_rom = {**params,
                                   'peso_media': kg_ent,        # kg ingresados totales
                                   'rendimiento': rend_real,
                                   'precio_compra_kg': precio_compra_real}

                    res_aplica = aplicar_precios_a_romaneo(
                        p_sel.get('cortes', []), precios_custom, params_rom
                    )

                    # ── Métricas ──
                    col_x1, col_x2, col_x3, col_x4 = st.columns(4)
                    col_x1.metric("Kg carne (real)",
                                   fmt_num(res_aplica['kg_total_carne'], 1))
                    col_x2.metric("Precio prom $/kg (real)",
                                   fmt_dinero(res_aplica['precio_prom_total']),
                                   help="Venta total / kg carne TOTAL del romaneo")
                    col_x3.metric("Venta total",
                                   fmt_dinero(res_aplica['venta_total']))
                    col_color = 'normal' if res_aplica['resultado'] >= 0 else 'inverse'
                    col_x4.metric("Resultado",
                                   fmt_dinero(res_aplica['resultado']),
                                   delta=fmt_pct(res_aplica['margen_pct'], 2),
                                   delta_color=col_color)

                    # Comparación contra precio teórico de la lista
                    delta_precio = res_aplica['precio_prom_total'] - rc['precio_prom']
                    pct_dif = (delta_precio / rc['precio_prom'] * 100) if rc['precio_prom'] else 0
                    col_y1, col_y2, col_y3 = st.columns(3)
                    col_y1.metric("Cobertura cortes",
                                   f"{res_aplica['cobertura_pct']:.1f}%",
                                   help="% de kg del romaneo que matchearon contra la lista")
                    col_y2.metric("Precio teórico (pricing)", fmt_dinero(rc['precio_prom']))
                    col_y3.metric(
                        "Diferencia vs teórico",
                        fmt_dinero(delta_precio),
                        delta=fmt_pct(pct_dif, 2),
                        delta_color='normal' if delta_precio >= 0 else 'inverse',
                        help="precio real - precio teórico"
                    )

                    if res_aplica['margen_pct'] >= rc['margen_objetivo_pct']:
                        st.success(
                            f"✅ Vendiendo este romaneo a la lista actual obtenés "
                            f"{fmt_pct(res_aplica['margen_pct'], 2)} de margen "
                            f"(objetivo {fmt_pct(rc['margen_objetivo_pct'], 2)})."
                        )
                    else:
                        st.warning(
                            f"⚠️ El margen real sería {fmt_pct(res_aplica['margen_pct'], 2)}, "
                            f"por debajo del objetivo {fmt_pct(rc['margen_objetivo_pct'], 2)}."
                        )

                    # Cortes sin match (gap importante para entender la diferencia)
                    if res_aplica['cortes_sin_match']:
                        with st.expander(
                            f"⚠️ {len(res_aplica['cortes_sin_match'])} cortes "
                            f"({res_aplica['kg_sin_match']:,.1f} kg) sin precio en la lista"
                            .replace(',', '.')
                        ):
                            df_sm = pd.DataFrame(res_aplica['cortes_sin_match'])
                            df_sm = df_sm.groupby('corte', as_index=False).agg({'kg': 'sum'})
                            df_sm = df_sm.sort_values('kg', ascending=False)
                            df_sm['kg'] = df_sm['kg'].apply(lambda x: fmt_num(x, 1))
                            st.dataframe(df_sm, use_container_width=True, hide_index=True)
                            st.caption(
                                "Estos cortes están en el romaneo pero no aparecen en "
                                "la lista de pricing. **No suman a la venta**, por eso "
                                "el precio promedio puede dar más bajo que el teórico."
                            )

                    # Detalle de los matcheados
                    with st.expander("📋 Detalle de cortes valorizados",
                                      expanded=False):
                        df_m = pd.DataFrame(res_aplica['cortes_match'])
                        if not df_m.empty:
                            df_m = df_m.groupby(['corte', 'precio'], as_index=False).agg(
                                {'kg': 'sum', 'valor': 'sum'}
                            ).sort_values('valor', ascending=False)
                            df_m['kg'] = df_m['kg'].apply(lambda x: fmt_num(x, 1))
                            df_m['precio'] = df_m['precio'].apply(fmt_dinero)
                            df_m['valor'] = df_m['valor'].apply(fmt_dinero)
                            st.dataframe(df_m, use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════════════
    # MODALIDAD: HAMBURGUESAS (parámetros propios)
    # ══════════════════════════════════════════════════════════════════
    else:
        st.markdown("#### 🍔 Costeo de Hamburguesas")

        receta_hb = st.radio(
            "**Receta**",
            ['TFC (recorte 80/20)', 'Kosher Diarco (carne picada)'],
            horizontal=True, key='hamb_receta_v3'
        )

        # ── 1. Lote y formulación ──
        st.markdown("#### 1. Lote y formulación")
        col_ha, col_hb, col_hc = st.columns(3)
        hb_lote = col_ha.number_input(
            "Lote (kg MP)", value=1800 if 'TFC' in receta_hb else 600,
            step=50, key='hb_lote_v3'
        )
        hb_peso_hamb = col_hb.number_input(
            "Peso hamburguesa (kg)", value=0.0835, step=0.005, format="%.4f", key='hb_peso_v3'
        )
        hb_und_flow = col_hc.number_input(
            "Unidades por flow pack", value=2, step=1, key='hb_und_v3'
        )

        st.markdown("##### Fórmula (en %) — incluye grasa")
        if 'TFC' in receta_hb:
            nombre_mp_hb = 'Recorte 80/20'
            col_f1, col_f2, col_f3, col_f4 = st.columns(4)
            pct_mp_h = col_f1.number_input(f"{nombre_mp_hb} (%)",
                                             value=90.00, step=0.5, format="%.2f", key='hb_pct_mp_v3')
            pct_grasa_h = col_f2.number_input("Grasa (%)",
                                               value=6.00, step=0.5, format="%.2f", key='hb_pct_gr_v3')
            pct_int_h = col_f3.number_input("Integral (%)",
                                              value=1.86, step=0.1, format="%.2f", key='hb_pct_int_v3')
            pct_agua_h = col_f4.number_input("Agua (%)",
                                               value=2.13, step=0.1, format="%.2f", key='hb_pct_ag_v3')
        else:
            nombre_mp_hb = 'Carne picada'
            col_f1, col_f2, col_f3, col_f4 = st.columns(4)
            pct_mp_h = col_f1.number_input(f"{nombre_mp_hb} (%)",
                                             value=88.00, step=0.5, format="%.2f", key='hb_pct_mp_k_v3')
            pct_grasa_h = col_f2.number_input("Grasa (%)",
                                               value=4.28, step=0.5, format="%.2f", key='hb_pct_gr_k_v3')
            pct_int_h = col_f3.number_input("Integral (%)",
                                              value=3.11, step=0.1, format="%.2f", key='hb_pct_int_k_v3')
            pct_agua_h = col_f4.number_input("Agua (%)",
                                               value=4.61, step=0.1, format="%.2f", key='hb_pct_ag_k_v3')

        suma_pct_hb = pct_mp_h + pct_grasa_h + pct_int_h + pct_agua_h
        if abs(suma_pct_hb - 100) > 0.5:
            st.warning(f"⚠️ La fórmula suma {fmt_pct(suma_pct_hb, 2)} (debería ser ~100%)")

        # ── 2. Costos de materia prima y producción + empaque y logística ──
        st.markdown("#### 2. Costos de MP, producción, empaque y logística (sin IVA)")
        st.caption("Todos los costos específicos del producto.")

        # Materia prima
        st.markdown("**Materia prima**")
        col_mp1, col_mp2, col_mp3 = st.columns(3)
        hb_cost_mp = col_mp1.number_input(
            f"{nombre_mp_hb} $/kg",
            value=5500 if 'TFC' in receta_hb else 15000,
            step=100, key='hb_cost_mp_v3'
        )
        hb_cost_grasa = col_mp2.number_input(
            "Grasa $/kg", value=1500, step=100, key='hb_cost_grasa_v3'
        )
        hb_cost_int = col_mp3.number_input(
            "Integral $/kg", value=7849, step=100, key='hb_cost_int_v3'
        )

        # Producción
        st.markdown("**Producción**")
        col_p1, col_p2 = st.columns(2)
        hb_cost_mo = col_p1.number_input(
            "Mano de obra $/kg producido",
            value=1117 if 'TFC' in receta_hb else 2000,
            step=50, key='hb_cost_mo_v3'
        )
        hb_merma_pct_in = col_p2.number_input(
            "Merma producción (%)", value=2.00, step=0.1, format="%.2f", key='hb_merma_v3'
        )
        hb_merma = hb_merma_pct_in / 100

        # Empaque y logística
        st.markdown("**Empaque y logística**")
        col_e1, col_e2, col_e3, col_e4 = st.columns(4)
        hb_cost_flow = col_e1.number_input(
            "Flow pack $/u",
            value=39.52 if 'TFC' in receta_hb else 31.20,
            step=1.0, format="%.2f", key='hb_cost_flow_v3'
        )
        hb_cost_caja = col_e2.number_input(
            "Caja 48u $/u",
            value=405 if 'TFC' in receta_hb else 315,
            step=10, key='hb_cost_caja_v3'
        )
        hb_cost_tte1 = col_e3.number_input(
            "Transporte Moreno→EXH $",
            value=270000 if 'TFC' in receta_hb else 250000,
            step=10000, key='hb_tte1_v3'
        )
        hb_cost_tte2 = col_e4.number_input(
            "Transporte EXH→Diarco $",
            value=350000 if 'TFC' in receta_hb else 250000,
            step=10000, key='hb_tte2_v3'
        )

        # ── 3. Costos operativos (con flete a destino) ──
        st.markdown("#### 3. Costos operativos (con flete a destino)")
        col_o1, col_o2, col_o3, col_o4 = st.columns(4)
        hb_flete_dest = col_o1.number_input(
            "Flete a destino $/kg", value=0, step=10, key='hb_flete_dest',
            help="Flete adicional al cliente final"
        )
        hb_tna_pct = col_o2.number_input(
            "TNA (%)", value=48.00, step=1.0, format="%.2f", key='hb_tna_v3'
        )
        hb_dias = col_o3.number_input(
            "Días financ.", value=25, step=1, key='hb_dias_v3'
        )
        hb_cost_cong = col_o4.number_input(
            "Congelado $/kg", value=0, step=10, key='hb_cong_v3'
        )

        col_o5, col_o6 = st.columns(2)
        hb_iibb_pct = col_o5.number_input(
            "IIBB+Ganancias (%)", value=3.50, step=0.1, format="%.2f", key='hb_iibb_v3'
        )
        hb_cheque_pct = col_o6.number_input(
            "Imp. Cheque (%)", value=1.20, step=0.1, format="%.2f", key='hb_cheque_v3'
        )

        hb_tna = hb_tna_pct / 100
        hb_iibb = hb_iibb_pct / 100
        hb_cheque = hb_cheque_pct / 100

        # ── 4. Rentabilidad ──
        st.markdown("#### 4. Rentabilidad objetivo")
        hb_margen_h_pct = st.number_input(
            "Rentabilidad objetivo (%)", value=10.00, step=0.5, format="%.2f", key='hb_margen_v4'
        )
        hb_margen_h = hb_margen_h_pct / 100

        # ══════ CÁLCULOS ══════
        # Proporciones normalizadas
        pct_mp = pct_mp_h / suma_pct_hb if suma_pct_hb > 0 else 0
        pct_gr = pct_grasa_h / suma_pct_hb if suma_pct_hb > 0 else 0
        pct_in = pct_int_h / suma_pct_hb if suma_pct_hb > 0 else 0
        pct_ag = pct_agua_h / suma_pct_hb if suma_pct_hb > 0 else 0

        kg_mp_h = hb_lote * pct_mp
        kg_grasa_h = hb_lote * pct_gr
        kg_int_h = hb_lote * pct_in
        kg_agua_h = hb_lote * pct_ag

        kg_producidos_h = hb_lote * (1 - hb_merma)
        cant_hamb_h = kg_producidos_h / hb_peso_hamb if hb_peso_hamb > 0 else 0
        cant_flowpacks_h = cant_hamb_h / hb_und_flow if hb_und_flow > 0 else 0
        cant_cajas_h = cant_flowpacks_h / 24

        # Costos MP
        c_mp = kg_mp_h * hb_cost_mp
        c_gr = kg_grasa_h * hb_cost_grasa
        c_in_h = kg_int_h * hb_cost_int
        c_ag_h = 0

        # Producción
        c_mo_h = kg_producidos_h * hb_cost_mo

        # Empaque + logística interna
        c_flow_h = cant_flowpacks_h * hb_cost_flow
        c_caja_h = cant_cajas_h * hb_cost_caja
        c_tte_interno = hb_cost_tte1 + hb_cost_tte2

        # Operativos
        c_flete_dest_h = kg_producidos_h * hb_flete_dest
        c_cong_h = kg_producidos_h * hb_cost_cong

        costos_directos_h = (c_mp + c_gr + c_in_h + c_ag_h + c_mo_h
                             + c_flow_h + c_caja_h + c_tte_interno
                             + c_flete_dest_h + c_cong_h)

        costo_financiero_h = costos_directos_h * hb_tna * hb_dias / 365

        # Iteración para impuestos sobre venta
        venta_est_h = costos_directos_h / (1 - hb_margen_h) if hb_margen_h < 1 else costos_directos_h
        for _ in range(5):
            c_iibb_h = venta_est_h * hb_iibb
            c_cheque_h = venta_est_h * hb_cheque
            costo_total_h = costos_directos_h + costo_financiero_h + c_iibb_h + c_cheque_h
            venta_est_h = costo_total_h / (1 - hb_margen_h) if hb_margen_h < 1 else costo_total_h

        c_iibb_h = venta_est_h * hb_iibb
        c_cheque_h = venta_est_h * hb_cheque
        costo_total_h = costos_directos_h + costo_financiero_h + c_iibb_h + c_cheque_h

        costo_kg_h = costo_total_h / kg_producidos_h if kg_producidos_h > 0 else 0
        costo_flow_h_u = costo_total_h / cant_flowpacks_h if cant_flowpacks_h > 0 else 0
        precio_kg_neto_h = costo_kg_h / (1 - hb_margen_h) if hb_margen_h < 1 else costo_kg_h
        precio_flow_neto_h = costo_flow_h_u / (1 - hb_margen_h) if hb_margen_h < 1 else costo_flow_h_u
        precio_kg_iva_h = precio_kg_neto_h * 1.21
        precio_flow_iva_h = precio_flow_neto_h * 1.21
        resultado_h = venta_est_h - costo_total_h
        margen_real_h = resultado_h / venta_est_h * 100 if venta_est_h > 0 else 0

        st.markdown("---")
        st.markdown("#### 5. Resultado del lote")
        col_k1, col_k2, col_k3, col_k4 = st.columns(4)
        col_k1.metric("Kg producidos", fmt_num(kg_producidos_h, 1))
        col_k2.metric("Hamburguesas", fmt_num(cant_hamb_h, 0))
        col_k3.metric("Flow packs", fmt_num(cant_flowpacks_h, 0))
        col_k4.metric("Cajas (48u)", fmt_num(cant_cajas_h, 0))

        col_t1, col_t2, col_t3, col_t4 = st.columns(4)
        col_t1.metric("Costo total", fmt_dinero(costo_total_h))
        col_t2.metric("Costo $/kg", fmt_dinero(costo_kg_h))
        col_t3.metric("Precio sugerido $/kg", fmt_dinero(precio_kg_neto_h))
        col_t4.metric("Precio $/flow pack", fmt_dinero(precio_flow_neto_h, 2))

        st.markdown("---")
        st.markdown("#### 6. Precios finales")
        col_pp1, col_pp2 = st.columns(2)
        with col_pp1:
            st.markdown("**Por kg**")
            st.markdown(f"""
| Concepto | Precio |
|---|---|
| Costo neto | {fmt_dinero(costo_kg_h)} |
| **Precio + rentabilidad** | **{fmt_dinero(precio_kg_neto_h)}** |
| Precio con IVA (21%) | {fmt_dinero(precio_kg_iva_h)} |
            """)
        with col_pp2:
            st.markdown("**Por flow pack**")
            st.markdown(f"""
| Concepto | Precio |
|---|---|
| Costo neto | {fmt_dinero(costo_flow_h_u, 2)} |
| **Precio + rentabilidad** | **{fmt_dinero(precio_flow_neto_h, 2)}** |
| Precio con IVA (21%) | {fmt_dinero(precio_flow_iva_h, 2)} |
            """)

        with st.expander("📊 Desglose completo", expanded=False):
            rows = [
                (f'{nombre_mp_hb} ({fmt_num(kg_mp_h, 2)} kg)', c_mp),
                (f'Grasa ({fmt_num(kg_grasa_h, 2)} kg × {fmt_dinero(hb_cost_grasa)})', c_gr),
                (f'Integral ({fmt_num(kg_int_h, 2)} kg)', c_in_h),
                (f'Agua ({fmt_num(kg_agua_h, 2)} kg)', c_ag_h),
                (f'Mano de obra ({fmt_num(kg_producidos_h, 0)} kg × {fmt_dinero(hb_cost_mo)})', c_mo_h),
                (f'Flow packs ({fmt_num(cant_flowpacks_h, 0)} × {fmt_dinero(hb_cost_flow, 2)})', c_flow_h),
                (f'Cajas ({fmt_num(cant_cajas_h, 0)})', c_caja_h),
                ('Transporte interno (Moreno↔EXH↔Diarco)', c_tte_interno),
                (f'Flete a destino ({fmt_dinero(hb_flete_dest)}/kg)', c_flete_dest_h),
                (f'Congelado ({fmt_dinero(hb_cost_cong)}/kg)', c_cong_h),
                ('COSTOS DIRECTOS', costos_directos_h),
                (f'Costo financiero ({fmt_pct(hb_tna*100, 2)} × {hb_dias} días)', costo_financiero_h),
                (f'IIBB ({fmt_pct(hb_iibb*100, 2)})', c_iibb_h),
                (f'Imp. Cheque ({fmt_pct(hb_cheque*100, 2)})', c_cheque_h),
                ('TOTAL', costo_total_h),
            ]
            df_desg_h = pd.DataFrame(rows, columns=['Concepto', 'Monto'])
            df_desg_h['% total'] = df_desg_h['Monto'].apply(
                lambda x: fmt_pct(x/costo_total_h*100, 2) if costo_total_h > 0 else '-')
            df_desg_h['Monto'] = df_desg_h['Monto'].apply(fmt_dinero)
            st.dataframe(df_desg_h, use_container_width=True, hide_index=True)


# ── TAB: REVOLEO ──
with tab_revoleo:
    from pricing import fmt_num, fmt_pct, fmt_dinero
    import pandas as pd

    st.markdown("### 🔄 Revoleo — Calculadora de compra/reventa")
    st.markdown("_Definí precio de venta objetivo y analizá la rentabilidad real._")

    modo_rev = st.radio(
        "**Modo**",
        ['Compra individual', 'Venta combinada (mismo producto)', 'Venta multi-producto (mismo cliente)'],
        horizontal=True, key='modo_revoleo'
    )

    st.markdown("---")

    def _calcular_revoleo(kg, costo_kg, iva_compra, flete, gastos_op,
                          com1, com2, iibb, imp_cheque, tna, dias, margen,
                          precio_real=None):
        """Calcula costos, precio objetivo y análisis de venta real."""
        total_kg = kg if kg > 0 else 1
        costo_base_kg = costo_kg + flete + gastos_op
        prop = com1 + com2 + iibb + imp_cheque  # % sobre venta
        cf = costo_base_kg * (tna / 365) * dias
        costo_fijo = costo_base_kg + cf
        pmin = costo_fijo / (1 - prop) if prop < 1 else costo_fijo
        d = 1 - (1 + margen) * prop
        if d <= 0.001:
            d = 0.001
        pobj = (1 + margen) * costo_fijo / d
        pobj_iva = pobj * (1 + iva_compra)

        resultado = {
            'total_kg': total_kg,
            'costo_base_kg': costo_base_kg,
            'prop': prop,
            'cf': cf,
            'costo_fijo': costo_fijo,
            'pmin': pmin,
            'pobj': pobj,
            'pobj_iva': pobj_iva,
            'iva_compra': iva_compra,
        }

        if precio_real and precio_real > 0:
            tasas = precio_real * prop
            ct = costo_base_kg + tasas + cf
            gk = precio_real - ct
            gt = gk * total_kg
            rent = (gk / ct) * 100 if ct > 0 else 0
            resultado.update({
                'precio_real': precio_real,
                'tasas': tasas,
                'ct': ct,
                'ganancia_kg': gk,
                'ganancia_total': gt,
                'rentabilidad': rent,
            })

        return resultado

    # ══════════════════════════════════════════════════════════════════
    # MODO: COMPRA INDIVIDUAL
    # ══════════════════════════════════════════════════════════════════
    if modo_rev == 'Compra individual':
        st.markdown("#### 1. Datos de compra")
        col_a1, col_a2, col_a3, col_a4 = st.columns(4)
        rv_producto = col_a1.text_input("Producto", value="BONDIOLA", key='rv_prod')
        rv_kg = col_a2.number_input("Kg totales", value=6000.0, step=100.0,
                                     format="%.2f", key='rv_kg')
        rv_costo = col_a3.number_input("Costo $/kg s/IVA", value=9650, step=100, key='rv_costo')
        rv_iva_pct = col_a4.number_input("IVA compra (%)", value=10.50, step=0.5,
                                          format="%.2f", key='rv_iva')

        st.markdown("#### 2. Costos operativos")
        col_b1, col_b2, col_b3, col_b4 = st.columns(4)
        rv_flete = col_b1.number_input("Flete $/kg", value=70, step=10, key='rv_flete')
        rv_gastos = col_b2.number_input("Gastos op. $/kg", value=1000, step=50, key='rv_gastos')
        rv_com1_pct = col_b3.number_input("Comisión 1 (%)", value=0.00, step=0.5, format="%.2f", key='rv_com1')
        rv_com2_pct = col_b4.number_input("Comisión 2 (%)", value=1.00, step=0.5, format="%.2f", key='rv_com2')

        col_c1, col_c2, col_c3, col_c4 = st.columns(4)
        rv_iibb_pct = col_c1.number_input("IIBB (%)", value=4.00, step=0.5, format="%.2f", key='rv_iibb')
        rv_cheque_pct = col_c2.number_input("Imp. Cheque (%)", value=1.20, step=0.1, format="%.2f", key='rv_chq')
        rv_tna_pct = col_c3.number_input("TNA (%)", value=48.00, step=1.0, format="%.2f", key='rv_tna')
        rv_dias = col_c4.number_input("Días financ.", value=10, step=1, key='rv_dias')

        st.markdown("#### 3. Precio objetivo")
        col_d1, col_d2 = st.columns(2)
        rv_margen_pct = col_d1.number_input("Margen deseado (%)", value=15.00, step=0.5, format="%.2f", key='rv_margen')
        rv_precio_real = col_d2.number_input("Precio real de venta $/kg s/IVA",
                                              value=0, step=100, key='rv_preal',
                                              help="Dejá 0 para solo ver precio objetivo")

        # Cálculo
        res = _calcular_revoleo(
            rv_kg, rv_costo, rv_iva_pct/100,
            rv_flete, rv_gastos,
            rv_com1_pct/100, rv_com2_pct/100, rv_iibb_pct/100, rv_cheque_pct/100,
            rv_tna_pct/100, rv_dias, rv_margen_pct/100,
            precio_real=rv_precio_real if rv_precio_real > 0 else None
        )

        st.markdown("---")

        # Resultado principal
        st.markdown("#### Precio objetivo de venta")
        col_r1, col_r2 = st.columns(2)
        col_r1.metric("💰 Precio objetivo $/kg s/IVA", fmt_dinero(res['pobj']))
        col_r2.metric("Con IVA", fmt_dinero(res['pobj_iva']))

        with st.expander("📊 Desglose de costos", expanded=True):
            # Comisiones e impuestos calculados al precio objetivo (lo que
            # realmente se cobra cuando vendés al precio sugerido)
            pobj = res['pobj']
            pmin = res['pmin']
            com1_kg_obj = pobj * rv_com1_pct / 100
            com2_kg_obj = pobj * rv_com2_pct / 100
            iibb_kg_obj = pobj * rv_iibb_pct / 100
            chq_kg_obj  = pobj * rv_cheque_pct / 100
            total_imp_obj = com1_kg_obj + com2_kg_obj + iibb_kg_obj + chq_kg_obj
            margen_kg = pobj - res['costo_fijo'] - total_imp_obj

            filas = [
                ('Total kg', fmt_num(res['total_kg'], 1) + ' kg'),
                ('— COSTOS DIRECTOS —', ''),
                ('Costo materia prima', fmt_dinero(rv_costo) + '/kg'),
                ('Flete + Gastos op.', fmt_dinero(rv_flete + rv_gastos) + '/kg'),
                (f'Costo financiero (TNA {rv_tna_pct:.1f}% × {rv_dias}d)',
                 fmt_dinero(res['cf']) + '/kg'),
                ('Subtotal directos (s/comisiones)',
                 fmt_dinero(res['costo_fijo']) + '/kg'),
                ('— COMISIONES E IMPUESTOS (sobre precio objetivo) —', ''),
            ]
            if rv_com1_pct > 0:
                filas.append((f'Comisión 1 ({fmt_pct(rv_com1_pct, 2)})',
                              fmt_dinero(com1_kg_obj) + '/kg'))
            if rv_com2_pct > 0:
                filas.append((f'Comisión 2 ({fmt_pct(rv_com2_pct, 2)})',
                              fmt_dinero(com2_kg_obj) + '/kg'))
            if rv_iibb_pct > 0:
                filas.append((f'IIBB ({fmt_pct(rv_iibb_pct, 2)})',
                              fmt_dinero(iibb_kg_obj) + '/kg'))
            if rv_cheque_pct > 0:
                filas.append((f'Imp. Cheque ({fmt_pct(rv_cheque_pct, 2)})',
                              fmt_dinero(chq_kg_obj) + '/kg'))
            filas.append((f'Subtotal impuestos+comisiones ({fmt_pct(res["prop"]*100, 2)} s/venta)',
                          fmt_dinero(total_imp_obj) + '/kg'))
            filas.extend([
                ('— TOTALES —', ''),
                ('Precio mínimo (cubre todos los costos, sin ganancia)',
                 fmt_dinero(pmin) + '/kg'),
                (f'Margen objetivo ({fmt_pct(rv_margen_pct, 2)})',
                 fmt_dinero(margen_kg) + '/kg'),
                ('💰 PRECIO OBJETIVO DE VENTA (s/IVA)',
                 fmt_dinero(pobj) + '/kg'),
                (f'PRECIO OBJETIVO c/IVA ({fmt_pct(rv_iva_pct, 2)})',
                 fmt_dinero(res['pobj_iva']) + '/kg'),
            ])
            df_d = pd.DataFrame(filas, columns=['Concepto', 'Valor'])
            st.dataframe(df_d, use_container_width=True, hide_index=True)

            # Totales en valor absoluto para todo el lote
            st.caption(
                f"**Sobre los {fmt_num(res['total_kg'], 1)} kg del lote:** "
                f"Costos directos: {fmt_dinero(res['costo_fijo'] * res['total_kg'])} · "
                f"Impuestos+comisiones: {fmt_dinero(total_imp_obj * res['total_kg'])} · "
                f"Margen: {fmt_dinero(margen_kg * res['total_kg'])} · "
                f"Venta total objetivo: {fmt_dinero(pobj * res['total_kg'])}"
            )

        if 'precio_real' in res:
            st.markdown("---")
            st.markdown("#### 4. Análisis de venta real")
            col_v1, col_v2, col_v3 = st.columns(3)
            col_v1.metric("Costo base/kg", fmt_dinero(res['costo_base_kg']))
            col_v2.metric("Comisiones + tasas", fmt_dinero(res['tasas']))
            col_v3.metric("Costo total/kg", fmt_dinero(res['ct']))

            col_g1, col_g2, col_g3 = st.columns(3)
            color_gk = 'normal' if res['ganancia_kg'] >= 0 else 'inverse'
            col_g1.metric("Ganancia/kg", fmt_dinero(res['ganancia_kg']),
                          delta=fmt_pct(res['rentabilidad'], 2), delta_color=color_gk)
            col_g2.metric("Ganancia total", fmt_dinero(res['ganancia_total']))
            col_g3.metric("Rentabilidad", fmt_pct(res['rentabilidad'], 2))

            if res['rentabilidad'] >= 15:
                st.success(f"🏆 Excelente — {fmt_pct(res['rentabilidad'], 2)} de rentabilidad")
            elif res['rentabilidad'] >= 8:
                st.info(f"👍 Bueno — {fmt_pct(res['rentabilidad'], 2)}")
            elif res['rentabilidad'] >= 3:
                st.warning(f"⚠️ Ajustado — {fmt_pct(res['rentabilidad'], 2)}")
            elif res['rentabilidad'] >= 0:
                st.warning(f"😐 Bajo — {fmt_pct(res['rentabilidad'], 2)}")
            else:
                st.error(f"🔴 Pérdida — {fmt_pct(res['rentabilidad'], 2)}")

    # ══════════════════════════════════════════════════════════════════
    # MODO: VENTA COMBINADA (mismo producto, varias ventas)
    # ══════════════════════════════════════════════════════════════════
    elif modo_rev == 'Venta combinada (mismo producto)':
        st.markdown("#### Venta combinada — mismo producto")
        st.caption("Un único producto vendido en 3 a 5 operaciones distintas (clientes/condiciones). "
                   "El producto es idéntico en todas; las comisiones, IIBB, cheque y precio pueden variar por venta.")

        # ── Producto fijo (igual para todas las ventas) ──
        st.markdown("#### 1. Producto (común a todas las ventas)")
        col_p1, col_p2, col_p3, col_p4 = st.columns(4)
        rc_producto = col_p1.text_input("Producto", value="BONDIOLA", key='rc_prod_nom')
        rc_costo = col_p2.number_input("Costo $/kg s/IVA", value=9650, step=100, key='rc_prod_costo')
        rc_iva_pct = col_p3.number_input("IVA compra (%)", value=10.50, step=0.5,
                                          format="%.2f", key='rc_prod_iva')
        rc_gastos = col_p4.number_input("Gastos op. $/kg", value=1000, step=50, key='rc_prod_gastos')

        # ── Impuestos y financiero comunes ──
        st.markdown("#### 2. Impuestos y financiero (comunes a todos los clientes)")
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        rc_iibb_pct = col_f1.number_input("IIBB (%)", value=4.00, step=0.5,
                                           format="%.2f", key='rc_iibb_com')
        rc_cheque_pct = col_f2.number_input("Imp. Cheque (%)", value=1.20, step=0.1,
                                              format="%.2f", key='rc_chq_com')
        rc_tna_pct = col_f3.number_input("TNA (%)", value=48.00, step=1.0,
                                           format="%.2f", key='rc_tna')
        rc_margen_pct = col_f4.number_input("Margen objetivo (%)", value=15.00, step=0.5,
                                              format="%.2f", key='rc_margen')
        st.caption("_Las comisiones y los días de financiamiento se definen por cliente en cada venta._")

        # ── Ventas (cada una con sus condiciones) ──
        st.markdown("#### 3. Ventas del producto")
        n_ventas = st.slider("Cantidad de ventas", min_value=3, max_value=5, value=3, key='rc_n_ventas')

        ventas_data = []
        for i in range(n_ventas):
            with st.expander(f"🛒 Venta {i+1}", expanded=(i < 3)):
                col_v1, col_v2, col_v3, col_v4 = st.columns(4)
                v_cliente = col_v1.text_input("Cliente / referencia", value=f"Cliente {i+1}", key=f'rc_v_cli_{i}')
                v_kg = col_v2.number_input("Kg de la venta", value=1500.0, step=100.0,
                                            format="%.2f", key=f'rc_v_kg_{i}')
                v_flete = col_v3.number_input("Flete $/kg", value=70, step=10, key=f'rc_v_flete_{i}')
                v_precio_real = col_v4.number_input("Precio venta real $/kg s/IVA",
                                                     value=0, step=100, key=f'rc_v_preal_{i}',
                                                     help="Dejá 0 para usar precio objetivo calculado")

                col_v5, col_v6, col_v7 = st.columns(3)
                v_com1 = col_v5.number_input("Comisión 1 (%)", value=0.00, step=0.5,
                                              format="%.2f", key=f'rc_v_c1_{i}')
                v_com2 = col_v6.number_input("Comisión 2 (%)", value=1.00, step=0.5,
                                              format="%.2f", key=f'rc_v_c2_{i}')
                v_dias = col_v7.number_input("Días financ.", value=10, step=1, key=f'rc_v_dias_{i}')

                ventas_data.append({
                    'cliente': v_cliente, 'kg': v_kg, 'flete': v_flete, 'precio_real': v_precio_real,
                    'com1': v_com1/100, 'com2': v_com2/100, 'dias': v_dias,
                })

        if st.button("🧮 Calcular venta combinada", type='primary', use_container_width=True, key='rc_calc'):
            resultados = []
            venta_total_all = 0
            costo_total_all = 0
            kg_total_all = 0
            ganancia_total_all = 0

            for v in ventas_data:
                r = _calcular_revoleo(
                    v['kg'], rc_costo, rc_iva_pct/100,
                    v['flete'], rc_gastos,
                    v['com1'], v['com2'], rc_iibb_pct/100, rc_cheque_pct/100,
                    rc_tna_pct/100, v['dias'], rc_margen_pct/100,
                    precio_real=v['precio_real'] if v['precio_real'] > 0 else None
                )
                precio_usado = v['precio_real'] if v['precio_real'] > 0 else r['pobj']
                tasas = precio_usado * r['prop']
                ct = r['costo_base_kg'] + tasas + r['cf']
                gk = precio_usado - ct
                gt = gk * r['total_kg']
                venta = precio_usado * r['total_kg']
                costo_v = ct * r['total_kg']
                rent = (gk / ct * 100) if ct > 0 else 0

                resultados.append({
                    'Cliente': v['cliente'],
                    'Kg': r['total_kg'],
                    'Costo $/kg': r['costo_base_kg'],
                    'Precio $/kg': precio_usado,
                    'Precio obj': r['pobj'],
                    'Venta total': venta,
                    'Costo total': costo_v,
                    'Ganancia $/kg': gk,
                    'Ganancia total': gt,
                    'Rentab %': rent,
                })
                venta_total_all += venta
                costo_total_all += costo_v
                kg_total_all += r['total_kg']
                ganancia_total_all += gt

            rent_global = (ganancia_total_all / costo_total_all * 100) if costo_total_all > 0 else 0

            st.markdown("---")
            st.markdown(f"#### Detalle por venta — {rc_producto}")
            df_res = pd.DataFrame(resultados)
            df_show = df_res.copy()
            df_show['Kg'] = df_show['Kg'].apply(lambda x: fmt_num(x, 1))
            df_show['Costo $/kg'] = df_show['Costo $/kg'].apply(fmt_dinero)
            df_show['Precio $/kg'] = df_show['Precio $/kg'].apply(fmt_dinero)
            df_show['Precio obj'] = df_show['Precio obj'].apply(fmt_dinero)
            df_show['Venta total'] = df_show['Venta total'].apply(fmt_dinero)
            df_show['Costo total'] = df_show['Costo total'].apply(fmt_dinero)
            df_show['Ganancia $/kg'] = df_show['Ganancia $/kg'].apply(fmt_dinero)
            df_show['Ganancia total'] = df_show['Ganancia total'].apply(fmt_dinero)
            df_show['Rentab %'] = df_show['Rentab %'].apply(lambda x: fmt_pct(x, 2))
            st.dataframe(df_show, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("#### 📊 Resultado consolidado")
            col_t1, col_t2, col_t3, col_t4 = st.columns(4)
            col_t1.metric("Kg totales", fmt_num(kg_total_all, 1))
            col_t2.metric("Venta total", fmt_dinero(venta_total_all))
            col_t3.metric("Costo total", fmt_dinero(costo_total_all))
            col_color = 'normal' if ganancia_total_all >= 0 else 'inverse'
            col_t4.metric("Ganancia total", fmt_dinero(ganancia_total_all),
                          delta=fmt_pct(rent_global, 2), delta_color=col_color)

            precio_prom_pond = venta_total_all / kg_total_all if kg_total_all > 0 else 0
            costo_prom_pond = costo_total_all / kg_total_all if kg_total_all > 0 else 0
            col_pp1, col_pp2, col_pp3 = st.columns(3)
            col_pp1.metric("Precio prom. ponderado $/kg", fmt_dinero(precio_prom_pond))
            col_pp2.metric("Costo prom. ponderado $/kg", fmt_dinero(costo_prom_pond))
            col_pp3.metric("Rentabilidad global", fmt_pct(rent_global, 2))

            if rent_global >= 15:
                st.success(f"🏆 Excelente — rentabilidad global {fmt_pct(rent_global, 2)}")
            elif rent_global >= 8:
                st.info(f"👍 Buena — {fmt_pct(rent_global, 2)}")
            elif rent_global >= 3:
                st.warning(f"⚠️ Ajustada — {fmt_pct(rent_global, 2)}")
            elif rent_global >= 0:
                st.warning(f"😐 Baja — {fmt_pct(rent_global, 2)}")
            else:
                st.error(f"🔴 Pérdida — {fmt_pct(rent_global, 2)}")

            st.markdown("##### 💡 Aporte de cada venta")
            aportes = []
            for r in resultados:
                aportes.append({
                    'Cliente': r['Cliente'],
                    '% de la venta': r['Venta total'] / venta_total_all * 100 if venta_total_all > 0 else 0,
                    '% de la ganancia': (r['Ganancia total'] / ganancia_total_all * 100)
                                         if ganancia_total_all != 0 else 0,
                    'Rentab individual': r['Rentab %'],
                })
            df_ap = pd.DataFrame(aportes)
            df_ap['% de la venta'] = df_ap['% de la venta'].apply(lambda x: fmt_pct(x, 2))
            df_ap['% de la ganancia'] = df_ap['% de la ganancia'].apply(lambda x: fmt_pct(x, 2))
            df_ap['Rentab individual'] = df_ap['Rentab individual'].apply(lambda x: fmt_pct(x, 2))
            st.dataframe(df_ap, use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════════════════
    # MODO: VENTA MULTI-PRODUCTO (distintos productos, mismo cliente)
    # ══════════════════════════════════════════════════════════════════
    else:
        st.markdown("#### Venta multi-producto — mismo cliente")
        st.caption("Varios productos distintos vendidos al mismo cliente. "
                   "Las condiciones del cliente (comisiones, IIBB, cheque, financiero, margen) son comunes; "
                   "las reglas de cada producto (costo, flete, gastos, precio) varían.")

        # ── Condiciones del cliente (comunes) ──
        st.markdown("#### 1. Condiciones del cliente (comunes a todos los productos)")
        col_c1, col_c2, col_c3, col_c4 = st.columns(4)
        rm_com1_pct = col_c1.number_input("Comisión 1 (%)", value=0.00, step=0.5, format="%.2f", key='rm_com1')
        rm_com2_pct = col_c2.number_input("Comisión 2 (%)", value=1.00, step=0.5, format="%.2f", key='rm_com2')
        rm_iibb_pct = col_c3.number_input("IIBB (%)", value=4.00, step=0.5, format="%.2f", key='rm_iibb')
        rm_cheque_pct = col_c4.number_input("Imp. Cheque (%)", value=1.20, step=0.1, format="%.2f", key='rm_chq')

        col_f1, col_f2, col_f3 = st.columns(3)
        rm_tna_pct = col_f1.number_input("TNA (%)", value=48.00, step=1.0, format="%.2f", key='rm_tna')
        rm_dias = col_f2.number_input("Días financ.", value=10, step=1, key='rm_dias')
        rm_margen_pct = col_f3.number_input("Margen objetivo (%)", value=15.00, step=0.5,
                                              format="%.2f", key='rm_margen')

        # ── Productos ──
        st.markdown("#### 2. Productos vendidos al cliente")
        n_prods = st.slider("Cantidad de productos", min_value=3, max_value=5, value=3, key='rm_n_prod')

        prods_data = []
        for i in range(n_prods):
            with st.expander(f"📦 Producto {i+1}", expanded=(i < 3)):
                col_p1, col_p2, col_p3 = st.columns(3)
                p_nombre = col_p1.text_input("Producto", value=f"Producto {i+1}", key=f'rm_p_nom_{i}')
                p_kg = col_p2.number_input("Kg totales", value=2000.0, step=100.0,
                                            format="%.2f", key=f'rm_p_kg_{i}')
                p_costo = col_p3.number_input("Costo $/kg s/IVA", value=9000, step=100, key=f'rm_p_costo_{i}')

                col_p4, col_p5, col_p6 = st.columns(3)
                p_iva = col_p4.number_input("IVA compra (%)", value=10.50, step=0.5,
                                             format="%.2f", key=f'rm_p_iva_{i}')
                p_flete = col_p5.number_input("Flete $/kg", value=70, step=10, key=f'rm_p_flete_{i}')
                p_gastos = col_p6.number_input("Gastos op. $/kg", value=1000, step=50, key=f'rm_p_gastos_{i}')

                p_preal = st.number_input("Precio venta real $/kg s/IVA",
                                           value=0, step=100, key=f'rm_p_preal_{i}',
                                           help="Dejá 0 para usar precio objetivo calculado")

                prods_data.append({
                    'nombre': p_nombre, 'kg': p_kg,
                    'costo': p_costo, 'iva': p_iva/100, 'flete': p_flete,
                    'gastos': p_gastos, 'precio_real': p_preal,
                })

        if st.button("🧮 Calcular venta multi-producto", type='primary',
                     use_container_width=True, key='rm_calc'):
            resultados = []
            venta_total_all = 0
            costo_total_all = 0
            kg_total_all = 0
            ganancia_total_all = 0

            for p in prods_data:
                r = _calcular_revoleo(
                    p['kg'], p['costo'], p['iva'],
                    p['flete'], p['gastos'],
                    rm_com1_pct/100, rm_com2_pct/100, rm_iibb_pct/100, rm_cheque_pct/100,
                    rm_tna_pct/100, rm_dias, rm_margen_pct/100,
                    precio_real=p['precio_real'] if p['precio_real'] > 0 else None
                )
                precio_usado = p['precio_real'] if p['precio_real'] > 0 else r['pobj']
                tasas = precio_usado * r['prop']
                ct = r['costo_base_kg'] + tasas + r['cf']
                gk = precio_usado - ct
                gt = gk * r['total_kg']
                venta = precio_usado * r['total_kg']
                costo_p = ct * r['total_kg']
                rent = (gk / ct * 100) if ct > 0 else 0

                resultados.append({
                    'Producto': p['nombre'],
                    'Kg': r['total_kg'],
                    'Costo $/kg': r['costo_base_kg'],
                    'Precio $/kg': precio_usado,
                    'Precio obj': r['pobj'],
                    'Venta total': venta,
                    'Costo total': costo_p,
                    'Ganancia $/kg': gk,
                    'Ganancia total': gt,
                    'Rentab %': rent,
                })
                venta_total_all += venta
                costo_total_all += costo_p
                kg_total_all += r['total_kg']
                ganancia_total_all += gt

            rent_global = (ganancia_total_all / costo_total_all * 100) if costo_total_all > 0 else 0

            st.markdown("---")
            st.markdown("#### Detalle por producto")
            df_res = pd.DataFrame(resultados)
            df_show = df_res.copy()
            df_show['Kg'] = df_show['Kg'].apply(lambda x: fmt_num(x, 1))
            df_show['Costo $/kg'] = df_show['Costo $/kg'].apply(fmt_dinero)
            df_show['Precio $/kg'] = df_show['Precio $/kg'].apply(fmt_dinero)
            df_show['Precio obj'] = df_show['Precio obj'].apply(fmt_dinero)
            df_show['Venta total'] = df_show['Venta total'].apply(fmt_dinero)
            df_show['Costo total'] = df_show['Costo total'].apply(fmt_dinero)
            df_show['Ganancia $/kg'] = df_show['Ganancia $/kg'].apply(fmt_dinero)
            df_show['Ganancia total'] = df_show['Ganancia total'].apply(fmt_dinero)
            df_show['Rentab %'] = df_show['Rentab %'].apply(lambda x: fmt_pct(x, 2))
            st.dataframe(df_show, use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("#### 📊 Resultado consolidado del cliente")
            col_t1, col_t2, col_t3, col_t4 = st.columns(4)
            col_t1.metric("Kg totales", fmt_num(kg_total_all, 1))
            col_t2.metric("Venta total", fmt_dinero(venta_total_all))
            col_t3.metric("Costo total", fmt_dinero(costo_total_all))
            col_color = 'normal' if ganancia_total_all >= 0 else 'inverse'
            col_t4.metric("Ganancia total", fmt_dinero(ganancia_total_all),
                          delta=fmt_pct(rent_global, 2), delta_color=col_color)

            precio_prom_pond = venta_total_all / kg_total_all if kg_total_all > 0 else 0
            costo_prom_pond = costo_total_all / kg_total_all if kg_total_all > 0 else 0
            col_pp1, col_pp2, col_pp3 = st.columns(3)
            col_pp1.metric("Precio prom. ponderado $/kg", fmt_dinero(precio_prom_pond))
            col_pp2.metric("Costo prom. ponderado $/kg", fmt_dinero(costo_prom_pond))
            col_pp3.metric("Rentabilidad global", fmt_pct(rent_global, 2))

            if rent_global >= 15:
                st.success(f"🏆 Excelente — rentabilidad global {fmt_pct(rent_global, 2)}")
            elif rent_global >= 8:
                st.info(f"👍 Buena — {fmt_pct(rent_global, 2)}")
            elif rent_global >= 3:
                st.warning(f"⚠️ Ajustada — {fmt_pct(rent_global, 2)}")
            elif rent_global >= 0:
                st.warning(f"😐 Baja — {fmt_pct(rent_global, 2)}")
            else:
                st.error(f"🔴 Pérdida — {fmt_pct(rent_global, 2)}")

            st.markdown("##### 💡 Aporte de cada producto")
            aportes = []
            for r in resultados:
                aportes.append({
                    'Producto': r['Producto'],
                    '% de la venta': r['Venta total'] / venta_total_all * 100 if venta_total_all > 0 else 0,
                    '% de la ganancia': (r['Ganancia total'] / ganancia_total_all * 100)
                                         if ganancia_total_all != 0 else 0,
                    'Rentab individual': r['Rentab %'],
                })
            df_ap = pd.DataFrame(aportes)
            df_ap['% de la venta'] = df_ap['% de la venta'].apply(lambda x: fmt_pct(x, 2))
            df_ap['% de la ganancia'] = df_ap['% de la ganancia'].apply(lambda x: fmt_pct(x, 2))
            df_ap['Rentab individual'] = df_ap['Rentab individual'].apply(lambda x: fmt_pct(x, 2))
            st.dataframe(df_ap, use_container_width=True, hide_index=True)


# ── TAB 5: EXPORTACIÓN CHINA ──
with tab_export:
    st.markdown("### 🌍 Exportación China — Proyección de negocio")
    st.markdown("_Costos reales de INTEGRACIONES CHINA. Editá dólar, USD FOB y split de facturación._")

    from config import COSTOS_EXPORT

    # ══════ SECCIÓN 1: PARÁMETROS EDITABLES ══════
    st.markdown("#### 1. Parámetros del negocio")

    col_a1, col_a2, col_a3 = st.columns(3)
    exp_cabezas = col_a1.number_input("Cabezas", value=222, step=10, key='exp_cab')
    exp_peso_vivo = col_a2.number_input("Peso prom. animal (kg vivo)", value=450, step=10, key='exp_peso')
    exp_rend_faena_pct = col_a3.number_input("Rendimiento faena (%)", value=45.00, step=0.5, format="%.2f", key='exp_rend_f_v2')
    exp_rend_faena = exp_rend_faena_pct / 100

    col_b1, col_b2, col_b3 = st.columns(3)
    exp_precio_hac = col_b1.number_input("Precio hacienda ($/kg vivo)", value=2150, step=50, key='exp_hac')
    exp_faena_kg = col_b2.number_input("Costo faena ($/kg gancho)", value=COSTOS_EXPORT['faena_kg_gancho'], step=5, key='exp_faena')
    exp_insumos_mo = col_b3.number_input("Insumos + MO ($/kg carne)", value=COSTOS_EXPORT['insumos_kg_carne'] + COSTOS_EXPORT['mo_kg_carne'], step=50, key='exp_insmo')

    st.markdown("---")
    st.markdown("#### 2. Tipo de cambio y facturación")

    col_tc1, col_tc2, col_tc3, col_tc4 = st.columns(4)
    exp_tc_ofi = col_tc1.number_input("TC Oficial ($/USD)", value=COSTOS_EXPORT['tc_oficial'], step=10, key='exp_tco')
    exp_tc_blue = col_tc2.number_input("TC Blue ($/USD)", value=COSTOS_EXPORT['tc_blue'], step=10, key='exp_tcb')
    exp_pct_oficial_p = col_tc3.number_input("Cobro aduana oficial (%)",
                                               value=COSTOS_EXPORT['pct_aduana_oficial']*100,
                                               step=5.0, format="%.2f", key='exp_pof_v2')
    exp_pct_oficial = exp_pct_oficial_p / 100
    exp_pct_blue_p = col_tc4.number_input("Cobro fuera blue (%)",
                                            value=COSTOS_EXPORT['pct_aduana_blue']*100,
                                            step=5.0, format="%.2f", key='exp_pbl_v2')
    exp_pct_blue = exp_pct_blue_p / 100

    st.markdown("---")
    st.markdown("#### 3. Precios de venta FOB (USD/kg)")

    col_v1, col_v2, col_v3, col_v4 = st.columns(4)
    exp_usd_carne = col_v1.number_input("USD/kg carne FOB", value=COSTOS_EXPORT['usd_carne_fob'], step=0.05, format="%.2f", key='exp_uc')
    exp_usd_huesos = col_v2.number_input("USD/kg huesos FOB", value=COSTOS_EXPORT['usd_huesos_fob'], step=0.01, format="%.2f", key='exp_uh')
    exp_usd_recortes = col_v3.number_input("USD/kg recortes FOB", value=COSTOS_EXPORT['usd_recortes_fob'], step=0.1, format="%.1f", key='exp_ur')
    exp_reintegro_p = col_v4.number_input("Reintegro export (%)",
                                            value=COSTOS_EXPORT['pct_reintegro']*100,
                                            step=0.05, format="%.2f", key='exp_reint_v2')
    exp_reintegro = exp_reintegro_p / 100

    st.markdown("---")
    st.markdown("#### 4. Costos logísticos fijos (ARS por operación)")
    st.markdown("_Estos vienen de INTEGRACIONES CHINA. Editá si cambiaron._")

    col_l1, col_l2, col_l3 = st.columns(3)
    exp_flete2 = col_l1.number_input("Flete Etapa 2", value=COSTOS_EXPORT['flete_etapa2'], step=50000, key='exp_fl2')
    exp_almacenaje = col_l2.number_input("Almacenaje Dodero", value=COSTOS_EXPORT['almacenaje_dodero'], step=50000, key='exp_alm')
    exp_flete_pto = col_l3.number_input("Flete puerto", value=COSTOS_EXPORT['flete_puerto'], step=50000, key='exp_flp')

    col_l4, col_l5, col_l6 = st.columns(3)
    exp_despachante = col_l4.number_input("Despachante", value=COSTOS_EXPORT['despachante'], step=10000, key='exp_desp')
    exp_docs = col_l5.number_input("Gastos docs SENASA", value=COSTOS_EXPORT['gastos_docs_senasa'], step=10000, key='exp_docs')
    exp_exolgan = col_l6.number_input("Exolgan coordinación", value=COSTOS_EXPORT['exolgan'], step=50000, key='exp_exo')

    col_l7, col_l8 = st.columns(2)
    exp_fwd = col_l7.number_input("FWD gastos locales", value=COSTOS_EXPORT['fwd_gastos_locales'], step=50000, key='exp_fwd')
    exp_consol = col_l8.number_input("Consolidación Dodero", value=COSTOS_EXPORT['consolidacion_dodero'], step=10000, key='exp_cons')

    total_logistica = exp_flete2 + exp_almacenaje + exp_flete_pto + exp_despachante + exp_docs + exp_exolgan + exp_fwd + exp_consol
    st.metric("Total logística por operación", f"${total_logistica:,.0f}")

    st.markdown("---")
    st.markdown("#### 5. Ingresos locales (RIS)")
    col_r1, col_r2 = st.columns(2)
    exp_huesos_ris = col_r1.number_input("Ingreso huesos RIS ($/lote)", value=COSTOS_EXPORT['ingreso_huesos_ris_lote'], step=10000, key='exp_hris')
    exp_grasa_ris = col_r2.number_input("Ingreso grasa RIS ($/lote)", value=COSTOS_EXPORT['ingreso_grasa_ris_lote'], step=50000, key='exp_gris')

    st.markdown("---")

    # ══════ CALCULAR ══════
    if st.button("📊 Calcular proyección", type="primary", use_container_width=True, key='exp_calc'):

        # Volúmenes
        kg_vivo_total = exp_cabezas * exp_peso_vivo
        kg_gancho = kg_vivo_total * exp_rend_faena
        kg_carne_china = kg_gancho * COSTOS_EXPORT['pct_carne_china']
        kg_recortes = kg_gancho * COSTOS_EXPORT['pct_recortes_china']
        kg_huesos_china = kg_gancho * COSTOS_EXPORT['pct_huesos_china']

        # COSTOS
        costo_hacienda = kg_vivo_total * exp_precio_hac
        costo_faena = kg_gancho * exp_faena_kg
        costo_insumos_mo = kg_carne_china * exp_insumos_mo
        costo_total = costo_hacienda + costo_faena + costo_insumos_mo + total_logistica

        # INGRESOS
        # Carne China
        venta_carne_usd = kg_carne_china * exp_usd_carne
        ingreso_carne_oficial = venta_carne_usd * exp_pct_oficial * exp_tc_ofi
        ingreso_carne_blue = venta_carne_usd * exp_pct_blue * exp_tc_blue
        ingreso_carne_total = ingreso_carne_oficial + ingreso_carne_blue

        # Recortes
        venta_recortes_usd = kg_recortes * exp_usd_recortes
        ingreso_recortes = venta_recortes_usd * (exp_pct_oficial * exp_tc_ofi + exp_pct_blue * exp_tc_blue)

        # Huesos China
        venta_huesos_usd = kg_huesos_china * exp_usd_huesos
        ingreso_huesos_china = venta_huesos_usd * (exp_pct_oficial * exp_tc_ofi + exp_pct_blue * exp_tc_blue)

        # RIS local
        ingreso_ris = exp_huesos_ris + exp_grasa_ris

        # Reintegro
        total_fob_usd = venta_carne_usd + venta_recortes_usd + venta_huesos_usd
        reintegro = total_fob_usd * exp_reintegro * exp_tc_ofi

        ingreso_total = ingreso_carne_total + ingreso_recortes + ingreso_huesos_china + ingreso_ris + reintegro

        # RESULTADO
        resultado = ingreso_total - costo_total
        margen = (resultado / ingreso_total * 100) if ingreso_total > 0 else 0
        resultado_kg_gancho = resultado / kg_gancho if kg_gancho > 0 else 0
        resultado_kg_vivo = resultado / kg_vivo_total if kg_vivo_total > 0 else 0
        resultado_cabeza = resultado / exp_cabezas if exp_cabezas > 0 else 0

        # TC ponderado real
        tc_ponderado = (exp_pct_oficial * exp_tc_ofi + exp_pct_blue * exp_tc_blue)

        # ── Mostrar resultado ──
        st.markdown("---")
        st.markdown("### Resultado de la proyección")

        col_r1, col_r2, col_r3 = st.columns(3)
        col_r1.metric("Kg vivo total", f"{kg_vivo_total:,.0f}")
        col_r2.metric("Kg gancho", f"{kg_gancho:,.0f}")
        col_r3.metric("Kg carne China", f"{kg_carne_china:,.0f}")

        col_s1, col_s2, col_s3, col_s4 = st.columns(4)
        col_s1.metric("Ingreso total", f"${ingreso_total:,.0f}")
        col_s2.metric("Costo total", f"${costo_total:,.0f}")
        emoji_res = "🏆" if margen >= 12 else "👍" if margen >= 5 else "⚠️" if margen >= 0 else "🔴"
        col_s3.metric(f"{emoji_res} Resultado", f"${resultado:,.0f}", delta=f"{margen:.1f}%",
                      delta_color="normal" if resultado >= 0 else "inverse")
        col_s4.metric("TC ponderado", f"${tc_ponderado:,.0f}")

        st.markdown("---")

        # Desglose completo
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            st.markdown("##### Costos")
            st.markdown(f"""
| Concepto | Monto |
|---|---|
| Hacienda ({exp_cabezas} cab × {exp_peso_vivo}kg × ${exp_precio_hac}) | ${costo_hacienda:,.0f} |
| Faena ({kg_gancho:,.0f}kg × ${exp_faena_kg}) | ${costo_faena:,.0f} |
| Insumos + MO ({kg_carne_china:,.0f}kg × ${exp_insumos_mo}) | ${costo_insumos_mo:,.0f} |
| Flete Etapa 2 | ${exp_flete2:,.0f} |
| Almacenaje Dodero | ${exp_almacenaje:,.0f} |
| Flete puerto | ${exp_flete_pto:,.0f} |
| Despachante | ${exp_despachante:,.0f} |
| Docs SENASA | ${exp_docs:,.0f} |
| Exolgan | ${exp_exolgan:,.0f} |
| FWD | ${exp_fwd:,.0f} |
| Consolidación | ${exp_consol:,.0f} |
| **TOTAL COSTOS** | **${costo_total:,.0f}** |
            """)

        with col_d2:
            st.markdown("##### Ingresos")
            st.markdown(f"""
| Concepto | USD | ARS |
|---|---|---|
| Carne China ({kg_carne_china:,.0f}kg × USD {exp_usd_carne}) | ${venta_carne_usd:,.0f} | ${ingreso_carne_total:,.0f} |
| → Aduana oficial ({exp_pct_oficial*100:.0f}% × TC {exp_tc_ofi}) | | ${ingreso_carne_oficial:,.0f} |
| → Blue ({exp_pct_blue*100:.0f}% × TC {exp_tc_blue}) | | ${ingreso_carne_blue:,.0f} |
| Recortes ({kg_recortes:,.0f}kg × USD {exp_usd_recortes}) | ${venta_recortes_usd:,.0f} | ${ingreso_recortes:,.0f} |
| Huesos China ({kg_huesos_china:,.0f}kg × USD {exp_usd_huesos}) | ${venta_huesos_usd:,.0f} | ${ingreso_huesos_china:,.0f} |
| Huesos RIS (lote) | — | ${exp_huesos_ris:,.0f} |
| Grasa RIS (lote) | — | ${exp_grasa_ris:,.0f} |
| Reintegro ({exp_reintegro*100:.1f}%) | ${total_fob_usd * exp_reintegro:,.0f} | ${reintegro:,.0f} |
| **TOTAL INGRESOS** | **${total_fob_usd:,.0f}** | **${ingreso_total:,.0f}** |
            """)

        st.markdown("---")
        st.markdown("##### Indicadores por unidad")
        st.markdown(f"""
| Indicador | Valor |
|---|---|
| Resultado por cabeza | ${resultado_cabeza:,.0f} |
| Resultado por kg gancho | ${resultado_kg_gancho:,.0f} |
| Resultado por kg vivo | ${resultado_kg_vivo:,.0f} |
| Rentabilidad / CM | **{margen:.1f}%** |
| FOB total USD | ${total_fob_usd:,.0f} |
| Precio hacienda break-even | ${(ingreso_total - costo_faena - costo_insumos_mo - total_logistica) / kg_vivo_total:,.0f}/kg vivo |
        """)

        if margen >= 12:
            st.success(f"🏆 ÓPTIMO — {margen:.1f}% rentabilidad")
        elif margen >= 5:
            st.info(f"👍 BUENO — {margen:.1f}% rentabilidad")
        elif margen >= 0:
            st.warning(f"⚠️ AJUSTADO — {margen:.1f}% rentabilidad")
        else:
            st.error(f"🔴 PÉRDIDA — {margen:.1f}%")

# ── TAB 5: HISTORIAL / INTELIGENCIA ──
with tab_historial:
    st.markdown("### 📈 Inteligencia del negocio")

    historial_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'ROMANEOS', 'historial_romaneos.json'
    )

    if not os.path.exists(historial_path):
        st.info("No se encontró el archivo de historial. Se creará al generar el primer análisis.")
    else:
        with open(historial_path, 'r') as f:
            historial = json.load(f)

        if not historial:
            st.info("No hay historial de romaneos aún.")
        else:
            df_raw = pd.DataFrame(historial)

            # Asegurar columnas numéricas ANTES de deduplicar
            for col in ['kg_entrada', 'kg_carne', 'rendimiento_pct', 'precio_compra',
                        'pct_amarilla', 'rend_vs_obj', 'medias', 'costo_hacienda',
                        'margen_pct', 'cm', 'ingreso_bruto', 'costo_total',
                        'precio_venta_promedio']:
                if col in df_raw.columns:
                    df_raw[col] = pd.to_numeric(df_raw[col], errors='coerce')

            # ══════ DEDUPLICAR mejorado ══════
            # Criterio: archivo_original (si existe) o archivo
            # Preferir registros más completos (con más columnas pobladas)
            def _key(row):
                ao = row.get('archivo_original') if 'archivo_original' in row else None
                if pd.notna(ao) and ao:
                    return str(ao).strip().upper().replace('.PDF', '')
                return str(row.get('archivo', '')).strip().upper().replace('.PDF', '')

            df_raw['_dedup_key'] = df_raw.apply(_key, axis=1)
            # Score = cantidad de celdas no-NaN
            df_raw['_completeness'] = df_raw.notna().sum(axis=1)
            # Ordenar por completeness desc, quedarse con el más completo
            df_raw = df_raw.sort_values('_completeness', ascending=False)
            df_hist = df_raw.drop_duplicates(subset='_dedup_key', keep='first').copy()
            df_hist = df_hist.drop(columns=['_dedup_key', '_completeness'])

            # Descartar registros sin kg_entrada (están rotos)
            if 'kg_entrada' in df_hist.columns:
                df_hist = df_hist[df_hist['kg_entrada'].notna() & (df_hist['kg_entrada'] > 0)]

            # Calcular costo_hacienda faltante
            if 'costo_hacienda' not in df_hist.columns:
                df_hist['costo_hacienda'] = df_hist['kg_entrada'] * df_hist['precio_compra']
            else:
                mask = df_hist['costo_hacienda'].isna() & df_hist['precio_compra'].notna()
                df_hist.loc[mask, 'costo_hacienda'] = (
                    df_hist.loc[mask, 'kg_entrada'] * df_hist.loc[mask, 'precio_compra'])

            n_total = len(df_hist)
            # Sumas ignorando NaN
            kg_total = df_hist['kg_entrada'].fillna(0).sum()
            kg_carne_total = df_hist['kg_carne'].fillna(0).sum() if 'kg_carne' in df_hist.columns else 0
            medias_total = df_hist['medias'].fillna(0).sum() if 'medias' in df_hist.columns else 0

            # ══════ RESUMEN GENERAL ══════
            st.markdown("#### Resumen acumulado")
            st.caption(f"Sobre {n_total} romaneos únicos (de {len(historial)} registros totales)")
            col1, col2, col3, col4, col5 = st.columns(5)
            col1.metric("Romaneos", n_total)
            col2.metric("Medias totales", f"{medias_total:,.0f}".replace(',', '.'))
            col3.metric("Kg entrada", f"{kg_total:,.0f}".replace(',', '.'))
            col4.metric("Kg carne", f"{kg_carne_total:,.0f}".replace(',', '.'))
            rend_pond = (kg_carne_total / kg_total * 100) if kg_total > 0 else 0
            col5.metric("Rend. ponderado", f"{rend_pond:.1f}%".replace('.', ','))

            col6, col7, col8 = st.columns(3)
            costo_hac_total = df_hist['costo_hacienda'].fillna(0).sum() if 'costo_hacienda' in df_hist.columns else 0
            precio_prom_pond = (costo_hac_total / kg_total) if kg_total > 0 else 0
            col6.metric("Precio compra prom. ponderado", f"${precio_prom_pond:,.0f}/kg".replace(',', '.'))
            col7.metric("Inversión hacienda", f"${costo_hac_total:,.0f}".replace(',', '.'))
            if 'pct_amarilla' in df_hist.columns:
                amarilla_prom = df_hist['pct_amarilla'].dropna().mean()
                if pd.notna(amarilla_prom):
                    col8.metric("Amarilla promedio", f"{amarilla_prom:.1f}%".replace('.', ','))

            st.markdown("---")

            # ══════ 1. RENDIMIENTO POR CATEGORÍA ══════
            st.markdown("#### 1. Rendimiento real por categoría vs objetivo")

            from config import REND_OBJETIVO
            rend_obj_std = REND_OBJETIVO.get('Standard', {})

            if 'categoria' in df_hist.columns:
                cats = df_hist.groupby('categoria').agg({
                    'kg_entrada': 'sum',
                    'kg_carne': 'sum',
                    'medias': 'sum',
                    'rendimiento_pct': 'mean',
                    'pct_amarilla': 'mean',
                    'precio_compra': 'mean',
                }).round(2)

                cats['Rend. ponderado %'] = (cats['kg_carne'] / cats['kg_entrada'] * 100).round(2)
                cats['Objetivo %'] = cats.index.map(lambda c: rend_obj_std.get(c, 0.66) * 100)
                cats['Desvío %'] = (cats['Rend. ponderado %'] - cats['Objetivo %']).round(2)
                cats['Romaneos'] = df_hist.groupby('categoria').size()

                cats_display = cats[['Romaneos', 'medias', 'kg_entrada', 'kg_carne',
                                     'Rend. ponderado %', 'Objetivo %', 'Desvío %',
                                     'pct_amarilla', 'precio_compra']].copy()
                cats_display.columns = ['Romaneos', 'Medias', 'Kg entrada', 'Kg carne',
                                        'Rend. real %', 'Objetivo %', 'Desvío %',
                                        'Amarilla %', 'Precio compra prom']

                def color_desvio(val):
                    if val >= 1: return 'background-color: #E8F5E9; color: #2E7D32'
                    elif val >= -1: return 'background-color: #FFF3E0; color: #E65100'
                    else: return 'background-color: #FCE4EC; color: #C62828'

                st.dataframe(
                    cats_display.style.applymap(color_desvio, subset=['Desvío %']),
                    use_container_width=True
                )

                # Gráfico rendimiento por categoría
                chart_rend = cats[['Rend. ponderado %', 'Objetivo %']].copy()
                st.bar_chart(chart_rend)

            st.markdown("---")

            # ══════ 2. ANÁLISIS DE EQUILIBRIO ══════
            st.markdown("#### 2. Precio máximo de compra por categoría")
            st.markdown(
                "_¿Hasta cuánto puedo pagar por kg de media res para obtener margen positivo, "
                "bueno u óptimo? Basado en rendimientos y precios de venta reales._"
            )

            from config import COSTOS_PERFILES
            costos_std = COSTOS_PERFILES['Standard']
            costo_var_kg = costos_std['mo'] + costos_std['insumos'] + costos_std['flete'] + costos_std['senasa']
            iibb = costos_std['iibb']

            # Precio venta promedio real (de historial si existe)
            if 'precio_venta_promedio' in df_hist.columns:
                pvp = df_hist['precio_venta_promedio'].mean()
            else:
                pvp = 13000  # estimado

            equilibrio_data = []
            for cat in df_hist['categoria'].unique():
                df_cat = df_hist[df_hist['categoria'] == cat]
                rend_real = df_cat['kg_carne'].sum() / df_cat['kg_entrada'].sum() if df_cat['kg_entrada'].sum() > 0 else 0.66
                amarilla = df_cat['pct_amarilla'].mean() / 100 if 'pct_amarilla' in df_cat.columns else 0

                # pvp ajustado por amarilla (la amarilla vende a ~$10.500, baja el promedio)
                pvp_cat = pvp * (1 - amarilla) + 9500 * amarilla
                ingreso_neto_por_kg_entrada = rend_real * pvp_cat * (1 - iibb)
                costo_var_por_kg_entrada = rend_real * costo_var_kg

                # Precio máx compra = ingreso neto - costo variable - margen deseado
                for margen_target, label in [(0, 'Equilibrio (0%)'), (0.03, 'Regular (3%)'),
                                              (0.08, 'Bueno (8%)'), (0.15, 'Óptimo (15%)')]:
                    ingreso_obj = ingreso_neto_por_kg_entrada
                    precio_max = ingreso_obj * (1 - margen_target) - costo_var_por_kg_entrada
                    equilibrio_data.append({
                        'Categoría': cat,
                        'Rend. real': f"{rend_real*100:.1f}%",
                        'Amarilla %': f"{amarilla*100:.1f}%",
                        'Margen objetivo': label,
                        'Precio máx compra $/kg': round(precio_max) if not (precio_max != precio_max) else 0,
                    })

            df_eq = pd.DataFrame(equilibrio_data)
            # Pivot
            df_pivot = df_eq.pivot(index='Categoría', columns='Margen objetivo', values='Precio máx compra $/kg')
            cols_order = ['Equilibrio (0%)', 'Regular (3%)', 'Bueno (8%)', 'Óptimo (15%)']
            df_pivot = df_pivot[[c for c in cols_order if c in df_pivot.columns]]
            st.dataframe(df_pivot, use_container_width=True)

            # Comparativa directa: ¿conviene pagar más por mejor animal?
            st.markdown("**¿Conviene comprar más caro un mejor animal?**")
            st.markdown(
                "_Un novillo/novillito rinde más y tiene mejor terneza (mayor precio de venta potencial), "
                "pero cuesta más. ¿Cuánto más puedo pagar?_"
            )

            # Construir tabla comparativa
            comp_data = []
            for cat in df_hist['categoria'].unique():
                df_cat = df_hist[df_hist['categoria'] == cat]
                kg_ent = df_cat['kg_entrada'].sum()
                kg_car = df_cat['kg_carne'].sum() if 'kg_carne' in df_cat.columns else 0
                rend = (kg_car / kg_ent * 100) if kg_ent > 0 else 0
                precio_real = df_cat['precio_compra'].mean()
                amarilla = df_cat['pct_amarilla'].mean() if 'pct_amarilla' in df_cat.columns else 0

                pvp_cat = df_cat['precio_venta_promedio'].mean() if 'precio_venta_promedio' in df_cat.columns else pvp

                # Ingreso neto por kg entrada (lo que genera cada categoría)
                rend_dec = rend / 100
                ingreso_neto_kg = rend_dec * pvp_cat * (1 - iibb)
                costo_var_kg_ent = rend_dec * costo_var_kg

                # Precio máximo para 8% margen
                precio_max_8 = ingreso_neto_kg * (1 - 0.08) - costo_var_kg_ent
                # Precio máximo para 10%
                precio_max_10 = ingreso_neto_kg * (1 - 0.10) - costo_var_kg_ent

                diff_8 = precio_max_8 - precio_real
                n_romaneos = len(df_cat)

                comp_data.append({
                    'Categoría': cat,
                    'Romaneos': n_romaneos,
                    'Rend. real': f"{rend:.1f}%",
                    'Amarilla': f"{amarilla:.1f}%",
                    'Compra real': f"${precio_real:,.0f}",
                    'Máx p/8% CM': f"${precio_max_8:,.0f}",
                    'Máx p/10% CM': f"${precio_max_10:,.0f}",
                    'Margen $/kg': f"${diff_8:+,.0f}",
                    '_precio_real': precio_real,
                    '_precio_max_8': precio_max_8,
                    '_rend': rend,
                })

            if len(comp_data) > 0:
                df_comp = pd.DataFrame(comp_data)
                display_cols = ['Categoría', 'Romaneos', 'Rend. real', 'Amarilla',
                                'Compra real', 'Máx p/8% CM', 'Máx p/10% CM', 'Margen $/kg']
                st.dataframe(df_comp[display_cols], use_container_width=True, hide_index=True)

                # Análisis narrativo
                if len(comp_data) >= 2:
                    best = max(comp_data, key=lambda x: x['_precio_max_8'])
                    worst = min(comp_data, key=lambda x: x['_precio_max_8'])

                    diff_precio_max = best['_precio_max_8'] - worst['_precio_max_8']
                    diff_rend = best['_rend'] - worst['_rend']

                    st.markdown(f"""
**Conclusión:**
- **{best['Categoría']}** permite pagar hasta **${best['_precio_max_8']:,.0f}/kg** y mantener 8% CM
  (rinde {best['Rend. real']}, amarilla {best['Amarilla']})
- **{worst['Categoría']}** solo permite hasta **${worst['_precio_max_8']:,.0f}/kg** para el mismo 8%
- La diferencia es **${diff_precio_max:,.0f}/kg** — eso es lo máximo que podés pagar de más
  por {best['Categoría']} vs {worst['Categoría']} y seguir con el mismo margen
- Si el mercado ofrece {best['Categoría']} a menos de ${diff_precio_max:,.0f}/kg más que {worst['Categoría']},
  **conviene comprar {best['Categoría']}** (mejor terneza + mismo o mejor margen)
                    """)

            st.markdown("---")

            # ══════ 3. IMPACTO AMARILLA ══════
            st.markdown("#### 3. Impacto de la amarilla en el resultado")

            if 'pct_amarilla' in df_hist.columns:
                # Clasificar en rangos
                df_am = df_hist[['archivo', 'pct_amarilla', 'rendimiento_pct', 'precio_compra']].copy()
                if 'margen_pct' in df_hist.columns:
                    df_am['margen_pct'] = df_hist['margen_pct']
                elif 'precio_venta_promedio' in df_hist.columns:
                    df_am['pvp'] = df_hist['precio_venta_promedio']

                df_am['Rango amarilla'] = pd.cut(
                    df_am['pct_amarilla'],
                    bins=[0, 5, 15, 30, 100],
                    labels=['< 5%', '5-15%', '15-30%', '> 30%']
                )

                am_impact = df_am.groupby('Rango amarilla', observed=True).agg({
                    'rendimiento_pct': 'mean',
                    'pct_amarilla': ['mean', 'count'],
                }).round(2)
                am_impact.columns = ['Rend. promedio %', 'Amarilla prom %', 'Romaneos']

                if 'margen_pct' in df_hist.columns:
                    am_margen = df_am.groupby('Rango amarilla', observed=True)['margen_pct'].mean().round(2)
                    am_impact['Margen prom %'] = am_margen

                st.dataframe(am_impact, use_container_width=True)
                st.caption(
                    "A mayor % de amarilla, el precio de venta promedio baja "
                    "(la amarilla vende a $10.500/kg vs cortes a ~$14.000/kg). "
                    "Romaneos con >15% amarilla castigan fuerte el margen."
                )

            st.markdown("---")

            # ══════ 4. RECURRENCIAS ══════
            st.markdown("#### 4. Patrones y recurrencias")

            # Rendimientos atípicos
            if 'rend_vs_obj' in df_hist.columns:
                bajo_obj = df_hist[df_hist['rend_vs_obj'] < -2]
                sobre_obj = df_hist[df_hist['rend_vs_obj'] > 3]

                if not bajo_obj.empty:
                    st.markdown(f"**Romaneos muy por debajo del objetivo** ({len(bajo_obj)}):")
                    for _, r in bajo_obj.iterrows():
                        st.markdown(
                            f"- {r.get('archivo', '?')} — {r.get('categoria', '?')} — "
                            f"Rend: {r['rendimiento_pct']:.1f}% (obj: {r.get('rend_objetivo_pct', '?')}%) — "
                            f"Desvío: {r['rend_vs_obj']:+.1f}% — Amarilla: {r.get('pct_amarilla', 0):.1f}%"
                        )

                if not sobre_obj.empty:
                    st.markdown(f"**Romaneos muy por encima del objetivo** ({len(sobre_obj)}):")
                    for _, r in sobre_obj.iterrows():
                        st.markdown(
                            f"- {r.get('archivo', '?')} — {r.get('categoria', '?')} — "
                            f"Rend: {r['rendimiento_pct']:.1f}% (obj: {r.get('rend_objetivo_pct', '?')}%) — "
                            f"Desvío: {r['rend_vs_obj']:+.1f}% — Amarilla: {r.get('pct_amarilla', 0):.1f}%"
                        )

            # Correlación peso promedio media vs rendimiento
            if 'peso_promedio_media' in df_hist.columns:
                df_peso = df_hist[df_hist['peso_promedio_media'].notna() & (df_hist['peso_promedio_media'] > 0)]
                if len(df_peso) > 3:
                    st.markdown("**Peso promedio de media res vs rendimiento:**")
                    chart_peso = df_peso[['peso_promedio_media', 'rendimiento_pct']].copy()
                    chart_peso.columns = ['Peso prom. media (kg)', 'Rendimiento %']
                    st.scatter_chart(chart_peso, x='Peso prom. media (kg)', y='Rendimiento %')

            # Tendencia temporal
            st.markdown("**Evolución temporal del rendimiento:**")
            df_time = df_hist.copy()
            if 'fecha' in df_time.columns:
                df_time['fecha_dt'] = pd.to_datetime(df_time['fecha'], format='%d/%m/%Y', errors='coerce')
                df_time = df_time.sort_values('fecha_dt')
                if df_time['fecha_dt'].notna().sum() > 1:
                    chart_time = df_time[['fecha_dt', 'rendimiento_pct']].dropna().set_index('fecha_dt')
                    st.line_chart(chart_time)

            st.markdown("---")

            # ══════ 5. INSIGHTS AUTOMÁTICOS ══════
            st.markdown("#### 5. Insights")

            insights = []

            # Mejor y peor rendimiento
            if 'rendimiento_pct' in df_hist.columns and len(df_hist) > 1:
                best = df_hist.loc[df_hist['rendimiento_pct'].idxmax()]
                worst = df_hist.loc[df_hist['rendimiento_pct'].idxmin()]
                insights.append(
                    f"**Mejor rendimiento:** {best.get('archivo', '?')} — "
                    f"{best['rendimiento_pct']:.1f}% ({best.get('categoria', '?')}, "
                    f"amarilla {best.get('pct_amarilla', 0):.1f}%)"
                )
                insights.append(
                    f"**Peor rendimiento:** {worst.get('archivo', '?')} — "
                    f"{worst['rendimiento_pct']:.1f}% ({worst.get('categoria', '?')}, "
                    f"amarilla {worst.get('pct_amarilla', 0):.1f}%)"
                )

            # Relación amarilla-rendimiento
            if 'pct_amarilla' in df_hist.columns:
                alta_am = df_hist[df_hist['pct_amarilla'] > 20]
                baja_am = df_hist[df_hist['pct_amarilla'] <= 5]
                if not alta_am.empty and not baja_am.empty:
                    rend_alta = alta_am['rendimiento_pct'].mean()
                    rend_baja = baja_am['rendimiento_pct'].mean()
                    insights.append(
                        f"**Efecto amarilla:** Romaneos con >20% amarilla promedian "
                        f"{rend_alta:.1f}% rend. vs {rend_baja:.1f}% con <5% amarilla. "
                        f"{'El rendimiento sube pero el precio de venta baja.' if rend_alta > rend_baja else 'Menor rendimiento y menor precio.'}"
                    )

            # Precio compra vs margen
            if 'precio_compra' in df_hist.columns and 'margen_pct' in df_hist.columns:
                corr = df_hist[['precio_compra', 'margen_pct']].corr().iloc[0, 1]
                if abs(corr) > 0.3:
                    sentido = "más caro compra → menor margen" if corr < 0 else "más caro compra → mayor margen"
                    insights.append(f"**Correlación precio/margen:** {sentido} (r={corr:.2f})")

            # Categoría más rentable
            if 'categoria' in df_hist.columns and 'margen_pct' in df_hist.columns:
                cat_margen = df_hist.dropna(subset=['margen_pct']).groupby('categoria')['margen_pct'].mean()
                if len(cat_margen) >= 2:
                    mejor_cat = cat_margen.idxmax()
                    peor_cat = cat_margen.idxmin()
                    if mejor_cat != peor_cat:
                        insights.append(
                            f"**Categoría más rentable:** {mejor_cat} ({cat_margen[mejor_cat]:.1f}% margen prom.) "
                            f"vs {peor_cat} ({cat_margen[peor_cat]:.1f}%)"
                        )

            for ins in insights:
                st.markdown(f"- {ins}")

            if not insights:
                st.caption("Se generarán insights cuando haya más romaneos procesados.")

            st.markdown("---")

            # ══════ TABLA DETALLE ══════
            with st.expander("📋 Tabla detalle completa", expanded=False):
                cols_display = [c for c in [
                    'archivo', 'fecha', 'categoria', 'medias',
                    'peso_promedio_media', 'kg_entrada', 'kg_carne',
                    'rendimiento_pct', 'rend_vs_obj', 'precio_compra',
                    'costo_hacienda', 'pct_amarilla'
                ] if c in df_hist.columns]
                st.dataframe(df_hist[cols_display], use_container_width=True, hide_index=True)

            # Botón para limpiar historial de duplicados y guardar
            if st.button("🧹 Limpiar duplicados y guardar", key='clean_hist'):
                clean = df_hist.to_dict('records')
                with open(historial_path, 'w') as f:
                    json.dump(clean, f, indent=2, ensure_ascii=False)
                st.success(f"✅ Historial limpiado: {len(clean)} registros únicos (de {len(historial)} originales)")
                st.rerun()

# ══════════════════════════════════════════════════════════════════════
# TAB 9: COSTOS OPERATIVOS (Frigorífico + Insumos — proyectado/teórico vs real)
# ══════════════════════════════════════════════════════════════════════
with tab_costos:
    st.markdown("### 🏭 Costos Operativos — Proyectado/Teórico vs Real")
    st.caption("El proyectado sale de los romaneos incluidos; el real se lee en vivo "
               "de Drive. Para comparar bien, cargá el **mes completo** de romaneos.")

    import frigorifico as _frig
    import frigorifico_real as _frigR
    import insumos as _ins
    import insumos_real as _insR
    import historial_costos as _hc

    _snap = {}  # foto del mes para el historial de pricing

    _incluidos = [p for p in st.session_state.get('parsed_files', [])
                  if 'error' not in p and not p.get('excluido')]

    if not _incluidos:
        st.info("Cargá romaneos en la solapa **📤 Cargar archivos** para calcular el proyectado.")
    else:
        _meses_stock = st.number_input("Meses de stock congelado (proyectado congelado)",
                                       value=1, min_value=0, step=1, key='costos_meses_stock')

        import diagnostico as _diag
        _kg_carne = sum(_frig._kg_carne(_p) for _p in _incluidos)
        st.caption(f"Base del mes cargado: {len(_incluidos)} romaneos · {_kg_carne:,.0f} kg carne")

        # ─── FRIGORÍFICO ───
        st.markdown("#### ❄️ Frigorífico — Proyectado vs Real ($/kg carne)")
        _proy = {'vep': 0.0, 'cuarteo': 0.0, 'despostada': 0.0, 'congelado': 0.0,
                 'cabezas': 0.0, 'medias': 0.0, 'kg_carne': 0.0, 'kg_congelado': 0.0}
        for _p in _incluidos:
            _r = _frig.costo_proyectado(_p, meses_stock=_meses_stock)
            for _k in _proy:
                _proy[_k] += _r.get(_k, 0)
        st.caption("Comparación por **cantidad física**: a igual cabezas/medias/kg, igual costo. "
                   "Si la cantidad facturada difiere de la de los romaneos, se levanta la alerta.")
        try:
            _, _meses_frig = _frigR.cargar_real_drive(CREDENTIALS_PATH)
            if _meses_frig:
                _op = {d['nombre']: d for d in _meses_frig.values()}
                _sel = st.selectbox("Mes real (frigorífico)", list(_op.keys()),
                                    index=len(_op) - 1, key='costos_mes_frig')
                _rf = _op[_sel]
                _alf = _diag.diagnostico_frigorifico(_proy, _rf, _kg_carne)
                _snap['mes'] = _sel
                _snap['frigo_kg'] = sum(a['real_kg'] for a in _alf)
                st.dataframe(pd.DataFrame([{
                    'Servicio': a['concepto'],
                    'Facturado': f"{a['cant_real']:,.0f} {a['unidad']}",
                    'Romaneos': f"{a['cant_proy']:,.0f} {a['unidad']}",
                    'Δ cantidad': f"{a['cant_desvio']:+,.0f} ({a['cant_desvio_pct']:+.0f}%)",
                    'Real $/kg': f"${a['real_kg']:,.0f}",
                    'Alerta': a['severidad'],
                } for a in _alf]), hide_index=True, use_container_width=True)
                for a in _alf:
                    _msg = (f"**{a['titulo']}**  \n{a['detalle']}"
                            f"  \n💡 {a['recomendacion']}")
                    if a['box'] == 'error':
                        st.error(_msg)          # te facturan de más → ROJO
                    elif a['box'] == 'success':
                        st.success(_msg)        # a tu favor → VERDE
        except Exception as _e:
            st.warning(f"No pude leer el real del frigorífico desde Drive: {_e}")

        st.markdown("---")
        # ─── INSUMOS ───
        st.markdown("#### 📦 Insumos — Teórico vs Real ($/kg carne)")
        _teo = _ins.costo_teorico(_incluidos)
        st.caption(f"Teórico por pieza = ${_teo['costo_por_kg']:,.0f}/kg  ·  piezas: "
                   f"chica {_teo['por_categoria']['chica']}, grande {_teo['por_categoria']['grande']}, "
                   f"hueso {_teo['por_categoria']['hueso']}, sin bolsa {_teo['por_categoria']['sin_bolsa']}")
        st.caption("Chequeo: 1 bolsa + 1 etiqueta por pieza. Consumo real = stock inicial + compras "
                   "del mes − stock actual. Si consumiste más que las piezas, deberías tener más stock.")
        try:
            _, _, _real_ins = _insR.cargar_insumos_real_drive(CREDENTIALS_PATH)
            _kgi = _teo['kg'] or _kg_carne
            _ali = _diag.diagnostico_insumos(_teo, _real_ins, _kgi)
            _snap['insumos_kg'] = sum(a['real_kg'] for a in _ali)
            st.dataframe(pd.DataFrame([{
                'Insumo': a['concepto'],
                'Consumidas': f"{a['real_und']:,.0f}",
                'Esperadas (1/pieza)': f"{a['esperado_und']:,.0f}",
                'Ratio': f"{a['ratio']:.2f}×",
                'Real $/kg': f"${a['real_kg']:,.0f}",
                'Alerta': a['severidad'],
            } for a in _ali]), hide_index=True, use_container_width=True)
            for a in _ali:
                _msg = f"**{a['titulo']}**  \n{a['detalle']}  \n💡 {a['recomendacion']}"
                if a['box'] == 'error':
                    st.error(_msg)
                elif a['box'] == 'warning':
                    st.warning(_msg)
                elif a['box'] == 'success':
                    st.success(_msg)
        except Exception as _e:
            st.warning(f"No pude leer el real de insumos desde Drive: {_e}")


    # ─── FLETES (independiente de los romaneos: lee todos los viajes del mes) ───
    st.markdown("---")
    st.markdown("#### 🚚 Fletes — eficiencia por segmento")
    st.caption("Cada viaje se compara contra el benchmark de su segmento "
               "(PEYA $60-70 · Corta $100-150 · Larga $600-800). Deficiente = arriba de su benchmark.")
    try:
        import fletes as _flt
        _vf, _cf, _af = _flt.cargar_fletes_drive(CREDENTIALS_PATH)
        _snap['fletes_kg'] = _af['costo_kg_prom']
        _cs = st.columns(4)
        _cs[0].metric("Kg movidos", f"{_af['kg_total']:,.0f}")
        _cs[1].metric("Costo prom (c/comisión)", f"${_af['costo_kg_prom']:,.0f}/kg")
        _cs[2].metric("Gestión (comisiones)", f"+${_af['comision_por_kg']:,.0f}/kg")
        _cs[3].metric("Sobrecosto vs benchmark", f"${_af['sobrecosto_total']:,.0f}")

        _seg_rows = [{'Segmento': _seg, 'Kg': f"{_d['kg']:,.0f}", '% vol': f"{_d['pct_kg']:.0f}%",
                      'Real $/kg': f"${_d['real_kg']:,.0f}", 'Benchmark': _d['benchmark'],
                      'Gap $/kg': f"${_d['gap_kg']:+,.0f}", 'Sobrecosto': f"${_d['sobrecosto']:,.0f}"}
                     for _seg, _d in _af['por_segmento'].items()]
        st.dataframe(pd.DataFrame(_seg_rows), hide_index=True, use_container_width=True)

        st.markdown(f"**Viajes deficientes: {_af['n_deficientes']}** (arriba del benchmark de su segmento)")
        if _af['deficientes']:
            _def_rows = [{'Fecha': _v['fecha'], 'Proveedor': _v['proveedor'], 'Segmento': _v['segmento'],
                          'Tipo': _v['tipo'], 'Kg': f"{_v['kg']:,.0f}", '$/kg': f"${_v['costo_kg']:,.0f}",
                          'Bench': f"${_v['benchmark']:,.0f}", 'Sobrecosto': f"${_v['exceso_kg'] * _v['kg']:,.0f}"}
                         for _v in _af['deficientes'][:40]]
            st.dataframe(pd.DataFrame(_def_rows), hide_index=True, use_container_width=True)
        import diagnostico as _diag2
        for _a in _diag2.resumen_fletes(_af):
            with st.expander(f"{_a['severidad']} · {_a['concepto']} — real ${_a['real_kg']:,.0f}/kg "
                             f"vs benchmark ${_a['benchmark']:,.0f} · sobrecosto ${_a['desvio_total']:,.0f}"):
                st.markdown(f"**¿Por qué?** {_a['causa']}")
                st.markdown(f"**Qué hacer / dónde investigar:** {_a['recomendacion']}")
    except Exception as _e:
        st.warning(f"No pude leer los fletes desde Drive: {_e}")


    # ─── HISTORIAL DE COSTOS (para pricing futuro) ───
    st.markdown("---")
    st.markdown("#### 📈 Historial de costos por mes — para pricing futuro")
    st.caption("Guardá cómo cerró el $/kg de cada mes. Con la serie estimás el costo del mes "
               "que viene (con inflación) para fijar precios con anticipación.")
    _histc = _hc.cargar(HIST_COSTOS_PATH)
    _cH = st.columns([2, 1, 2])
    _mes_lbl = _cH[0].text_input("Mes a guardar (ej: 2026-06)", value=_snap.get('mes', ''), key='hc_mes')
    if _cH[1].button("💾 Guardar mes", key='hc_save'):
        _reg = {'mes': _mes_lbl, 'frigo_kg': round(_snap.get('frigo_kg', 0), 1),
                'insumos_kg': round(_snap.get('insumos_kg', 0), 1),
                'fletes_kg': round(_snap.get('fletes_kg', 0), 1)}
        _histc = _hc.guardar_mes(HIST_COSTOS_PATH, _reg)
        st.success(f"Guardado {_mes_lbl}. Frigo ${_reg['frigo_kg']:,.0f} · "
                   f"Insumos ${_reg['insumos_kg']:,.0f} · Fletes ${_reg['fletes_kg']:,.0f}/kg")
    if _histc:
        st.dataframe(pd.DataFrame([{
            'Mes': h['mes'], 'Frigo $/kg': f"${h.get('frigo_kg', 0):,.0f}",
            'Insumos $/kg': f"${h.get('insumos_kg', 0):,.0f}",
            'Fletes $/kg': f"${h.get('fletes_kg', 0):,.0f}",
            'Total oper. $/kg': f"${h.get('total_op_kg', 0):,.0f}",
        } for h in _histc]), hide_index=True, use_container_width=True)
        _cP = st.columns([1, 1, 2])
        _infl = _cP[0].number_input("Inflación estim. % próx. mes", value=5.0, step=0.5, key='hc_infl')
        _met = _cP[1].selectbox("Método", ['ultimo', 'tendencia'], key='hc_met')
        _proj = _hc.proyeccion(_histc, _infl, _met)
        if _proj:
            st.info(f"**Proyección próximo mes** (base {_proj['base_mes']}, método {_met}, +{_infl:.1f}%): "
                    f"Frigo ${_proj['frigo_kg']:,.0f} · Insumos ${_proj['insumos_kg']:,.0f} · "
                    f"Fletes ${_proj['fletes_kg']:,.0f} · **Total operativo ${_proj['total_op_kg']:,.0f}/kg** "
                    f"→ usalo como piso de costo para el pricing del mes que viene.")
    else:
        st.caption("Todavía no hay meses guardados. Cargá un mes, revisá los costos y apretá 'Guardar mes'.")


# ══════════════════════════════════════════════════════════════════════
# FOOTER
# ══════════════════════════════════════════════════════════════════════
st.markdown("---")
st.markdown(
    '<p style="text-align: center; color: #999; font-size: 0.8rem;">'
    'Romaneo v2.1 — TF Carnes S.A. — Desarrollado con Streamlit</p>',
    unsafe_allow_html=True
)
