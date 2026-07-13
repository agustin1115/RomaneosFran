"""
excel_builder.py — Genera Excel de análisis de romaneo con soporte multi-calidad.
Evolución de build_analisis.py v4 con perfiles Búfalo, Premium Black y Exportación.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from config import (
    GRUPOS_POR_CALIDAD, CORTES_CAROS, REND_OBJETIVO,
    COSTOS_PERFILES, AMARILLA_CONTRAMARCAS, AMARILLA_PRECIO_DEFAULT,
    CONTRAMARCA_MAP, PRECIOS_FIJOS_RECORTE, BUBALINO_PRECIO_DEFAULT,
)

# ── Styles ──
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SEC_FILL = PatternFill('solid', fgColor='D6E4F0')
SEC_FONT = Font(name='Arial', bold=True, size=10)
DATA = Font(name='Arial', size=10)
BLUE = Font(name='Arial', size=10, color='0000FF')
TOTAL_FILL = PatternFill('solid', fgColor='E2EFDA')
TOTAL_FONT = Font(name='Arial', bold=True, size=10)
RED = Font(name='Arial', size=10, color='FF0000')
GREEN = Font(name='Arial', size=10, color='008000')
CF_GREEN = PatternFill('solid', fgColor='E2EFDA')
CF_BLUE = PatternFill('solid', fgColor='D6E4F0')
CF_YELLOW = PatternFill('solid', fgColor='FFF2CC')
CF_RED = PatternFill('solid', fgColor='FCE4EC')
ORANGE_FONT = Font(name='Arial', bold=True, size=10, color='FF6600')
AMARILLA_FILL = PatternFill('solid', fgColor='FFFFCC')

# Colores por perfil de calidad
PERFIL_COLORS = {
    'Standard':      '1F4E79',
    'Búfalo':        '6D4C41',
    'Premium Black': '212121',
    'Exportación':   '1B5E20',
}

PERFIL_LABELS = {
    'Standard':      'ESTÁNDAR — Mercado Interno',
    'Búfalo':        'BÚFALO — Carne de Búfalo',
    'Premium Black': 'PREMIUM BLACK — Selección Premium',
    'Exportación':   'EXPORTACIÓN — Mercado Internacional',
}


def hdr_row(ws, row, headers, fill=HDR_FILL, font=HDR_FONT):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.fill, c.font = fill, font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def build_analisis(data, output_path, calidad='Standard', costos_override=None,
                   precio_amarilla=None, price_matrix=None, extra_info=None):
    """
    Genera un Excel de 6 hojas con análisis de romaneo.

    Args:
        data: dict con 'romaneo' y 'cortes' (output de pdf_parser)
        output_path: ruta del Excel de salida
        calidad: 'Standard' | 'Búfalo' | 'Premium Black' | 'Exportación'
        costos_override: dict opcional para sobreescribir costos del perfil
        precio_amarilla: precio $/kg amarilla (default 10500)
        price_matrix: dict {GRUPO: {CLIENTE_COL: precio}}
    """
    if price_matrix is None:
        price_matrix = {}

    GRUPOS = GRUPOS_POR_CALIDAD.get(calidad, GRUPOS_POR_CALIDAD['Standard'])
    costos_base = COSTOS_PERFILES.get(calidad, COSTOS_PERFILES['Standard']).copy()
    if costos_override:
        costos_base.update(costos_override)
    costos = costos_base
    precio_am = precio_amarilla or AMARILLA_PRECIO_DEFAULT
    color_perfil = PERFIL_COLORS.get(calidad, '1F4E79')
    label_perfil = PERFIL_LABELS.get(calidad, calidad)

    rom = data['romaneo'] if 'romaneo' in data else data
    cortes = data.get('cortes', rom.get('cortes', []))

    # Contramarca → pricing column mapping
    cm_to_client = {}
    for c in cortes:
        cm = str(c.get('contramarca', ''))
        if cm and cm in CONTRAMARCA_MAP:
            _, pricing_col = CONTRAMARCA_MAP[cm]
            cm_to_client[cm] = pricing_col

    # Rendimiento objetivo
    categoria = rom.get('categoria', 'Vaca')
    cat_clean = categoria.split('(')[0].strip()
    rend_tabla = REND_OBJETIVO.get(calidad, REND_OBJETIVO['Standard'])
    rend_obj_base = rend_tabla.get(cat_clean, 0.66)

    meat_cortes = [c for c in cortes if c.get('grupo') != 'GRASA']
    kg_total_meat = sum(c['kg'] for c in meat_cortes)
    kg_anatomico = sum(c['kg'] for c in meat_cortes if c.get('tipo', '') == 'ANATÓMICO')
    pct_anatomico = kg_anatomico / kg_total_meat if kg_total_meat > 0 else 0
    real_rend = kg_total_meat / rom.get('kg_entrada', 1) if rom.get('kg_entrada', 0) > 0 else 0
    ajuste_anatomico = 0.01 if pct_anatomico > 0.50 else 0
    rend_obj = rend_obj_base + ajuste_anatomico

    wb = Workbook()
    N_P = len(GRUPOS)

    # Clientes únicos
    clientes_romaneo = {}
    for c in cortes:
        if c.get('grupo') == 'GRASA':
            continue
        cli = c.get('cliente', 'SIN ASIGNAR')
        cm = str(c.get('contramarca', ''))
        if cli not in clientes_romaneo:
            pricing_col = cm_to_client.get(cm, 'RESTO CLIENTES AMBA')
            es_amarilla = cm in AMARILLA_CONTRAMARCAS
            clientes_romaneo[cli] = {'contramarca': cm, 'pricing_col': pricing_col, 'es_amarilla': es_amarilla}
    cli_list = sorted(clientes_romaneo.keys())

    # ══════ PARAMETROS ══════
    ws = wb.active
    ws.title = "PARAMETROS"
    ws.sheet_properties.tabColor = color_perfil
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20

    ws.merge_cells('A1:B1')
    ws['A1'] = f'PARÁMETROS — {label_perfil}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color=color_perfil)

    ws['A3'] = 'PERFIL DE CALIDAD'
    ws['A3'].font = Font(name='Arial', bold=True, size=11, color=color_perfil)
    ws['A3'].fill = SEC_FILL; ws['B3'].fill = SEC_FILL
    ws.cell(4, 1, 'Calidad seleccionada').font = DATA
    ws.cell(4, 2, calidad).font = Font(name='Arial', bold=True, size=12, color=color_perfil)

    ws['A6'] = 'COSTOS VARIABLES (por kg de carne producida)'
    ws['A6'].font = Font(name='Arial', bold=True, size=11, color=color_perfil)
    ws['A6'].fill = SEC_FILL; ws['B6'].fill = SEC_FILL

    cost_items = [
        (7, 'Mano de obra (despostada)', costos['mo']),
        (8, 'Insumos (envasado, etiquetado, etc.)', costos['insumos']),
        (9, 'Flete / transporte', costos['flete']),
        (10, 'SENASA + cuarteo', costos['senasa']),
    ]
    for r, label, val in cost_items:
        ws.cell(r, 1, label).font = DATA
        ws.cell(r, 2, val).font = BLUE; ws[f'B{r}'].number_format = '#,##0'

    ws.cell(11, 1, 'Impuestos y costos financieros (ver resumen abajo)').font = Font(
        name='Arial', italic=True, size=9, color='888888')
    ws['B11'] = '=B40'
    ws['B11'].font = Font(name='Arial', italic=True, size=9, color='888888')
    ws['B11'].number_format = '0.00%'
    ws.cell(12, 1, 'TOTAL COSTO VARIABLE / kg prod. (sin impuestos)').font = TOTAL_FONT
    ws.cell(12, 1).fill = TOTAL_FILL
    ws['B12'] = '=SUM(B7:B10)'; ws['B12'].font = TOTAL_FONT; ws['B12'].fill = TOTAL_FILL
    ws['B12'].number_format = '#,##0'

    ws['A14'] = 'REFERENCIA MERCADO (MAG Cañuelas)'
    ws['A14'].font = Font(name='Arial', bold=True, size=11, color=color_perfil)
    ws['A14'].fill = SEC_FILL; ws['B14'].fill = SEC_FILL
    ws.cell(15, 1, 'Precio MAG referencia $/kg media res').font = DATA
    ws.cell(15, 2, rom.get('precio_mag', 0)).font = BLUE; ws['B15'].number_format = '#,##0'

    ws['A17'] = 'PRECIO AMARILLA ($/kg, todos los cortes)'
    ws['A17'].font = Font(name='Arial', bold=True, size=11, color='FF6600')
    ws['A17'].fill = AMARILLA_FILL; ws['B17'].fill = AMARILLA_FILL
    ws.cell(17, 2, precio_am).font = Font(name='Arial', bold=True, size=12, color='0000FF')
    ws['B17'].number_format = '#,##0'

    ws['A19'] = 'RENDIMIENTO OBJETIVO'
    ws['A19'].font = Font(name='Arial', bold=True, size=11, color=color_perfil)
    ws['A19'].fill = SEC_FILL; ws['B19'].fill = SEC_FILL
    ws.cell(20, 1, f'Categoría: {categoria}').font = DATA
    ws.cell(20, 2, rend_obj_base).font = BLUE; ws['B20'].number_format = '0.0%'
    ws.cell(21, 1, f'Ajuste anatómico (>50% = +1%): {pct_anatomico*100:.0f}% anatómico').font = DATA
    ws.cell(21, 2, ajuste_anatomico).font = BLUE; ws['B21'].number_format = '+0.0%;-0.0%;0.0%'
    ws.cell(22, 1, 'RENDIMIENTO OBJETIVO FINAL').font = TOTAL_FONT; ws.cell(22, 1).fill = TOTAL_FILL
    ws.cell(22, 2, rend_obj).font = Font(name='Arial', bold=True, size=12, color='0000FF')
    ws['B22'].number_format = '0.0%'; ws.cell(22, 2).fill = TOTAL_FILL

    # Tabla comparativa de costos por perfil
    ws['A24'] = 'COMPARATIVA DE PERFILES'
    ws['A24'].font = Font(name='Arial', bold=True, size=11, color=color_perfil)
    ws['A24'].fill = SEC_FILL; ws['B24'].fill = SEC_FILL
    hdr_row(ws, 25, ['Concepto', 'Standard', 'Búfalo', 'Premium Black', 'Exportación'])
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14

    for i, (key, label) in enumerate([
        ('mo', 'Mano de obra'),
        ('insumos', 'Insumos'),
        ('flete', 'Flete'),
        ('senasa', 'SENASA'),
        ('iibb', 'IIBB'),
    ]):
        r = 26 + i
        ws.cell(r, 1, label).font = DATA
        for j, perfil in enumerate(['Standard', 'Búfalo', 'Premium Black', 'Exportación']):
            val = COSTOS_PERFILES[perfil][key]
            ws.cell(r, 2 + j, val).font = DATA
            if key == 'iibb':
                ws.cell(r, 2 + j).number_format = '0.0%'
            else:
                ws.cell(r, 2 + j).number_format = '#,##0'
            # Resaltar perfil activo
            if perfil == calidad:
                ws.cell(r, 2 + j).font = Font(name='Arial', bold=True, size=10, color=color_perfil)

    # ══════ IMPUESTOS Y COSTO FINANCIERO (editables) ══════
    ws['A33'] = 'IMPUESTOS Y COSTO FINANCIERO (editables)'
    ws['A33'].font = Font(name='Arial', bold=True, size=11, color=color_perfil)
    ws['A33'].fill = SEC_FILL; ws['B33'].fill = SEC_FILL

    ws.cell(34, 1, 'Ganancias (% sobre venta)').font = DATA
    ws.cell(34, 2, 0.02).font = BLUE; ws['B34'].number_format = '0.00%'

    ws.cell(35, 1, 'IIBB (% sobre venta)').font = DATA
    ws.cell(35, 2, 0.015).font = BLUE; ws['B35'].number_format = '0.00%'

    ws.cell(36, 1, 'Impuesto Débitos/Créditos (% sobre venta)').font = DATA
    ws.cell(36, 2, 0.012).font = BLUE; ws['B36'].number_format = '0.00%'

    ws.cell(37, 1, 'TNA costo financiero').font = DATA
    ws.cell(37, 2, 0.30).font = BLUE; ws['B37'].number_format = '0.00%'

    ws.cell(38, 1, 'Días de financiamiento').font = DATA
    ws.cell(38, 2, 20).font = BLUE; ws['B38'].number_format = '0'

    ws.cell(39, 1, 'Costo financiero (% sobre venta) = TNA × días/365').font = Font(name='Arial', italic=True, size=9)
    ws['B39'] = '=B37*B38/365'
    ws['B39'].font = BLUE; ws['B39'].number_format = '0.00%'

    ws.cell(40, 1, 'TOTAL IMPUESTOS + FINANCIERO').font = TOTAL_FONT
    ws.cell(40, 1).fill = TOTAL_FILL
    ws['B40'] = '=B34+B35+B36+B39'
    ws['B40'].font = TOTAL_FONT; ws['B40'].fill = TOTAL_FILL; ws['B40'].number_format = '0.00%'

    # ══════ PRECIOS ══════
    ws_p = wb.create_sheet("PRECIOS")
    ws_p.sheet_properties.tabColor = "2E7D32"
    ws_p.column_dimensions['A'].width = 22
    ws_p.column_dimensions['B'].width = 14
    ws_p.column_dimensions['C'].width = 10

    ws_p.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(cli_list))
    ws_p['A1'] = f'PRECIOS DE VENTA — {calidad.upper()}'
    ws_p['A1'].font = Font(name='Arial', bold=True, size=12, color='2E7D32')

    hdrs = ['Corte', '% x Tropa', 'Tipo $'] + [
        f'{cli}\n(ctm {clientes_romaneo[cli]["contramarca"]})' for cli in cli_list
    ]
    P_HDR = 3
    hdr_row(ws_p, P_HDR, hdrs, PatternFill('solid', fgColor='2E7D32'))
    for i in range(len(cli_list)):
        col = 4 + i
        ws_p.column_dimensions[get_column_letter(col)].width = 18
        cli = cli_list[i]
        if clientes_romaneo[cli]['es_amarilla']:
            ws_p.cell(P_HDR, col).fill = PatternFill('solid', fgColor='FF8C00')

    PREC_FIRST = P_HDR + 1
    PREC_LAST = PREC_FIRST + N_P - 1

    for i, (grupo, pct) in enumerate(GRUPOS):
        r = PREC_FIRST + i
        ws_p.cell(r, 1, grupo).font = DATA
        ws_p.cell(r, 2, pct).font = BLUE; ws_p[f'B{r}'].number_format = '0.00%'
        es_caro = grupo in CORTES_CAROS
        ws_p.cell(r, 3, 'CARO' if es_caro else 'BARATO').font = Font(
            name='Arial', bold=True, size=9,
            color='C62828' if es_caro else '2E7D32'
        )
        ws_p.cell(r, 3).alignment = Alignment(horizontal='center')

        for j, cli in enumerate(cli_list):
            col = 4 + j
            cli_info = clientes_romaneo[cli]
            # Recortes: precio fijo independiente del cliente
            if grupo in PRECIOS_FIJOS_RECORTE:
                price = PRECIOS_FIJOS_RECORTE[grupo]
                ws_p.cell(r, col, price).font = BLUE
                ws_p.cell(r, col).number_format = '#,##0'
            elif cli_info['es_amarilla']:
                ws_p.cell(r, col).value = '=PARAMETROS!$B$17'
                ws_p.cell(r, col).font = ORANGE_FONT
                ws_p.cell(r, col).fill = AMARILLA_FILL
            else:
                pricing_col = cli_info['pricing_col']
                price = None
                if grupo in price_matrix and pricing_col in price_matrix[grupo]:
                    price = price_matrix[grupo][pricing_col]
                elif grupo in price_matrix:
                    for fb in ['NETOS PEYA', 'RESTO CLIENTES AMBA']:
                        if fb in price_matrix[grupo]:
                            price = price_matrix[grupo][fb]; break
                ws_p.cell(r, col, round(price) if price else 0).font = BLUE if price else RED
                ws_p.cell(r, col).number_format = '#,##0'

    r_tot_p = PREC_LAST + 1
    ws_p.cell(r_tot_p, 1, 'TOTAL').font = TOTAL_FONT
    ws_p.cell(r_tot_p, 1).fill = TOTAL_FILL; ws_p.cell(r_tot_p, 2).fill = TOTAL_FILL
    ws_p[f'B{r_tot_p}'] = f'=SUM(B{PREC_FIRST}:B{PREC_LAST})'
    ws_p[f'B{r_tot_p}'].number_format = '0.00%'
    ws_p[f'B{r_tot_p}'].font = TOTAL_FONT; ws_p[f'B{r_tot_p}'].fill = TOTAL_FILL

    # ══════ ROMANEO ══════
    ws_r = wb.create_sheet("ROMANEO")
    ws_r.sheet_properties.tabColor = "E65100"
    for col, w in {'A': 30, 'B': 22, 'C': 14, 'D': 12, 'E': 12,
                   'F': 14, 'G': 12, 'H': 14, 'I': 18, 'J': 14}.items():
        ws_r.column_dimensions[col].width = w

    ws_r.merge_cells('A1:J1')
    titulo_rom = rom.get('numero', 'S/N')
    fecha_rom = rom.get('fecha', '')
    ws_r['A1'] = f'ROMANEO N° {titulo_rom} — {fecha_rom} — {calidad.upper()}'
    ws_r['A1'].font = Font(name='Arial', bold=True, size=14, color='E65100')

    hdr_items = [
        (3, 'N° Romaneo', rom.get('numero', '')),
        (4, 'Fecha', rom.get('fecha', '')),
        (5, 'Medias reses', rom.get('medias_reses', 0)),
        (6, 'Kg entrada (media res)', rom.get('kg_entrada', 0)),
        (7, 'Categoría (principal)', rom.get('categoria', '')),
        (8, 'Precio compra $/kg media res (s/IVA)', rom.get('precio_compra', 0)),
        (9, f'Rendimiento objetivo ({cat_clean})', rend_obj),
        (10, f'Rendimiento real', round(real_rend, 4)),
        (11, 'Perfil de calidad', calidad),
    ]
    for r, label, val in hdr_items:
        ws_r.cell(r, 1, label).font = SEC_FONT; ws_r.cell(r, 1).fill = SEC_FILL
        ws_r.cell(r, 2, val).font = BLUE
        if isinstance(val, (int, float)) and val != 0:
            ws_r[f'B{r}'].number_format = '#,##0' if val > 1 else '0.0%'

    # Desglose categoría (col D-F)
    desglose_cat = rom.get('desglose_categoria', data.get('desglose_categoria', {}))
    if desglose_cat:
        ws_r.cell(3, 4, 'Desglose categoría').font = SEC_FONT; ws_r.cell(3, 4).fill = SEC_FILL
        ws_r.column_dimensions['D'].width = 20
        ws_r.column_dimensions['E'].width = 12
        r_cat = 4
        for cat_name, cat_data in desglose_cat.items():
            ws_r.cell(r_cat, 4, cat_name).font = DATA
            ws_r.cell(r_cat, 5, round(cat_data['pct'], 1) / 100).font = BLUE
            ws_r[f'E{r_cat}'].number_format = '0.0%'
            r_cat += 1

    # Desglose tipificación (col G-H)
    desglose_tip = rom.get('desglose_tipificacion', data.get('desglose_tipificacion', {}))
    if desglose_tip:
        ws_r.cell(3, 7, 'Tipificación').font = SEC_FONT; ws_r.cell(3, 7).fill = SEC_FILL
        ws_r.column_dimensions['G'].width = 14
        total_tip = sum(desglose_tip.values())
        r_tip = 4
        for tip, kg in sorted(desglose_tip.items()):
            pct = kg / total_tip if total_tip > 0 else 0
            ws_r.cell(r_tip, 7, tip).font = DATA
            ws_r.cell(r_tip, 8, round(pct, 3)).font = BLUE
            ws_r[f'H{r_tip}'].number_format = '0.0%'
            r_tip += 1

    # Datos de tropa y cruce con planilla de compras
    if extra_info is None:
        extra_info = {}
    tropas_match = extra_info.get('tropas_match', [])
    tropas_sin_match = extra_info.get('tropas_sin_match', [])
    dias_fp = extra_info.get('dias_faena_produccion')

    if tropas_match or tropas_sin_match or dias_fp is not None:
        ws_r.cell(13, 1, 'DATOS DE COMPRA (cruce con planilla)').font = Font(name='Arial', bold=True, size=11, color='1F4E79')
        ws_r.cell(13, 1).fill = SEC_FILL
        r_extra = 14
        if tropas_match:
            tropas_txt = ', '.join(m['tropa'] for m in tropas_match)
            ws_r.cell(r_extra, 1, 'Tropas matcheadas').font = SEC_FONT
            ws_r.cell(r_extra, 1).fill = SEC_FILL
            ws_r.cell(r_extra, 2, tropas_txt).font = BLUE
            r_extra += 1
            # Detalle por tropa: N° | $/kg | Kg | Fecha faena | Subtotal
            hdr_row(ws_r, r_extra, ['Tropa', 'Precio $/kg', 'Kg', 'Fecha faena', 'Subtotal'])
            r_extra += 1
            for m in tropas_match:
                ws_r.cell(r_extra, 1, m['tropa']).font = DATA
                ws_r.cell(r_extra, 2, m['precio']).font = DATA
                ws_r[f'B{r_extra}'].number_format = '$#,##0'
                ws_r.cell(r_extra, 3, m.get('kg', 0)).font = DATA
                ws_r[f'C{r_extra}'].number_format = '#,##0'
                ws_r.cell(r_extra, 4, m.get('fecha', '-')).font = DATA
                subtotal = m['precio'] * m.get('kg', 0)
                ws_r.cell(r_extra, 5, subtotal).font = DATA
                ws_r[f'E{r_extra}'].number_format = '$#,##0'
                r_extra += 1
            # Precio ponderado
            total_kg_m = sum(m.get('kg', 0) for m in tropas_match)
            if total_kg_m > 0:
                pond = sum(m['precio'] * m.get('kg', 0) for m in tropas_match) / total_kg_m
                ws_r.cell(r_extra, 1, 'Precio promedio PONDERADO por kg').font = TOTAL_FONT
                ws_r.cell(r_extra, 1).fill = TOTAL_FILL
                ws_r.cell(r_extra, 2, round(pond)).font = TOTAL_FONT
                ws_r[f'B{r_extra}'].number_format = '$#,##0'; ws_r.cell(r_extra, 2).fill = TOTAL_FILL
                r_extra += 1

        if tropas_sin_match:
            r_extra += 1
            ws_r.cell(r_extra, 1, '⚠️ Tropas SIN matcheo en planilla').font = Font(
                name='Arial', bold=True, size=10, color='C62828')
            r_extra += 1
            for m in tropas_sin_match:
                kg_sm = m.get('kg', 0) if isinstance(m, dict) else 0
                tropa_sm = m.get('tropa', m) if isinstance(m, dict) else m
                ws_r.cell(r_extra, 1, f"  Tropa {tropa_sm}").font = Font(name='Arial', size=10, color='C62828')
                ws_r.cell(r_extra, 2, f"{kg_sm:,.0f} kg — sin precio de referencia").font = Font(
                    name='Arial', italic=True, size=9, color='C62828')
                r_extra += 1
        if dias_fp is not None:
            ws_r.cell(r_extra, 1, 'Días faena → producción').font = SEC_FONT
            ws_r.cell(r_extra, 1).fill = SEC_FILL
            dias_font = Font(name='Arial', bold=True, size=11,
                             color='2E7D32' if dias_fp < 3 else 'E65100' if dias_fp < 5 else 'C62828')
            ws_r.cell(r_extra, 2, f"{dias_fp:.0f} días").font = dias_font
            r_extra += 1

        # Análisis de rendimiento
        r_extra += 1
        ws_r.cell(r_extra, 1, 'ANÁLISIS DE RENDIMIENTO').font = Font(name='Arial', bold=True, size=11, color='6A1B9A')
        ws_r.cell(r_extra, 1).fill = SEC_FILL
        r_extra += 1
        diff_pct = real_rend - rend_obj
        ws_r.cell(r_extra, 1, f'Objetivo: {rend_obj*100:.1f}%  |  Real: {real_rend*100:.1f}%  |  Diferencia: {diff_pct*100:+.1f}%').font = Font(
            name='Arial', bold=True, size=10, color='2E7D32' if diff_pct >= 0 else 'C62828')
        r_extra += 1

        # Inferencia de por qué la diferencia
        kg_am = sum(c['kg'] for c in cortes if str(c.get('contramarca', '')) in AMARILLA_CONTRAMARCAS)
        pct_am = (kg_am / kg_total_meat * 100) if kg_total_meat > 0 else 0
        pct_anat_val = pct_anatomico * 100

        inferencias = []
        if diff_pct < -0.02:
            if pct_am > 15:
                inferencias.append(f"Alto % amarilla ({pct_am:.0f}%) puede indicar tropas de menor calidad")
            if pct_anat_val < 50:
                inferencias.append(f"Solo {pct_anat_val:.0f}% anatómico — mucho porcionado puede generar merma")
            if dias_fp and dias_fp >= 5:
                inferencias.append(f"{dias_fp:.0f} días entre faena y producción — posible pérdida de peso por oreo")
            if not inferencias:
                inferencias.append("Revisar calidad de la tropa y condiciones de desposte")
        elif diff_pct > 0.02:
            if pct_am < 5:
                inferencias.append(f"Baja amarilla ({pct_am:.0f}%) — tropa de buena calidad")
            if pct_anat_val > 70:
                inferencias.append(f"Alto % anatómico ({pct_anat_val:.0f}%) — menos merma por procesamiento")
            if not inferencias:
                inferencias.append("Buen rendimiento — tropa y desposte dentro de lo esperado")
        else:
            inferencias.append("Rendimiento en línea con el objetivo")

        for inf in inferencias:
            ws_r.cell(r_extra, 1, f"→ {inf}").font = Font(name='Arial', italic=True, size=9, color='555555')
            r_extra += 1

        # ── Rendimiento por categoría (útil en acumulados) ──
        if desglose_cat and len(desglose_cat) > 1:
            r_extra += 1
            ws_r.cell(r_extra, 1, 'RENDIMIENTO POR CATEGORÍA (sobre kg entrada)').font = Font(
                name='Arial', bold=True, size=11, color='6A1B9A')
            ws_r.cell(r_extra, 1).fill = SEC_FILL
            r_extra += 1
            hdr_row(ws_r, r_extra, ['Categoría', '% del total kg entrada', 'Kg entrada', 'Rend obj'])
            r_extra += 1
            rend_obj_map = REND_OBJETIVO.get(calidad, REND_OBJETIVO['Standard'])
            for cat_name, cat_data in desglose_cat.items():
                kg_cat = cat_data.get('kg', 0)
                pct_cat = cat_data.get('pct', 0)
                obj_cat = rend_obj_map.get(cat_name, 0.66)
                ws_r.cell(r_extra, 1, cat_name).font = DATA
                ws_r.cell(r_extra, 2, round(pct_cat, 1) / 100).font = BLUE
                ws_r[f'B{r_extra}'].number_format = '0.0%'
                ws_r.cell(r_extra, 3, kg_cat).font = DATA
                ws_r[f'C{r_extra}'].number_format = '#,##0'
                ws_r.cell(r_extra, 4, obj_cat).font = BLUE
                ws_r[f'D{r_extra}'].number_format = '0.0%'
                r_extra += 1

        # ── Mix de cortes (caros/baratos/picada/recorte) ──
        r_extra += 1
        ws_r.cell(r_extra, 1, 'MIX DE CORTES (sobre kg carne)').font = Font(
            name='Arial', bold=True, size=11, color='6A1B9A')
        ws_r.cell(r_extra, 1).fill = SEC_FILL
        r_extra += 1

        kg_caros_v = sum(c['kg'] for c in cortes
                         if c.get('grupo') in CORTES_CAROS)
        kg_picada_v = sum(c['kg'] for c in cortes
                          if c.get('grupo') == 'CARNE PICADA')
        kg_recorte_v = sum(c['kg'] for c in cortes
                           if c.get('grupo', '').startswith('RECORTE'))
        kg_baratos_v = kg_total_meat - kg_caros_v - kg_picada_v - kg_recorte_v

        mix_items = [
            ('Cortes caros', kg_caros_v, 'C62828'),
            ('Cortes baratos', kg_baratos_v, '1F4E79'),
            ('Carne picada', kg_picada_v, 'E65100'),
            ('Recorte', kg_recorte_v, '999999'),
        ]
        for label, kg_v, color in mix_items:
            pct_v = kg_v / kg_total_meat if kg_total_meat > 0 else 0
            ws_r.cell(r_extra, 1, label).font = Font(name='Arial', size=10, color=color)
            ws_r.cell(r_extra, 2, pct_v).font = Font(name='Arial', bold=True, size=11, color=color)
            ws_r[f'B{r_extra}'].number_format = '0.0%'
            ws_r.cell(r_extra, 3, round(kg_v, 2)).font = DATA
            ws_r[f'C{r_extra}'].number_format = '#,##0.00'
            r_extra += 1

        DET_HDR = r_extra + 1
    else:
        diff_pct = real_rend - rend_obj
        ws_r.cell(13, 1, f'Rendimiento: Obj {rend_obj*100:.1f}% | Real {real_rend*100:.1f}% | Dif {diff_pct*100:+.1f}%').font = Font(
            name='Arial', bold=True, size=10, color='2E7D32' if diff_pct >= 0 else 'C62828')
        DET_HDR = 15

    hdr_row(ws_r, DET_HDR,
            ['Corte (detalle)', 'Grupo Precio', 'Tipo', 'Piezas', 'Unidades',
             'Kg', 'Destino', 'Rend %', 'Cliente', 'Nro.Venta'],
            PatternFill('solid', fgColor='E65100'))

    N_DET = len(cortes)
    DET_FIRST = DET_HDR + 1
    DET_LAST = DET_FIRST + N_DET - 1

    for i, c in enumerate(cortes):
        r = DET_FIRST + i
        ws_r.cell(r, 1, c['corte']).font = DATA
        ws_r.cell(r, 2, c['grupo']).font = DATA
        ws_r.cell(r, 3, c.get('tipo', 'ANATÓMICO')).font = DATA
        ws_r.cell(r, 4, c.get('piezas', 0)).font = DATA; ws_r[f'D{r}'].number_format = '#,##0'
        ws_r.cell(r, 5, c.get('unidades', 0)).font = DATA
        ws_r.cell(r, 6, c['kg']).font = DATA; ws_r[f'F{r}'].number_format = '#,##0.00'
        ws_r.cell(r, 7, c.get('destino', '')).font = DATA
        ws_r[f'H{r}'] = f'=IF($B$6=0,0,F{r}/$B$6)'
        ws_r[f'H{r}'].number_format = '0.00%'; ws_r[f'H{r}'].font = DATA
        ws_r.cell(r, 9, c.get('cliente', '')).font = DATA
        ws_r.cell(r, 10, c.get('nro_venta', '')).font = DATA
        cm = str(c.get('contramarca', ''))
        if cm in AMARILLA_CONTRAMARCAS:
            for col in range(1, 11):
                ws_r.cell(r, col).fill = AMARILLA_FILL

    R_TOT = DET_LAST + 1
    ws_r.cell(R_TOT, 1, 'TOTAL SALIDAS').font = TOTAL_FONT
    for col in range(1, 11):
        ws_r.cell(R_TOT, col).fill = TOTAL_FILL
    ws_r[f'D{R_TOT}'] = f'=SUM(D{DET_FIRST}:D{DET_LAST})'
    ws_r[f'D{R_TOT}'].number_format = '#,##0'; ws_r[f'D{R_TOT}'].font = TOTAL_FONT
    ws_r[f'E{R_TOT}'] = f'=SUM(E{DET_FIRST}:E{DET_LAST})'
    ws_r[f'E{R_TOT}'].font = TOTAL_FONT
    ws_r[f'F{R_TOT}'] = f'=SUM(F{DET_FIRST}:F{DET_LAST})'
    ws_r[f'F{R_TOT}'].number_format = '#,##0.00'; ws_r[f'F{R_TOT}'].font = TOTAL_FONT
    ws_r[f'H{R_TOT}'] = f'=IF($B$6=0,0,F{R_TOT}/$B$6)'
    ws_r[f'H{R_TOT}'].number_format = '0.00%'; ws_r[f'H{R_TOT}'].font = TOTAL_FONT

    R_MERMA = R_TOT + 1
    ws_r.cell(R_MERMA, 1, 'MERMA (oreo/proceso)').font = Font(
        name='Arial', bold=True, size=10, color='FF0000')
    ws_r[f'F{R_MERMA}'] = f'=B6-F{R_TOT}'
    ws_r[f'F{R_MERMA}'].number_format = '#,##0.00'
    ws_r[f'F{R_MERMA}'].font = Font(name='Arial', bold=True, size=10, color='FF0000')
    ws_r[f'H{R_MERMA}'] = f'=IF($B$6=0,0,F{R_MERMA}/$B$6)'
    ws_r[f'H{R_MERMA}'].number_format = '0.00%'
    ws_r[f'H{R_MERMA}'].font = Font(name='Arial', bold=True, size=10, color='FF0000')

    R_CARNE = R_MERMA + 1
    ws_r.cell(R_CARNE, 1, 'KG CARNE PRODUCIDA (sin grasa/decomiso)').font = Font(
        name='Arial', bold=True, size=10, color='2E7D32')
    ws_r[f'F{R_CARNE}'] = f'=F{R_TOT}-SUMIF(B{DET_FIRST}:B{DET_LAST},"GRASA",F{DET_FIRST}:F{DET_LAST})'
    ws_r[f'F{R_CARNE}'].number_format = '#,##0.00'
    ws_r[f'F{R_CARNE}'].font = Font(name='Arial', bold=True, size=10, color='2E7D32')
    ws_r[f'H{R_CARNE}'] = f'=IF($B$6=0,0,F{R_CARNE}/$B$6)'
    ws_r[f'H{R_CARNE}'].number_format = '0.00%'
    ws_r[f'H{R_CARNE}'].font = Font(name='Arial', bold=True, size=10, color='2E7D32')

    R_ROBJ = R_CARNE + 1
    ws_r.cell(R_ROBJ, 1, 'REND vs OBJETIVO').font = Font(
        name='Arial', bold=True, size=10, color='1F4E79')
    diff_rend = real_rend - rend_obj
    ws_r.cell(R_ROBJ, 6, diff_rend).font = Font(
        name='Arial', bold=True, size=10,
        color='008000' if diff_rend >= 0 else 'FF0000'
    )
    ws_r[f'F{R_ROBJ}'].number_format = '+0.0%;-0.0%'
    ws_r.cell(R_ROBJ, 8, 'OK' if diff_rend >= 0 else 'BAJO OBJETIVO').font = Font(
        name='Arial', bold=True, size=10,
        color='008000' if diff_rend >= 0 else 'FF0000'
    )

    # Amarillas summary
    R_AM = R_ROBJ + 2
    ws_r.cell(R_AM, 1, 'AMARILLAS (ctmarcas 47/73/74)').font = ORANGE_FONT
    ws_r.cell(R_AM, 1).fill = AMARILLA_FILL
    kg_amarilla = sum(c['kg'] for c in cortes
                      if str(c.get('contramarca', '')) in AMARILLA_CONTRAMARCAS)
    kg_total = sum(c['kg'] for c in cortes)
    ws_r.cell(R_AM, 6, kg_amarilla).font = ORANGE_FONT
    ws_r[f'F{R_AM}'].number_format = '#,##0.00'
    pct_am = kg_amarilla / kg_total if kg_total > 0 else 0
    ws_r.cell(R_AM, 8, pct_am).font = ORANGE_FONT
    ws_r[f'H{R_AM}'].number_format = '0.0%'

    # ══════ ANALISIS ══════
    ws_a = wb.create_sheet("ANALISIS")
    ws_a.sheet_properties.tabColor = "6A1B9A"
    for col, w in [('A', 24), ('B', 14), ('C', 14), ('D', 14),
                   ('E', 14), ('F', 14), ('G', 10), ('H', 16)]:
        ws_a.column_dimensions[col].width = w

    ws_a.merge_cells('A1:H1')
    ws_a['A1'] = f'ANÁLISIS DE RENDIMIENTO — {titulo_rom} — {calidad.upper()}'
    ws_a['A1'].font = Font(name='Arial', bold=True, size=14, color='6A1B9A')

    ws_a['A3'] = 'Kg entrada'; ws_a['A3'].font = SEC_FONT
    ws_a['B3'] = '=ROMANEO!B6'; ws_a['B3'].number_format = '#,##0'
    ws_a['B3'].font = Font(name='Arial', bold=True, size=12)
    ws_a['C3'] = 'Kg carne'; ws_a['C3'].font = SEC_FONT
    ws_a['D3'] = f'=ROMANEO!F{R_CARNE}'; ws_a['D3'].number_format = '#,##0.00'
    ws_a['D3'].font = Font(name='Arial', bold=True, size=12, color='2E7D32')
    ws_a['E3'] = 'Rend. cárnico'; ws_a['E3'].font = SEC_FONT
    ws_a['F3'] = '=IF(B3=0,0,D3/B3)'; ws_a['F3'].number_format = '0.00%'
    ws_a['F3'].font = Font(name='Arial', bold=True, size=12, color='6A1B9A')
    ws_a['G3'] = 'Obj.'; ws_a['G3'].font = SEC_FONT
    ws_a['H3'] = rend_obj; ws_a['H3'].number_format = '0.0%'
    ws_a['H3'].font = Font(name='Arial', bold=True, size=12, color='1F4E79')

    AN_HDR = 6
    hdr_row(ws_a, AN_HDR,
            ['Grupo Corte', 'Kg Real', '% Real s/Carne', '% Esperado',
             'Desvío', '% Cumplim.', 'Tipo $', 'Calificación'],
            PatternFill('solid', fgColor='6A1B9A'))

    AN_FIRST = AN_HDR + 1
    # Subcortes se muestran como filas COLAPSABLES debajo del parent (group/
    # outline). El parent absorbe la suma de sus subcortes en col B.
    try:
        from config import SUBCORTE_TO_PARENT, SUBCORTES
    except Exception:
        SUBCORTE_TO_PARENT = {}
        SUBCORTES = {}
    subcortes_set = set(SUBCORTE_TO_PARENT.keys())
    GRUPOS_PARENTS = [(g, p) for (g, p) in GRUPOS if g not in subcortes_set]

    # Construir la lista expandida: parent + sus subcortes (si tiene)
    # filas_an[i] = (kind, name, parent_name_or_None)
    filas_an = []
    for grupo_name, _ in GRUPOS_PARENTS:
        filas_an.append(('parent', grupo_name, None))
        for sc_info in SUBCORTES.get(grupo_name, []):
            sc_nombre = sc_info[0]
            filas_an.append(('sub', sc_nombre, grupo_name))

    AN_LAST = AN_FIRST + len(filas_an) - 1
    grp = f'ROMANEO!$B${DET_FIRST}:$B${DET_LAST}'
    kgr = f'ROMANEO!$F${DET_FIRST}:$F${DET_LAST}'

    SUB_FONT = Font(name='Arial', size=9, color='666666', italic=True)
    parent_rows = []  # para el TOTAL

    # Habilitar outline summary sobre los subcortes (parent arriba)
    ws_a.sheet_properties.outlinePr.summaryBelow = False

    for i, (kind, name, parent) in enumerate(filas_an):
        r = AN_FIRST + i

        if kind == 'parent':
            parent_rows.append(r)
            try:
                p_r = PREC_FIRST + next(
                    idx for idx, (g, _) in enumerate(GRUPOS) if g == name)
            except StopIteration:
                p_r = PREC_FIRST
            es_caro = name in CORTES_CAROS

            ws_a[f'A{r}'] = f'=PRECIOS!A{p_r}'; ws_a[f'A{r}'].font = DATA
            subcs = [sc[0] for sc in SUBCORTES.get(name, [])]
            if subcs:
                partes = [f'SUMIF({grp},"{name}",{kgr})']
                for sc in subcs:
                    partes.append(f'SUMIF({grp},"{sc}",{kgr})')
                ws_a[f'B{r}'] = '=' + '+'.join(partes)
            else:
                ws_a[f'B{r}'] = f'=SUMIF({grp},A{r},{kgr})'
            ws_a[f'B{r}'].number_format = '#,##0.00'; ws_a[f'B{r}'].font = DATA
            ws_a[f'C{r}'] = f'=IF(ROMANEO!$F${R_CARNE}=0,0,B{r}/ROMANEO!$F${R_CARNE})'
            ws_a[f'C{r}'].number_format = '0.00%'; ws_a[f'C{r}'].font = DATA
            ws_a[f'D{r}'] = f'=PRECIOS!B{p_r}'
            ws_a[f'D{r}'].number_format = '0.00%'; ws_a[f'D{r}'].font = DATA
            ws_a[f'E{r}'] = f'=C{r}-D{r}'
            ws_a[f'E{r}'].number_format = '+0.00%;-0.00%;0.00%'; ws_a[f'E{r}'].font = DATA
            ws_a[f'F{r}'] = f'=IF(D{r}=0,IF(B{r}=0,1,1.5),C{r}/D{r})'
            ws_a[f'F{r}'].number_format = '0.0%'; ws_a[f'F{r}'].font = DATA

            ws_a.cell(r, 7, 'CARO' if es_caro else 'BARATO').font = Font(
                name='Arial', bold=True, size=9,
                color='C62828' if es_caro else '2E7D32'
            )
            ws_a.cell(r, 7).alignment = Alignment(horizontal='center')

            if es_caro:
                formula = (
                    f'=IF(D{r}=0,IF(B{r}=0,"N/A","EXTRA"),'
                    f'IF(F{r}>=1.1,"ÓPTIMO",IF(F{r}>=0.95,"BUENO",IF(F{r}>=0.8,"REGULAR","MALO"))))'
                )
            else:
                formula = (
                    f'=IF(D{r}=0,IF(B{r}=0,"N/A","EXTRA"),'
                    f'IF(F{r}<=0.8,"ÓPTIMO",IF(F{r}<=0.95,"BUENO",IF(F{r}<=1.1,"REGULAR","MALO"))))'
                )
            ws_a[f'H{r}'] = formula
            ws_a[f'H{r}'].font = Font(name='Arial', bold=True, size=10)
            ws_a[f'H{r}'].alignment = Alignment(horizontal='center')
        else:
            # Subcorte — fila colapsable bajo el parent
            ws_a[f'A{r}'] = f'   └─ {name}'
            ws_a[f'A{r}'].font = SUB_FONT
            ws_a[f'B{r}'] = f'=SUMIF({grp},"{name}",{kgr})'
            ws_a[f'B{r}'].number_format = '#,##0.00'; ws_a[f'B{r}'].font = SUB_FONT
            ws_a[f'C{r}'] = f'=IF(ROMANEO!$F${R_CARNE}=0,0,B{r}/ROMANEO!$F${R_CARNE})'
            ws_a[f'C{r}'].number_format = '0.00%'; ws_a[f'C{r}'].font = SUB_FONT
            # D-H en blanco / italic gray para no confundir
            ws_a[f'D{r}'].font = SUB_FONT
            ws_a[f'E{r}'].font = SUB_FONT
            ws_a[f'F{r}'].font = SUB_FONT
            ws_a.cell(r, 7).font = SUB_FONT
            ws_a[f'H{r}'].font = SUB_FONT
            # Outline level 1 + colapsado por defecto
            ws_a.row_dimensions[r].outlineLevel = 1
            ws_a.row_dimensions[r].hidden = True
            ws_a.row_dimensions[r].collapsed = True

    # ── CORTES SIN CLASIFICAR (cortes que no matchearon ningún grupo) ──
    # Lista de cortes individuales sin clasificar (para mostrar nombres + kg)
    sin_clasif_list = [c for c in cortes if c.get('grupo') == 'SIN CLASIFICAR']
    sin_clasif_kg_total = sum(c.get('kg', 0) for c in sin_clasif_list)

    R_SC = AN_LAST + 1
    if sin_clasif_list:
        # Sub-header
        ws_a.cell(R_SC, 1, '⚠️ CORTES SIN CLASIFICAR').font = Font(
            name='Arial', bold=True, size=10, color='E65100')
        ws_a.cell(R_SC, 1).fill = PatternFill('solid', fgColor='FFF3E0')
        for cc in range(2, 9):
            ws_a.cell(R_SC, cc).fill = PatternFill('solid', fgColor='FFF3E0')
        # Una línea por cada corte único sin clasificar (suma kg si nombre repetido)
        from collections import defaultdict as _dd
        agrup_sc = _dd(float)
        for c in sin_clasif_list:
            nombre = c.get('corte', '?').strip()
            agrup_sc[nombre] += c.get('kg', 0)
        for nombre, kg_val in sorted(agrup_sc.items(), key=lambda x: -x[1]):
            R_SC += 1
            ws_a.cell(R_SC, 1, f'   • {nombre}').font = Font(
                name='Arial', size=9, color='666666', italic=True)
            ws_a.cell(R_SC, 2, kg_val).font = Font(
                name='Arial', size=9, color='666666', italic=True)
            ws_a.cell(R_SC, 2).number_format = '#,##0.00'
            ws_a.cell(R_SC, 3).number_format = '0.00%'
            ws_a[f'C{R_SC}'] = f'=IF(ROMANEO!$F${R_CARNE}=0,0,B{R_SC}/ROMANEO!$F${R_CARNE})'
            ws_a[f'C{R_SC}'].font = Font(name='Arial', size=9, color='666666', italic=True)
        # Subtotal SIN CLASIFICAR
        R_SC += 1
        ws_a.cell(R_SC, 1, 'Subtotal sin clasificar').font = Font(
            name='Arial', bold=True, size=10, color='E65100')
        ws_a.cell(R_SC, 2, sin_clasif_kg_total).font = Font(
            name='Arial', bold=True, size=10, color='E65100')
        ws_a.cell(R_SC, 2).number_format = '#,##0.00'
        ws_a[f'C{R_SC}'] = f'=IF(ROMANEO!$F${R_CARNE}=0,0,B{R_SC}/ROMANEO!$F${R_CARNE})'
        ws_a[f'C{R_SC}'].number_format = '0.00%'
        ws_a[f'C{R_SC}'].font = Font(name='Arial', bold=True, size=10, color='E65100')
        R_SC_SUBTOTAL = R_SC
    else:
        R_SC_SUBTOTAL = None

    R_TOT_AN = (R_SC + 1) if sin_clasif_list else (AN_LAST + 1)
    ws_a.cell(R_TOT_AN, 1, 'TOTAL CARNE').font = TOTAL_FONT
    for col in range(1, 9):
        ws_a.cell(R_TOT_AN, col).fill = TOTAL_FILL
    # Total = parents + subtotal sin clasificar para llegar al 100%
    partes = [f'B{pr}' for pr in parent_rows] if parent_rows else []
    if R_SC_SUBTOTAL:
        partes.append(f'B{R_SC_SUBTOTAL}')
    if partes:
        ws_a[f'B{R_TOT_AN}'] = '=' + '+'.join(partes)
    else:
        ws_a[f'B{R_TOT_AN}'] = f'=SUM(B{AN_FIRST}:B{AN_LAST})'
    ws_a[f'B{R_TOT_AN}'].number_format = '#,##0.00'; ws_a[f'B{R_TOT_AN}'].font = TOTAL_FONT
    ws_a[f'C{R_TOT_AN}'] = f'=IF(ROMANEO!$F${R_CARNE}=0,0,B{R_TOT_AN}/ROMANEO!$F${R_CARNE})'
    ws_a[f'C{R_TOT_AN}'].number_format = '0.00%'; ws_a[f'C{R_TOT_AN}'].font = TOTAL_FONT

    # Conditional formatting sólo sobre parents (col H)
    for pr in parent_rows:
        ws_a.conditional_formatting.add(f'H{pr}', CellIsRule(operator='equal', formula=['"ÓPTIMO"'], fill=CF_GREEN))
        ws_a.conditional_formatting.add(f'H{pr}', CellIsRule(operator='equal', formula=['"BUENO"'], fill=CF_BLUE))
        ws_a.conditional_formatting.add(f'H{pr}', CellIsRule(operator='equal', formula=['"REGULAR"'], fill=CF_YELLOW))
        ws_a.conditional_formatting.add(f'H{pr}', CellIsRule(operator='equal', formula=['"MALO"'], fill=CF_RED))

    # ══════ VENTAS ══════
    ws_v = wb.create_sheet("VENTAS")
    ws_v.sheet_properties.tabColor = "FF6600"
    for col, w in [('A', 22), ('B', 14), ('C', 14), ('D', 16),
                   ('E', 14), ('F', 18), ('G', 16)]:
        ws_v.column_dimensions[col].width = w

    ws_v.merge_cells('A1:G1')
    ws_v['A1'] = f'DESGLOSE POR CLIENTE — {titulo_rom} — {calidad.upper()}'
    ws_v['A1'].font = Font(name='Arial', bold=True, size=14, color='FF6600')

    V_HDR = 3
    hdr_row(ws_v, V_HDR,
            ['Cliente', 'Ctmarca', 'Kg Totales', '% del Total', 'Amarilla', 'Ingreso Est. $', 'Alerta'],
            PatternFill('solid', fgColor='FF6600'))

    client_data = {}
    for c in cortes:
        if c.get('grupo') == 'GRASA':
            continue
        cli = c.get('cliente', 'SIN ASIGNAR')
        cm = str(c.get('contramarca', ''))
        if cli not in client_data:
            client_data[cli] = {'kg': 0, 'ingreso': 0, 'cm': cm}
        client_data[cli]['kg'] += c['kg']
        grupo = c['grupo']
        if grupo in PRECIOS_FIJOS_RECORTE:
            price = PRECIOS_FIJOS_RECORTE[grupo]
        elif cm in AMARILLA_CONTRAMARCAS:
            price = precio_am
        elif c.get('es_bubalino'):
            price = BUBALINO_PRECIO_DEFAULT
        else:
            pricing_col = cm_to_client.get(cm, 'RESTO CLIENTES AMBA')
            price = 0
            if grupo in price_matrix and pricing_col in price_matrix[grupo]:
                price = price_matrix[grupo][pricing_col]
            elif grupo in price_matrix:
                for fb in ['NETOS PEYA', 'RESTO CLIENTES AMBA']:
                    if fb in price_matrix[grupo]:
                        price = price_matrix[grupo][fb]; break
        client_data[cli]['ingreso'] += c['kg'] * price

    v_row = V_HDR + 1
    v_first = v_row
    for cli_name in sorted(client_data.keys()):
        cd = client_data[cli_name]
        ws_v.cell(v_row, 1, cli_name).font = DATA
        ws_v.cell(v_row, 2, cd['cm']).font = DATA
        ws_v.cell(v_row, 3, round(cd['kg'], 2)).font = DATA
        ws_v[f'C{v_row}'].number_format = '#,##0.00'
        ws_v[f'D{v_row}'] = f'=IF(C{v_row}=0,0,C{v_row}/SUM(C{v_first}:C{v_first+len(client_data)-1}))'
        ws_v[f'D{v_row}'].number_format = '0.0%'; ws_v[f'D{v_row}'].font = DATA
        es_am = cd['cm'] in AMARILLA_CONTRAMARCAS
        ws_v.cell(v_row, 5, 'SÍ' if es_am else 'NO').font = ORANGE_FONT if es_am else DATA
        ws_v.cell(v_row, 6, round(cd['ingreso'])).font = DATA
        ws_v[f'F{v_row}'].number_format = '$#,##0'
        if es_am:
            ws_v.cell(v_row, 7, f'AMARILLA ${precio_am:,.0f}/kg').font = ORANGE_FONT
            for col in range(1, 8):
                ws_v.cell(v_row, col).fill = AMARILLA_FILL
        v_row += 1

    ws_v.cell(v_row, 1, 'TOTAL').font = TOTAL_FONT
    for col in range(1, 8):
        ws_v.cell(v_row, col).fill = TOTAL_FILL
    ws_v[f'C{v_row}'] = f'=SUM(C{v_first}:C{v_row-1})'
    ws_v[f'C{v_row}'].number_format = '#,##0.00'; ws_v[f'C{v_row}'].font = TOTAL_FONT
    ws_v[f'F{v_row}'] = f'=SUM(F{v_first}:F{v_row-1})'
    ws_v[f'F{v_row}'].number_format = '$#,##0'; ws_v[f'F{v_row}'].font = TOTAL_FONT

    # Precio venta promedio
    v_pvp = v_row + 2
    ws_v.cell(v_pvp, 1, 'PRECIO VENTA PROMEDIO ($/kg carne)').font = Font(
        name='Arial', bold=True, size=11, color='1F4E79')
    ws_v.cell(v_pvp, 1).fill = SEC_FILL
    ws_v[f'F{v_pvp}'] = f'=IF(C{v_row}=0,0,F{v_row}/C{v_row})'
    ws_v[f'F{v_pvp}'].number_format = '$#,##0'
    ws_v[f'F{v_pvp}'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws_v[f'F{v_pvp}'].fill = SEC_FILL

    # ══════ RESULTADO (P&L primero, detalle después) ══════
    ws_res = wb.create_sheet("RESULTADO")
    ws_res.sheet_properties.tabColor = "C62828"
    ws_res.column_dimensions['A'].width = 46
    ws_res.column_dimensions['B'].width = 24
    for c_col in ['C', 'D', 'E']:
        ws_res.column_dimensions[c_col].width = 18

    ws_res.merge_cells('A1:E1')
    ws_res['A1'] = f'RESULTADO ECONÓMICO — {titulo_rom} — {calidad.upper()}'
    ws_res['A1'].font = Font(name='Arial', bold=True, size=14, color='C62828')

    # ── Primero: ingresos por línea (oculto, para que las fórmulas funcionen) ──
    # Los pongo en una zona auxiliar que después referenciamos
    meat_only = [c for c in cortes if c.get('grupo') != 'GRASA']
    grupo_to_prec_row = {g: PREC_FIRST + i for i, (g, _) in enumerate(GRUPOS)}
    cli_to_col_idx = {cli: 4 + j for j, cli in enumerate(cli_list)}

    # Calcular ingreso total hardcoded para las fórmulas del P&L
    # (el detalle va al final)
    ingreso_total_calc = 0
    for c in meat_only:
        grupo = c['grupo']
        cli = c.get('cliente', 'SIN ASIGNAR')
        cm = str(c.get('contramarca', ''))
        if grupo in PRECIOS_FIJOS_RECORTE:
            p = PRECIOS_FIJOS_RECORTE[grupo]
        elif cm in AMARILLA_CONTRAMARCAS:
            p = precio_am
        elif c.get('es_bubalino'):
            p = BUBALINO_PRECIO_DEFAULT
        else:
            pricing_col = cm_to_client.get(cm, 'RESTO CLIENTES AMBA')
            p = 0
            if grupo in price_matrix and pricing_col in price_matrix[grupo]:
                p = price_matrix[grupo][pricing_col]
            elif grupo in price_matrix:
                for fb in ['NETOS PEYA', 'RESTO CLIENTES AMBA']:
                    if fb in price_matrix[grupo]:
                        p = price_matrix[grupo][fb]; break
        ingreso_total_calc += c['kg'] * p

    # ── RESULTADO DEL NEGOCIO (arriba, todo con fórmulas) ──
    # Nota: ING_FIRST/ING_LAST se calculan abajo en el detalle,
    # pero necesitamos saber dónde va el total de ingresos para referenciarlo.
    # El detalle empieza en fila 29, así que el total estará en 29 + len(meat_only)
    meat_only = [c for c in cortes if c.get('grupo') != 'GRASA']
    # Calcular posición final del total de ingresos
    # R_DET=30, header=31, cortes_con_precio empiezan en 32
    # Después: sep + subheader + cortes_sin_precio + 1 fila + TOTAL

    def _tiene_precio_lista_pre(c):
        grupo = c['grupo']
        cli = c.get('cliente', 'SIN ASIGNAR')
        cm_str = str(c.get('contramarca', ''))
        if grupo in PRECIOS_FIJOS_RECORTE: return True
        if cm_str in AMARILLA_CONTRAMARCAS: return True
        if c.get('es_bubalino'): return True
        _prec_row_map = {g: 1 for g, _ in GRUPOS}
        _cli_map = {cli_n: 1 for cli_n in cli_list}
        if not (grupo in _prec_row_map and cli in _cli_map):
            return False
        pricing_col = cm_to_client.get(cm_str, 'RESTO CLIENTES AMBA')
        if grupo in price_matrix:
            if pricing_col in price_matrix[grupo] and price_matrix[grupo][pricing_col] > 0:
                return True
            for fb in ['NETOS PEYA', 'RESTO CLIENTES AMBA']:
                if fb in price_matrix[grupo] and price_matrix[grupo][fb] > 0:
                    return True
        return False

    n_con = sum(1 for c in meat_only if _tiene_precio_lista_pre(c))
    n_sin = len(meat_only) - n_con

    # Cortes con precio: filas 32 a 31+n_con
    # Si hay sin precio: +2 (separador + subheader) + n_sin + 1 (total en r+1)
    # Si no hay sin precio: total en 32+n_con+1
    if n_sin > 0:
        _R_TOT_ING = 32 + n_con + 3 + n_sin + 1
    else:
        _R_TOT_ING = 32 + n_con + 1

    ws_res['A3'] = 'RESULTADO DEL NEGOCIO'
    ws_res['A3'].font = Font(name='Arial', bold=True, size=13, color='1F4E79')
    ws_res['A3'].fill = SEC_FILL; ws_res['B3'].fill = SEC_FILL

    R_ING_BRUTO = 4
    ws_res.cell(R_ING_BRUTO, 1, 'Ingresos brutos').font = DATA
    ws_res[f'B{R_ING_BRUTO}'] = f'=E{_R_TOT_ING}'
    ws_res[f'B{R_ING_BRUTO}'].number_format = '$#,##0'; ws_res[f'B{R_ING_BRUTO}'].font = GREEN

    # Impuestos desglosados
    ws_res.cell(5, 1, 'Ganancias (2%)').font = DATA
    ws_res['B5'] = f'=-B{R_ING_BRUTO}*PARAMETROS!B34'
    ws_res['B5'].number_format = '$#,##0'; ws_res['B5'].font = RED

    ws_res.cell(6, 1, 'IIBB (1,5%)').font = DATA
    ws_res['B6'] = f'=-B{R_ING_BRUTO}*PARAMETROS!B35'
    ws_res['B6'].number_format = '$#,##0'; ws_res['B6'].font = RED

    ws_res.cell(7, 1, 'Impuesto Débitos/Créditos (1,2%)').font = DATA
    ws_res['B7'] = f'=-B{R_ING_BRUTO}*PARAMETROS!B36'
    ws_res['B7'].number_format = '$#,##0'; ws_res['B7'].font = RED

    ws_res.cell(8, 1, 'Costo financiero (TNA × días/365)').font = DATA
    ws_res['B8'] = f'=-B{R_ING_BRUTO}*PARAMETROS!B39'
    ws_res['B8'].number_format = '$#,##0'; ws_res['B8'].font = RED

    R_ING_NETO = 9
    ws_res.cell(R_ING_NETO, 1, 'INGRESOS NETOS').font = Font(name='Arial', bold=True, size=11)
    ws_res[f'B{R_ING_NETO}'] = f'=B{R_ING_BRUTO}+B5+B6+B7+B8'
    ws_res[f'B{R_ING_NETO}'].number_format = '$#,##0'
    ws_res[f'B{R_ING_NETO}'].font = Font(name='Arial', bold=True, size=11, color='008000')

    ws_res['A11'] = 'COSTOS'
    ws_res['A11'].font = Font(name='Arial', bold=True, size=12, color='C62828')
    ws_res['A11'].fill = PatternFill('solid', fgColor='FCE4EC')
    ws_res['B11'].fill = PatternFill('solid', fgColor='FCE4EC')

    ws_res.cell(12, 1, 'Compra hacienda (Kg entrada x Precio compra)').font = DATA
    ws_res['B12'] = '=ROMANEO!B6*ROMANEO!B8'
    ws_res['B12'].number_format = '$#,##0'; ws_res['B12'].font = RED

    ws_res.cell(13, 1, 'Mano de obra').font = DATA
    ws_res['B13'] = f'=ROMANEO!F{R_CARNE}*PARAMETROS!B7'
    ws_res['B13'].number_format = '$#,##0'; ws_res['B13'].font = RED

    ws_res.cell(14, 1, 'Insumos').font = DATA
    ws_res['B14'] = f'=ROMANEO!F{R_CARNE}*PARAMETROS!B8'
    ws_res['B14'].number_format = '$#,##0'; ws_res['B14'].font = RED

    ws_res.cell(15, 1, 'Flete').font = DATA
    ws_res['B15'] = f'=ROMANEO!F{R_CARNE}*PARAMETROS!B9'
    ws_res['B15'].number_format = '$#,##0'; ws_res['B15'].font = RED

    ws_res.cell(16, 1, 'SENASA + cuarteo').font = DATA
    ws_res['B16'] = f'=ROMANEO!F{R_CARNE}*PARAMETROS!B10'
    ws_res['B16'].number_format = '$#,##0'; ws_res['B16'].font = RED

    ws_res.cell(18, 1, 'COSTO TOTAL').font = Font(name='Arial', bold=True, size=11, color='C62828')
    ws_res.cell(18, 1).fill = PatternFill('solid', fgColor='FCE4EC')
    ws_res['B18'] = '=SUM(B12:B16)'
    ws_res['B18'].number_format = '$#,##0'
    ws_res['B18'].font = Font(name='Arial', bold=True, size=11, color='C62828')
    ws_res['B18'].fill = PatternFill('solid', fgColor='FCE4EC')

    # CM
    R_CM = 20
    ws_res.cell(R_CM, 1, 'CONTRIBUCIÓN MARGINAL').font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws_res.cell(R_CM, 1).fill = SEC_FILL
    ws_res[f'B{R_CM}'] = f'=B{R_ING_NETO}-B18'
    ws_res[f'B{R_CM}'].number_format = '$#,##0'
    ws_res[f'B{R_CM}'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws_res[f'B{R_CM}'].fill = SEC_FILL

    R_MARGEN = 21
    ws_res.cell(R_MARGEN, 1, 'Margen %').font = Font(name='Arial', bold=True, size=10)
    ws_res[f'B{R_MARGEN}'] = f'=IF(B{R_ING_NETO}=0,0,B{R_CM}/B{R_ING_NETO})'
    ws_res[f'B{R_MARGEN}'].number_format = '0.0%'
    ws_res[f'B{R_MARGEN}'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')

    ws_res.cell(22, 1, 'CM por kg carne').font = DATA
    ws_res['B22'] = f'=IF(ROMANEO!F{R_CARNE}=0,0,B{R_CM}/ROMANEO!F{R_CARNE})'
    ws_res['B22'].number_format = '$#,##0'

    ws_res.cell(23, 1, 'CM por kg entrada').font = DATA
    ws_res['B23'] = f'=IF(ROMANEO!B6=0,0,B{R_CM}/ROMANEO!B6)'
    ws_res['B23'].number_format = '$#,##0'

    ws_res.cell(24, 1, 'Precio venta promedio $/kg carne').font = DATA
    ws_res['B24'] = f'=IF(ROMANEO!F{R_CARNE}=0,0,B{R_ING_BRUTO}/ROMANEO!F{R_CARNE})'
    ws_res['B24'].number_format = '$#,##0'

    # Calificación con fórmula
    ws_res['A26'] = 'CALIFICACIÓN'
    ws_res['A26'].font = Font(name='Arial', bold=True, size=13, color='1F4E79')
    ws_res['A26'].fill = SEC_FILL; ws_res['B26'].fill = SEC_FILL
    ws_res.cell(27, 1, 'RESULTADO').font = Font(name='Arial', bold=True, size=14)
    ws_res['B27'] = (
        f'=IF(B{R_CM}<=0,"PÉRDIDA",'
        f'IF(B{R_MARGEN}>=0.15,"ÓPTIMO",'
        f'IF(B{R_MARGEN}>=0.08,"BUENO",'
        f'IF(B{R_MARGEN}>=0.03,"REGULAR","MALO"))))'
    )
    ws_res['B27'].font = Font(name='Arial', bold=True, size=16)
    ws_res['B27'].alignment = Alignment(horizontal='center')
    rng = 'B27:B27'
    ws_res.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"ÓPTIMO"'], fill=CF_GREEN))
    ws_res.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"BUENO"'], fill=CF_BLUE))
    ws_res.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"REGULAR"'], fill=CF_YELLOW))
    ws_res.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"MALO"'], fill=CF_RED))
    ws_res.conditional_formatting.add(rng, CellIsRule(
        operator='equal', formula=['"PÉRDIDA"'], fill=PatternFill('solid', fgColor='FF0000')))

    # ── DETALLE: Ingresos por línea de corte (abajo) ──
    # Separar cortes con precio en lista vs SIN precio (editables)
    meat_only = [c for c in cortes if c.get('grupo') != 'GRASA']

    # Subcortes que SIEMPRE van a la sección editable (precios manuales),
    # aunque tengan default heredado del parent. Permite override por usuario.
    try:
        from config import SUBCORTE_TO_PARENT
    except Exception:
        SUBCORTE_TO_PARENT = {}

    def _tiene_precio_lista(c):
        grupo = c['grupo']
        cli = c.get('cliente', 'SIN ASIGNAR')
        cm_str = str(c.get('contramarca', ''))
        # Subcortes → SIEMPRE a la sección editable (con precio default)
        if grupo in SUBCORTE_TO_PARENT:
            return False
        if grupo in PRECIOS_FIJOS_RECORTE: return True
        if cm_str in AMARILLA_CONTRAMARCAS: return True
        if c.get('es_bubalino'): return True
        prec_row = grupo_to_prec_row.get(grupo)
        col_idx = cli_to_col_idx.get(cli)
        if not (prec_row and col_idx):
            return False
        pricing_col = cm_to_client.get(cm_str, 'RESTO CLIENTES AMBA')
        if grupo in price_matrix:
            if pricing_col in price_matrix[grupo] and price_matrix[grupo][pricing_col] > 0:
                return True
            for fb in ['NETOS PEYA', 'RESTO CLIENTES AMBA']:
                if fb in price_matrix[grupo] and price_matrix[grupo][fb] > 0:
                    return True
        return False

    def _precio_default_subcorte(c):
        """Precio default para pre-llenar el campo editable de un subcorte.
        Usa NETOS PEYA del subcorte si existe, sino del parent."""
        grupo = c.get('grupo', '')
        if grupo not in SUBCORTE_TO_PARENT:
            return 0
        try:
            from config import PRECIOS_BASE
        except Exception:
            return 0
        # Subcorte propio
        sc_dict = PRECIOS_BASE.get(grupo, {})
        if 'NETOS PEYA' in sc_dict and sc_dict['NETOS PEYA']:
            return sc_dict['NETOS PEYA']
        # Fallback: parent
        parent = SUBCORTE_TO_PARENT[grupo]
        p_dict = PRECIOS_BASE.get(parent, {})
        if 'NETOS PEYA' in p_dict and p_dict['NETOS PEYA']:
            return p_dict['NETOS PEYA']
        if 'RESTO CLIENTES AMBA' in p_dict and p_dict['RESTO CLIENTES AMBA']:
            return p_dict['RESTO CLIENTES AMBA']
        return 0

    cortes_con_precio = [c for c in meat_only if _tiene_precio_lista(c)]
    cortes_sin_precio = [c for c in meat_only if not _tiene_precio_lista(c)]

    # Si cambió el orden, actualizar ING_FIRST (ahora empieza con los que tienen precio)
    R_DET = 30
    ws_res.cell(R_DET, 1, 'DETALLE — INGRESOS POR LÍNEA DE CORTE').font = Font(
        name='Arial', bold=True, size=12, color='2E7D32')
    ws_res.cell(R_DET, 1).fill = PatternFill('solid', fgColor='E2EFDA')
    for cc in range(2, 6):
        ws_res.cell(R_DET, cc).fill = PatternFill('solid', fgColor='E2EFDA')

    hdr_row(ws_res, R_DET + 1, ['Corte', 'Kg', 'Cliente', 'Precio $/kg', 'Ingreso $'],
            PatternFill('solid', fgColor='2E7D32'))

    ING_FIRST = R_DET + 2
    # Primero los cortes con precio
    r = ING_FIRST
    for c in cortes_con_precio:
        ws_res.cell(r, 1, c['corte']).font = DATA
        ws_res.cell(r, 2, c['kg']).font = DATA; ws_res[f'B{r}'].number_format = '#,##0.00'
        cli = c.get('cliente', 'SIN ASIGNAR')
        ws_res.cell(r, 3, cli).font = DATA
        prec_row = grupo_to_prec_row.get(c['grupo'])
        col_idx = cli_to_col_idx.get(cli)
        if prec_row and col_idx:
            col_letter = get_column_letter(col_idx)
            ws_res[f'D{r}'] = f'=PRECIOS!{col_letter}{prec_row}'
        else:
            # Recorte/amarilla/bubalino
            grupo_c = c['grupo']
            cm_c = str(c.get('contramarca', ''))
            if grupo_c in PRECIOS_FIJOS_RECORTE:
                ws_res[f'D{r}'] = PRECIOS_FIJOS_RECORTE[grupo_c]
            elif cm_c in AMARILLA_CONTRAMARCAS:
                ws_res[f'D{r}'] = '=PARAMETROS!$B$17'
            elif c.get('es_bubalino'):
                ws_res[f'D{r}'] = BUBALINO_PRECIO_DEFAULT
        ws_res[f'D{r}'].number_format = '$#,##0'; ws_res[f'D{r}'].font = DATA
        ws_res[f'E{r}'] = f'=B{r}*D{r}'
        ws_res[f'E{r}'].number_format = '$#,##0'; ws_res[f'E{r}'].font = GREEN
        cm_str = str(c.get('contramarca', ''))
        if cm_str in AMARILLA_CONTRAMARCAS:
            for col in range(1, 6):
                ws_res.cell(r, col).fill = AMARILLA_FILL
        r += 1

    R_CON_LAST = r - 1  # última fila con precio de lista

    # ── Sección de cortes SIN precio en lista (editables) ──
    if cortes_sin_precio:
        # Separador
        r += 1
        ws_res.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws_res.cell(r, 1, '⚠️ CORTES SIN PRECIO EN LISTA — Completar precio manual').font = Font(
            name='Arial', bold=True, size=11, color='E65100')
        for cc in range(1, 6):
            ws_res.cell(r, cc).fill = PatternFill('solid', fgColor='FFF3E0')
        r += 1
        # Sub-header
        hdr_row(ws_res, r, ['Corte', 'Kg', 'Cliente', 'Precio manual', 'Ingreso $'],
                PatternFill('solid', fgColor='E65100'))
        r += 1
        R_SIN_FIRST = r
        # Ordenar: subcortes (con default) primero, luego sin precio
        cortes_sub = [c for c in cortes_sin_precio if c.get('grupo') in SUBCORTE_TO_PARENT]
        cortes_otros = [c for c in cortes_sin_precio if c.get('grupo') not in SUBCORTE_TO_PARENT]
        for c in cortes_sub + cortes_otros:
            es_subcorte = c.get('grupo') in SUBCORTE_TO_PARENT
            label = c['corte']
            if es_subcorte:
                label = f"⭐ {label}"  # marcar como subcorte editable
            ws_res.cell(r, 1, label).font = DATA
            ws_res.cell(r, 2, c['kg']).font = DATA; ws_res[f'B{r}'].number_format = '#,##0.00'
            cli = c.get('cliente', 'SIN ASIGNAR')
            ws_res.cell(r, 3, cli).font = DATA
            # Precio editable: subcortes pre-llenados con default; otros con 0
            precio_default = _precio_default_subcorte(c) if es_subcorte else 0
            ws_res.cell(r, 4, precio_default).font = Font(
                name='Arial', bold=True, size=10, color='E65100')
            ws_res[f'D{r}'].number_format = '$#,##0'
            # Color distinto para subcortes (amarillo claro) vs sin precio (naranja claro)
            ws_res[f'D{r}'].fill = PatternFill('solid',
                fgColor='FFF9C4' if es_subcorte else 'FFF3E0')
            ws_res[f'E{r}'] = f'=B{r}*D{r}'
            ws_res[f'E{r}'].number_format = '$#,##0'; ws_res[f'E{r}'].font = GREEN
            r += 1
        R_SIN_LAST = r - 1
    else:
        R_SIN_FIRST = None
        R_SIN_LAST = None

    # Total ingresos (suma de ambas secciones)
    R_TOT_ING = r + 1
    ws_res.cell(R_TOT_ING, 1, 'TOTAL INGRESOS BRUTOS').font = TOTAL_FONT
    for col in range(1, 6):
        ws_res.cell(R_TOT_ING, col).fill = TOTAL_FILL

    # Suma de kg
    if R_SIN_FIRST:
        ws_res[f'B{R_TOT_ING}'] = f'=SUM(B{ING_FIRST}:B{R_CON_LAST})+SUM(B{R_SIN_FIRST}:B{R_SIN_LAST})'
        ws_res[f'E{R_TOT_ING}'] = f'=SUM(E{ING_FIRST}:E{R_CON_LAST})+SUM(E{R_SIN_FIRST}:E{R_SIN_LAST})'
    else:
        ws_res[f'B{R_TOT_ING}'] = f'=SUM(B{ING_FIRST}:B{R_CON_LAST})'
        ws_res[f'E{R_TOT_ING}'] = f'=SUM(E{ING_FIRST}:E{R_CON_LAST})'
    ws_res[f'B{R_TOT_ING}'].number_format = '#,##0.00'; ws_res[f'B{R_TOT_ING}'].font = TOTAL_FONT
    ws_res[f'E{R_TOT_ING}'].number_format = '$#,##0'; ws_res[f'E{R_TOT_ING}'].font = TOTAL_FONT

    R_MARGEN = 18  # para el return
    # Desactivar gridlines en todas las hojas
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    wb.save(output_path)
    return {
        'R_CARNE': R_CARNE, 'R_CM': R_CM, 'R_MARGEN': R_MARGEN,
        'R_ING_NETO': R_ING_NETO, 'R_TOT_ING': R_TOT_ING,
        'AN_FIRST': AN_FIRST, 'AN_LAST': AN_LAST,
        'R_ING_BRUTO': R_ING_BRUTO, 'rend_obj': rend_obj,
        'calidad': calidad,
    }
