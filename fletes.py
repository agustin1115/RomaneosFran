"""
fletes.py — Analiza el costo de fletes (hoja 'Datos facturas'): calcula $/kg por
viaje, prorratea las comisiones de gestión (Silvina, Baltazar, etc.), y marca los
viajes DEFICIENTES (los que superan el umbral de $/kg).

Columnas detalle: MES FLETE, MES FACTURA, FECHA, PROVEEDOR, FACTURA, TIPO,
ID/Remito, N° KGS, MONTO, Costo x KG.
"""
import io
import re

FILE_ID_DEFAULT = '1zjIsUEJC5-Yb7a7NrhCju0YQqouhUNcUzJZQtmlgjKg'

# Proveedor → segmento (editable). GAMYT = PEYA; larga distancia = varios.
SEGMENTO_PROVEEDOR = {
    'GAMYT': 'PEYA',
    'SURFRIGO': 'Larga', 'CRUZ DEL VALLE': 'Larga', 'CRUZ DEL VALLE': 'Larga',
    'TRANSCONT': 'Larga',
}
# Benchmark $/kg esperado por segmento (min, max)
BENCHMARK = {'PEYA': (60, 70), 'Corta': (100, 150), 'Larga': (600, 800)}


def segmento(proveedor):
    p = (proveedor or '').upper().strip()
    for k, v in SEGMENTO_PROVEEDOR.items():
        if k in p:
            return v
    return 'Corta'  # el resto, por defecto, corta distancia


def _num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('$', '').replace(',', '').strip()
    if not s or s == '-':
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_fletes(rows):
    """Devuelve (viajes, comisiones)."""
    viajes, comisiones = [], []
    for row in rows:
        if not row:
            continue
        cells = [('' if c is None else c) for c in row]
        up = [str(c).strip().upper() for c in cells]
        # saltar encabezados
        if 'PROVEEDOR' in up or 'MES FLETE' in up or 'SUM DE' in ' '.join(up):
            continue

        tipo = 'DIRECTO' if 'DIRECTO' in up else ('COMPUESTO' if 'COMPUESTO' in up else '')
        montos = [_num(c) for c in cells if '$' in str(c)]
        monto = montos[-1] if montos else 0.0

        if tipo and len(cells) >= 9:
            proveedor = str(cells[3]).strip()
            factura = str(cells[4]).strip()
            remito = str(cells[6]).strip()
            kg = _num(cells[7])
            mto = _num(cells[8]) if _num(cells[8]) > 0 else monto
            costo_kg = _num(cells[9]) if len(cells) > 9 and _num(cells[9]) > 0 else (mto / kg if kg else 0)
            if kg > 0:
                viajes.append({
                    'fecha': str(cells[2]).strip(), 'proveedor': proveedor, 'factura': factura,
                    'tipo': tipo, 'remito': remito, 'kg': kg, 'monto': mto, 'costo_kg': costo_kg,
                    'n_remitos': len([r for r in re.split(r'[/\s]+', remito) if r.strip().isdigit()]),
                })
        elif monto > 0:
            # comisión de gestión: hay un nombre (texto) y un monto, sin kg
            nombre = ''
            for c in cells[3:6]:
                t = str(c).strip()
                if t and not re.match(r'^[\d\-/.]+$', t) and '$' not in t:
                    nombre = t
                    break
            comisiones.append({'fecha': str(cells[2]).strip(), 'nombre': nombre, 'monto': monto})
    return viajes, comisiones


def analizar(viajes, comisiones):
    """Clasifica cada viaje por segmento (según proveedor) y lo compara contra su
    benchmark. Deficiente = $/kg por encima del máximo esperado de su segmento."""
    for v in viajes:
        v['segmento'] = segmento(v['proveedor'])
        v['benchmark'] = BENCHMARK.get(v['segmento'], (0, 0))[1]
        v['deficiente'] = v['costo_kg'] > v['benchmark']
        v['exceso_kg'] = max(0.0, v['costo_kg'] - v['benchmark'])

    kg_total = sum(v['kg'] for v in viajes)
    monto_viajes = sum(v['monto'] for v in viajes)
    monto_comisiones = sum(c['monto'] for c in comisiones)
    comision_por_kg = monto_comisiones / kg_total if kg_total else 0
    costo_kg_prom = (monto_viajes + monto_comisiones) / kg_total if kg_total else 0

    # Desglose por segmento
    por_segmento = {}
    for seg in ('PEYA', 'Corta', 'Larga'):
        vs = [v for v in viajes if v['segmento'] == seg]
        kg = sum(v['kg'] for v in vs)
        mto = sum(v['monto'] for v in vs)
        real = mto / kg if kg else 0
        bmin, bmax = BENCHMARK[seg]
        por_segmento[seg] = {
            'kg': kg, 'monto': mto, 'real_kg': real, 'benchmark': f"${bmin}-{bmax}",
            'bmax': bmax, 'gap_kg': real - bmax, 'n_viajes': len(vs),
            'sobrecosto': sum(v['exceso_kg'] * v['kg'] for v in vs),
            'pct_kg': kg / kg_total * 100 if kg_total else 0,
        }

    deficientes = sorted([v for v in viajes if v['deficiente']],
                         key=lambda x: -(x['exceso_kg'] * x['kg']))
    sobrecosto_total = sum(v['exceso_kg'] * v['kg'] for v in viajes)

    return {
        'kg_total': kg_total, 'monto_viajes': monto_viajes, 'monto_comisiones': monto_comisiones,
        'comision_por_kg': comision_por_kg, 'costo_kg_prom': costo_kg_prom,
        'por_segmento': por_segmento, 'n_viajes': len(viajes), 'n_deficientes': len(deficientes),
        'deficientes': deficientes, 'sobrecosto_total': sobrecosto_total,
    }


def cargar_fletes_drive(credentials_path, file_id=FILE_ID_DEFAULT, hoja='Datos facturas'):
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    import openpyxl
    scopes = ['https://www.googleapis.com/auth/drive.readonly']
    creds = Credentials.from_service_account_file(credentials_path, scopes=scopes)
    service = build('drive', 'v3', credentials=creds)
    content = service.files().export(
        fileId=file_id,
        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ).execute()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[hoja] if hoja in wb.sheetnames else wb.worksheets[-1]
    rows = list(ws.iter_rows(values_only=True))
    viajes, comisiones = parse_fletes(rows)
    return viajes, comisiones, analizar(viajes, comisiones)
